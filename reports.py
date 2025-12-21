"""
Модуль генерации отчетов
"""
from typing import List, Dict, Tuple, Optional
from collections import defaultdict
import csv
from io import StringIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


class ReportGenerator:
    """Генератор отчетов"""
    
    @staticmethod
    def calculate_report(operations: List[Dict], sb_name_merges: Optional[Dict[str, str]] = None, 
                        stylist_expenses: Optional[List[Dict]] = None) -> Tuple[List[Dict], Dict, Dict, bool]:
        """
        Расчет отчета по операциям
        sb_name_merges: словарь для объединения имен СБ {старое_имя: новое_имя}
        stylist_expenses: список расходов на стилистов [{'code': 'Д14', 'name': 'Бритни', 'amount': 2000}, ...]
        Возвращает: (строки_отчета, итоги_по_строкам, итоги_пересчет, проверка_ок)
        """
        # Группируем по сотрудникам
        # Для СБ группируем по (код, имя), для остальных по коду
        employee_data = defaultdict(lambda: {
            'names': set(),
            'nal': 0.0,
            'beznal': 0.0,
            'stylist': 0.0
        })
        
        # Пересчет для проверки
        total_nal_raw = 0.0
        total_beznal_raw = 0.0
        
        for op in operations:
            code = op['code']
            name = op['name']
            channel = op['channel']
            amount = op['amount']
            
            # Применяем объединение имен СБ (только для отчета)
            if sb_name_merges and code == 'СБ' and name in sb_name_merges:
                name = sb_name_merges[name]
            
            # ДЛЯ СБ группируем по комбинации (код + имя), чтобы разные СБ не объединялись
            if code == 'СБ':
                group_key = f"СБ_{name}" if name else "СБ"
            else:
                group_key = code
            
            employee_data[group_key]['names'].add(name)
            
            if channel == 'нал':
                employee_data[group_key]['nal'] += amount
                total_nal_raw += amount
            elif channel == 'безнал':
                employee_data[group_key]['beznal'] += amount
                total_beznal_raw += amount
        
        # Добавляем расходы на стилистов
        if stylist_expenses:
            print(f"🔍 DEBUG: Всего расходов на стилистов: {len(stylist_expenses)}")
            for expense in stylist_expenses:
                exp_code = expense['code']
                exp_name = expense['name']
                exp_amount = expense['amount']
                
                print(f"🔍 DEBUG: Обрабатываю стилиста - код: {exp_code}, имя: {exp_name}, сумма: {exp_amount}")
                
                # Ищем соответствующего сотрудника
                # Для СБ не ищем (СБ не может иметь расходов на стилистов)
                if exp_code == 'СБ':
                    print(f"  ⚠️ Пропускаем СБ")
                    continue
                
                # Для обычных сотрудников ищем по коду
                group_key = exp_code
                
                print(f"  🔑 group_key: {group_key}")
                print(f"  📋 Ключи в employee_data: {list(employee_data.keys())}")
                
                # Проверяем, есть ли такой сотрудник в данных
                if group_key in employee_data:
                    print(f"  ✅ Сотрудник НАЙДЕН в данных")
                    # Добавляем расход по коду (имя не проверяем)
                    print(f"  ✅ Добавляю {exp_amount}₽ к коду {group_key}")
                    employee_data[group_key]['stylist'] += exp_amount
                else:
                    print(f"  ❌ Сотрудник НЕ НАЙДЕН в данных (код {group_key} отсутствует в operations)")
        
        # Формируем строки отчета
        report_rows = []
        total_nal = 0.0
        total_beznal = 0.0
        total_minus10 = 0.0
        total_stylist = 0.0
        total_itog = 0.0
        
        for group_key in sorted(employee_data.keys()):
            data = employee_data[group_key]
            
            # Определяем реальный код (убираем префикс "СБ_" если есть)
            if group_key.startswith('СБ_'):
                code = 'СБ'
            else:
                code = group_key
            
            # Имя (если разные - берем первое и помечаем)
            names_list = list(data['names'])
            name = names_list[0]
            name_comment = " (⚠️ разные имена)" if len(names_list) > 1 else ""
            
            nal = round(data['nal'], 2)
            beznal = round(data['beznal'], 2)
            minus10 = round(beznal * 0.10, 2)
            stylist = round(data['stylist'], 2)
            itog = round(nal + (beznal - minus10) - stylist, 2)  # НОВАЯ ФОРМУЛА
            
            report_rows.append({
                'name': name + name_comment,
                'code': code,
                'nal': nal,
                'beznal': beznal,
                'minus10': minus10,
                'stylist': stylist,  # НОВОЕ ПОЛЕ
                'itog': itog
            })
            
            total_nal += nal
            total_beznal += beznal
            total_minus10 += minus10
            total_stylist += stylist
            total_itog += itog
        
        # Итоги по строкам
        totals_by_rows = {
            'nal': round(total_nal, 2),
            'beznal': round(total_beznal, 2),
            'minus10': round(total_minus10, 2),
            'stylist': round(total_stylist, 2),  # НОВОЕ ПОЛЕ
            'itog': round(total_itog, 2)
        }
        
        # Пересчет для проверки
        recalc_minus10 = round(total_beznal_raw * 0.10, 2)
        recalc_itog = round(total_nal_raw + (total_beznal_raw - recalc_minus10), 2)
        
        totals_recalc = {
            'nal': round(total_nal_raw, 2),
            'beznal': round(total_beznal_raw, 2),
            'minus10': recalc_minus10,
            'itog': recalc_itog
        }
        
        # Проверка совпадения
        check_ok = (
            totals_by_rows['nal'] == totals_recalc['nal'] and
            totals_by_rows['beznal'] == totals_recalc['beznal'] and
            totals_by_rows['minus10'] == totals_recalc['minus10'] and
            totals_by_rows['itog'] == totals_recalc['itog']
        )
        
        return report_rows, totals_by_rows, totals_recalc, check_ok
    
    @staticmethod
    def format_report_text(report_rows: List[Dict], totals: Dict, 
                          check_ok: bool, totals_recalc: Dict,
                          club: str, period: str) -> str:
        """
        Форматирование отчета для вывода в Telegram
        """
        if not report_rows:
            return f"📊 Отчет по клубу {club} за {period}\n\nДанных нет."
        
        result = []
        result.append(f"📊 ОТЧЕТ")
        result.append(f"Клуб: {club}")
        result.append(f"Период: {period}")
        result.append("")
        
        # Заголовок таблицы
        result.append("```")
        result.append(f"{'Имя':<20} {'Код':<6} {'Нал':>10} {'Безнал':>10} {'10%':>10} {'Итог':>12}")
        result.append("-" * 80)
        
        # Строки отчета
        for row in report_rows:
            name_display = row['name'][:20]  # Обрезаем длинные имена
            result.append(
                f"{name_display:<20} {row['code']:<6} "
                f"{row['nal']:>10.2f} {row['beznal']:>10.2f} "
                f"{row['minus10']:>10.2f} {row['itog']:>12.2f}"
            )
        
        # Итоги
        result.append("-" * 80)
        result.append(
            f"{'ИТОГО':<20} {'':<6} "
            f"{totals['nal']:>10.2f} {totals['beznal']:>10.2f} "
            f"{totals['minus10']:>10.2f} {totals['itog']:>12.2f}"
        )
        result.append("```")
        result.append("")
        
        # Проверка
        if check_ok:
            result.append("✅ Сверка столбцов: совпадает")
        else:
            result.append("❗ Сверка столбцов: РАСХОЖДЕНИЕ")
            result.append("Пересчёт:")
            result.append(f"  Нал: {totals['nal']} vs {totals_recalc['nal']}")
            result.append(f"  Безнал: {totals['beznal']} vs {totals_recalc['beznal']}")
            result.append(f"  10%: {totals['minus10']} vs {totals_recalc['minus10']}")
            result.append(f"  Итог: {totals['itog']} vs {totals_recalc['itog']}")
        
        return '\n'.join(result)
    
    @staticmethod
    def generate_csv(report_rows: List[Dict], totals: Dict) -> str:
        """
        Генерация CSV
        """
        output = StringIO()
        writer = csv.writer(output)
        
        # Заголовок
        writer.writerow(['Имя', 'Код', 'Нал', 'Безнал', '10% от безнала', 'Итог (нал + безнал − 10%)'])
        
        # Данные
        for row in report_rows:
            writer.writerow([
                row['name'],
                row['code'],
                row['nal'],
                row['beznal'],
                row['minus10'],
                row['itog']
            ])
        
        # Итоги
        writer.writerow([
            'ИТОГО',
            '',
            totals['nal'],
            totals['beznal'],
            totals['minus10'],
            totals['itog']
        ])
        
        return output.getvalue()
    
    @staticmethod
    def generate_xlsx(report_rows: List[Dict], totals: Dict, 
                     club: str, period: str, filename: str, db=None) -> str:
        """
        Генерация XLSX файла
        db: экземпляр Database для проверки статуса самозанятости
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет"
        
        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        ws['A1'] = f"Отчет по клубу {club}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Период: {period}"
        ws['A2'].font = Font(size=11)
        
        # Шапка таблицы
        headers = [
            'Имя', 'Код', 'Нал', 'Безнал', '10% от безнала', 'Стилисты',
            'ИТОГО', 'Самозанятость', 'К выплате (самозанятый)'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Данные
        row_num = 5
        for row_data in report_rows:
            # Имя (текст, выравнивание влево)
            cell = ws.cell(row=row_num, column=1, value=row_data['name'])
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
            
            # Код (текст, выравнивание влево)
            cell = ws.cell(row=row_num, column=2, value=row_data['code'])
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
            
            # Числовые колонки (выравнивание вправо)
            for col, key in enumerate(['nal', 'beznal', 'minus10', 'stylist', 'itog'], 3):
                cell = ws.cell(row=row_num, column=col, value=row_data[key])
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = border
            
            # Проверяем статус самозанятости
            if db:
                normalized_code = row_data['code'].upper().strip()
                is_self_employed = db.is_self_employed(normalized_code)
                if is_self_employed:
                    cell = ws.cell(row=row_num, column=8, value='✓')
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                    
                    payout = round(row_data['itog'] / 0.94, 2)
                    cell = ws.cell(row=row_num, column=9, value=payout)
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = border
                else:
                    cell = ws.cell(row=row_num, column=8, value='')
                    cell.border = border
                    cell = ws.cell(row=row_num, column=9, value='')
                    cell.border = border
            else:
                cell = ws.cell(row=row_num, column=8, value='')
                cell.border = border
                cell = ws.cell(row=row_num, column=9, value='')
                cell.border = border
            
            row_num += 1
        
        # Итоги
        cell = ws.cell(row=row_num, column=1, value='ИТОГО')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        
        cell = ws.cell(row=row_num, column=2, value='')
        cell.border = border
        
        for col, key in enumerate(['nal', 'beznal', 'minus10', 'stylist', 'itog'], 3):
            cell = ws.cell(row=row_num, column=col, value=totals[key])
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border
        
        # Пустые ячейки в строке итогов
        for col in [8, 9]:
            cell = ws.cell(row=row_num, column=col, value='')
            cell.border = border
        
        # Автоматическая подгонка ширины столбцов
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем
        wb.save(filename)
        return filename
    
    @staticmethod
    def generate_merged_xlsx(report_moskvich: Tuple[List[Dict], Dict],
                            report_anora: Tuple[List[Dict], Dict],
                            report_merged: Tuple[List[Dict], Dict],
                            period: str, filename: str, db=None) -> str:
        """
        Генерация сводного XLSX файла с тремя листами:
        - Лист 1: Москвич
        - Лист 2: Анора
        - Лист 3: Сводный
        
        report_moskvich: (report_rows, totals) для Москвича
        report_anora: (report_rows, totals) для Аноры
        report_merged: (report_rows, totals) для сводного
        """
        wb = Workbook()
        
        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def fill_sheet(ws, club_name, report_rows, totals):
            """Заполнить лист данными"""
            # Заголовок
            ws['A1'] = f"Отчет по клубу {club_name}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f"Период: {period}"
            ws['A2'].font = Font(size=11)
            
            # Шапка таблицы
            headers = [
                'Имя', 'Код', 'Нал', 'Безнал', '10% от безнала', 'Стилисты',
                'Итог (нал + безнал − 10% − стилисты)', 'Самозанятость', 'К выплате (самозанятый)'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Данные
            row_num = 5
            for row_data in report_rows:
                # Имя
                cell = ws.cell(row=row_num, column=1, value=row_data['name'])
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
                
                # Код
                cell = ws.cell(row=row_num, column=2, value=row_data['code'])
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = border
                
                # Числовые колонки
                for col, key in enumerate(['nal', 'beznal', 'minus10', 'stylist', 'itog'], 3):
                    cell = ws.cell(row=row_num, column=col, value=row_data[key])
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    cell.border = border
                
                # Самозанятость
                if db:
                    normalized_code = row_data['code'].upper().strip()
                    is_self_employed = db.is_self_employed(normalized_code)
                    if is_self_employed:
                        cell = ws.cell(row=row_num, column=8, value='✓')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                        
                        payout = round(row_data['itog'] / 0.94, 2)
                        cell = ws.cell(row=row_num, column=9, value=payout)
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                        cell.border = border
                    else:
                        cell = ws.cell(row=row_num, column=8, value='')
                        cell.border = border
                        cell = ws.cell(row=row_num, column=9, value='')
                        cell.border = border
                else:
                    cell = ws.cell(row=row_num, column=8, value='')
                    cell.border = border
                    cell = ws.cell(row=row_num, column=9, value='')
                    cell.border = border
                
                row_num += 1
            
            # Итоги
            cell = ws.cell(row=row_num, column=1, value='ИТОГО')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
            
            cell = ws.cell(row=row_num, column=2, value='')
            cell.border = border
            
            for col, key in enumerate(['nal', 'beznal', 'minus10', 'stylist', 'itog'], 3):
                cell = ws.cell(row=row_num, column=col, value=totals[key])
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = border
            
            for col in [8, 9]:
                cell = ws.cell(row=row_num, column=col, value='')
                cell.border = border
            
            # Автоматическая подгонка ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Лист 1: Москвич
        ws1 = wb.active
        ws1.title = "Москвич"
        fill_sheet(ws1, "Москвич", report_moskvich[0], report_moskvich[1])
        
        # Лист 2: Анора
        ws2 = wb.create_sheet(title="Анора")
        fill_sheet(ws2, "Анора", report_anora[0], report_anora[1])
        
        # Лист 3: Сводный
        ws3 = wb.create_sheet(title="Сводный")
        fill_sheet(ws3, "СВОДНЫЙ (Москвич + Анора)", report_merged[0], report_merged[1])
        
        # Сохраняем
        wb.save(filename)
        return filename

