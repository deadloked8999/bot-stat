"""
Главный модуль Telegram бота для учета статистики
"""
import os
import re
import uuid
import tempfile
from datetime import datetime
from typing import Dict, Optional, Tuple, List
from openpyxl import Workbook
from difflib import SequenceMatcher

from telegram import Update, ReplyKeyboardMarkup, InlineKeyboardMarkup, InlineKeyboardButton, ReplyKeyboardRemove
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    CallbackQueryHandler,
    filters
)

import config
from database import Database
from parser import DataParser
from reports import ReportGenerator
from excel_parser import ExcelProcessor
from utils import (
    get_current_date,
    parse_date,
    parse_short_date,
    parse_date_range,
    get_week_range,
    parse_period,
    normalize_command,
    extract_club_from_text,
    format_operations_list
)


# Состояния пользователя
USER_STATES = {}

# Пин-код для удаления всех данных
RESET_PIN_CODE = "6002147"


class UserState:
    """Класс для хранения состояния пользователя"""
    
    def __init__(self):
        self.club: Optional[str] = None
        self.mode: Optional[str] = None
        self.temp_nal_data: list = []
        self.temp_beznal_data: list = []
        self.current_date: str = get_current_date()
        self.limited_access: bool = False  # Ограниченный доступ (только выплаты)
        
        # Для команды отчет
        self.report_club: Optional[str] = None
        self.pending_report_period: Optional[tuple] = None  # Для хранения периода при обработке "оба"
        
        # Для команды исправить
        self.edit_code: Optional[str] = None
        self.edit_date: Optional[str] = None
        self.edit_current_data: Optional[dict] = None
        
        # Для команды удалить
        self.delete_code: Optional[str] = None
        self.delete_date: Optional[str] = None
        self.delete_records: Optional[dict] = None
        self.delete_mass_club: Optional[str] = None
        self.delete_mass_date_from: Optional[str] = None
        self.delete_mass_date_to: Optional[str] = None
        self.delete_mass_preview: Optional[dict] = None
        
        # Для команды экспорт
        self.export_club: Optional[str] = None
        
        # Для команды список
        self.list_club: Optional[str] = None
        
        # Для сводного отчета
        self.merge_candidates: Optional[list] = None
        self.merge_period: Optional[tuple] = None
        
        # Для проверки дубликатов в отчёте
        self.duplicate_check_data: Optional[dict] = None
        
        # Для объединения СБ с похожими именами
        self.sb_merge_data: Optional[dict] = None
        
        # Сохранённые объединения СБ для каждого клуба (для сводного отчёта)
        self.sb_merges_moskvich: Optional[dict] = None
        self.sb_merges_anora: Optional[dict] = None
        
        # Для объединения сотрудников
        self.employees_list: Optional[list] = None
        self.employees_club: Optional[str] = None
        self.merge_employee_indices: Optional[list] = None
        
        # Для редактирования сотрудников
        self.edit_employees_list: Optional[list] = None
        self.edit_employees_club: Optional[str] = None
        self.edit_employee_selected: Optional[dict] = None
        
        # Для добавления сотрудника
        self.add_employee_club: Optional[str] = None
        
        # Для режима сотрудника
        self.employee_mode: bool = False
        self.employee_code: Optional[str] = None
        self.employee_club: Optional[str] = None
        self.employee_name: Optional[str] = None
        
        # Для предпросмотра данных
        self.preview_date: Optional[str] = None
        self.preview_duplicates: Optional[list] = None
        self.edit_line_number: Optional[int] = None
        
        # Для загрузки файла Excel
        self.upload_file_club: Optional[str] = None
        self.upload_file_date: Optional[str] = None
        self.upload_file_data: Optional[dict] = None
        
        # Для загрузки листа выплат (ЗП)
        self.payments_upload_club: Optional[str] = None
        self.payments_upload_date: Optional[str] = None
        self.payments_upload_data: Optional[list] = None
        self.payments_preview_data: Optional[list] = None
        self.payments_name_changes: Optional[list] = None
        
        # Для расходов на стилистов
        self.stylist_club: Optional[str] = None
        self.stylist_period_from: Optional[str] = None
        self.stylist_period_to: Optional[str] = None
        self.stylist_expenses: Optional[list] = None
        self.stylist_errors: Optional[list] = None
        self.stylist_edit_index: Optional[int] = None  # Индекс редактируемой записи
        self.stylist_clarification_queue: Optional[list] = None  # Очередь записей требующих уточнения
        self.stylist_clarification_index: Optional[int] = None  # Текущий индекс в очереди
        
        # Для просмотра/управления стилистами
        self.stylist_view_club: Optional[str] = None
        self.stylist_view_from: Optional[str] = None
        self.stylist_view_to: Optional[str] = None
        
        # ID сообщений бота для удаления
        self.bot_messages: list = []
    
    def reset_input(self):
        """Сброс блочного ввода"""
        self.mode = None
        self.temp_nal_data = []
        self.temp_beznal_data = []
        self.preview_date = None
        self.preview_duplicates = None
        self.edit_line_number = None
        self.delete_mass_club = None
        self.delete_mass_date_from = None
        self.delete_mass_date_to = None
        self.delete_mass_preview = None
        self.edit_employees_list = None
        self.edit_employees_club = None
        self.edit_employee_selected = None
        self.employee_mode = False
        self.employee_code = None
        self.employee_club = None
        self.employee_name = None
    
    def has_data(self) -> bool:
        """Проверка наличия данных"""
        return len(self.temp_nal_data) > 0 or len(self.temp_beznal_data) > 0


def get_user_state(user_id: int) -> UserState:
    """Получить состояние пользователя"""
    if user_id not in USER_STATES:
        USER_STATES[user_id] = UserState()
    return USER_STATES[user_id]


async def send_and_save(update: Update, state: UserState, text: str, **kwargs):
    """Отправить сообщение и сохранить его ID для возможного удаления"""
    msg = await update.message.reply_text(text, **kwargs)
    state.bot_messages.append(msg.message_id)
    # Ограничиваем список (храним только последние 100 сообщений)
    if len(state.bot_messages) > 100:
        state.bot_messages = state.bot_messages[-100:]
    return msg


def get_main_keyboard():
    """Главная клавиатура с основными командами"""
    keyboard = [
        ['📥 НАЛ', '📥 БЕЗНАЛ'],
        ['📎 ЗАГРУЗИТЬ ФАЙЛ', '💰 ЗАГРУЗИТЬ ЗП'],
        ['✅ ГОТОВО', '❌ ОТМЕНА'],
        ['📊 ОТЧЁТ', '💰 ВЫПЛАТЫ', '💵 ЗП'],
        ['📋 СПИСОК', '📤 ЭКСПОРТ'],
        ['✏️ ИСПРАВИТЬ', '🗑️ УДАЛИТЬ'],
        ['📜 ЖУРНАЛ', '👔 САМОЗАНЯТЫЕ'],
        ['👥 СОТРУДНИКИ', '💄 СТИЛИСТЫ'],
        ['❓ ПОМОЩЬ', '🚪 ЗАВЕРШИТЬ']
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def get_club_keyboard():
    """Клавиатура для выбора клуба (Inline кнопки)"""
    keyboard = [
        [InlineKeyboardButton("🏢 Москвич", callback_data='club_moskvich')],
        [InlineKeyboardButton("🏢 Анора", callback_data='club_anora')]
    ]
    return InlineKeyboardMarkup(keyboard)


def get_club_choice_keyboard():
    """Постоянная клавиатура для выбора клуба (Reply кнопки)"""
    keyboard = [
        ['🏢 СТАРТ МОСКВИЧ'],
        ['🏢 СТАРТ АНОРА']
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def get_employee_menu_keyboard():
    """Клавиатура для сотрудника (ограниченный доступ)"""
    keyboard = [
        ['💰 Моя ЗП'],  # Последняя начисленная ЗП
        ['💵 История выплат'],  # История последних выплат
        ['❌ Выход']
    ]
    return ReplyKeyboardMarkup(keyboard, resize_keyboard=True)


def get_club_report_keyboard():
    """Клавиатура для выбора клуба в отчёте"""
    keyboard = [
        [InlineKeyboardButton("🏢 Москвич", callback_data='report_club_moskvich')],
        [InlineKeyboardButton("🏢 Анора", callback_data='report_club_anora')],
        [InlineKeyboardButton("🏢🏢 ОБА", callback_data='report_club_both')]
    ]
    return InlineKeyboardMarkup(keyboard)


def get_club_employees_keyboard():
    """Клавиатура выбора клуба для списка сотрудников"""
    keyboard = [
        [InlineKeyboardButton("🏢 Москвич", callback_data='employees_club_moskvich')],
        [InlineKeyboardButton("🏢 Анора", callback_data='employees_club_anora')]
    ]
    return InlineKeyboardMarkup(keyboard)


def get_employees_menu_keyboard():
    """Меню управления сотрудниками"""
    keyboard = [
        [InlineKeyboardButton("🔗 Объединить сотрудников", callback_data='employees_merge')],
        [InlineKeyboardButton("✏️ Редактировать сотрудника", callback_data='employees_edit')],
        [InlineKeyboardButton("➕ Добавить сотрудника", callback_data='employees_add')],
        [InlineKeyboardButton("❌ Назад", callback_data='employees_cancel')]
    ]
    return InlineKeyboardMarkup(keyboard)


def get_delete_keyboard():
    """Клавиатура для выбора что удалить"""
    keyboard = [
        [InlineKeyboardButton("📗 НАЛ", callback_data='delete_nal')],
        [InlineKeyboardButton("📘 БЕЗНАЛ", callback_data='delete_beznal')],
        [InlineKeyboardButton("🗑️ ОБЕ", callback_data='delete_both')]
    ]
    return InlineKeyboardMarkup(keyboard)


def get_delete_mode_keyboard():
    """Клавиатура выбора режима удаления"""
    keyboard = [
        [InlineKeyboardButton("🧍 Удалить сотрудника", callback_data='delete_mode_employee')],
        [InlineKeyboardButton("🗑️ Удалить все", callback_data='delete_mode_mass')]
    ]
    return InlineKeyboardMarkup(keyboard)


def get_delete_mass_confirm_keyboard():
    """Клавиатура подтверждения массового удаления"""
    keyboard = [
        [
            InlineKeyboardButton("✅ Да, удалить", callback_data='delete_mass_confirm_yes'),
            InlineKeyboardButton("❌ Нет", callback_data='delete_mass_confirm_no')
        ]
    ]
    return InlineKeyboardMarkup(keyboard)


def get_self_employed_action_keyboard():
    """Клавиатура для управления самозанятыми"""
    keyboard = [
        [
            InlineKeyboardButton("➕ Добавить код", callback_data='self_employed_add'),
            InlineKeyboardButton("➖ Удалить код", callback_data='self_employed_remove')
        ],
        [InlineKeyboardButton("❌ Закрыть", callback_data='self_employed_close')]
    ]
    return InlineKeyboardMarkup(keyboard)


def get_merge_confirmation_keyboard():
    """Клавиатура для подтверждения объединения совпадений"""
    keyboard = [
        [
            InlineKeyboardButton("✅ Объединить все", callback_data='merge_all'),
            InlineKeyboardButton("❌ Не объединять", callback_data='merge_none')
        ],
        [InlineKeyboardButton("📄 Показать список", callback_data='merge_show_list')]
    ]
    return InlineKeyboardMarkup(keyboard)


def make_processed_key(code: str, name: Optional[str]) -> Tuple[str, str]:
    """Нормализованный ключ для отслеживания уже обработанных записей"""
    return code, (name or "").strip()


# Инициализация базы данных
db = Database()


async def start_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка команды /start и старт"""
    user_id = update.effective_user.id
    state = get_user_state(user_id)
    
    # Проверка: админ или сотрудник?
    if not db.is_admin(user_id):
        employee = db.get_employee_by_telegram_id(user_id)
        if employee and employee['is_active']:
            # Сотрудник - показываем персональное меню
            state.employee_mode = True
            state.employee_code = employee['code']
            state.employee_club = employee['club']
            state.employee_name = employee['full_name'] or employee['code']
            
            await update.message.reply_text(
                f"👋 Привет, {state.employee_name}!\n\n"
                f"🏢 Клуб: {state.employee_club}\n"
                f"💼 Код: {state.employee_code}\n\n"
                f"Используйте кнопки меню:",
                reply_markup=get_employee_menu_keyboard()
            )
            return
        else:
            # Не админ и не сотрудник - игнорируем
            return
    
    # Блокируем /start для ограниченного доступа
    if state.limited_access:
        await update.message.reply_text(
            "❌ Доступ запрещён\n\n"
            "У вас ограниченный доступ.\n"
            "Доступна только функция 'Выплаты'."
        )
        return
    
    # Получаем текст команды
    if update.message:
        text = update.message.text.lower()
    else:
        text = ""
    
    # Если команда /start без параметров - показываем выбор клуба
    if text.strip() == '/start':
        await update.message.reply_text(
            "Выберите клуб:",
            reply_markup=get_club_choice_keyboard()
        )
        return
    
    # Определяем клуб
    club = None
    if 'москвич' in text:
        club = 'Москвич'
    elif 'анора' in text or 'anora' in text:
        club = 'Анора'
    
    if not club:
        await update.message.reply_text(
            "Выберите клуб, нажав на кнопку ниже:",
            reply_markup=get_club_choice_keyboard()
        )
        return
    
    # Устанавливаем клуб и дату
    state.club = club
    state.current_date = get_current_date()
    state.reset_input()
    # Сбрасываем ВСЕ режимы и данные объединения
    state.mode = None
    state.duplicate_check_data = None
    state.sb_merge_data = None
    state.report_club = None
    
    await update.message.reply_text(
        f"✅ Выбран клуб: {club}\n"
        f"📅 Текущая дата: {state.current_date}\n\n"
        f"🎯 ЧТО ДАЛЬШЕ?\n\n"
        f"📥 Для ввода данных:\n"
        f"   • Нажмите НАЛ или БЕЗНАЛ\n"
        f"   • Вставьте список данных\n"
        f"   • Нажмите ГОТОВО\n\n"
        f"📊 Для просмотра отчётов:\n"
        f"   • Нажмите ОТЧЁТ, ВЫПЛАТЫ или СПИСОК\n\n"
        f"❓ Полная справка: нажмите ПОМОЩЬ\n\n"
        f"Используйте кнопки меню ⬇️",
        reply_markup=get_main_keyboard()
    )


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка текстовых сообщений"""
    user_id = update.effective_user.id
    state = get_user_state(user_id)
    text = update.message.text.strip()
    text_lower = normalize_command(text)
    
    print(f"DEBUG: Получена команда: '{text}', mode={state.mode}, limited_access={state.limited_access}")
    
    # Проверка авторизации
    if not db.is_admin(user_id) and not state.employee_mode:
        # Специальный код для limited_access
        if text == "0001":
            state.limited_access = True
            
            keyboard = [[InlineKeyboardButton("❌ Выход", callback_data="quick_exit")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            await update.message.reply_text(
                "🔐 Быстрый доступ к выплатам\n\n"
                "📋 Инструкция:\n"
                "1️⃣ Введите код сотрудника и период\n"
                "2️⃣ Получите Excel файл с выплатами\n\n"
                "📝 Примеры:\n"
                "• Д7 12,12\n"
                "• Д7 10,06-11,08\n\n"
                "💡 Введите данные:",
                reply_markup=reply_markup
            )
            state.mode = 'awaiting_payments_input'
            return
        else:
            # Проверяем: может это сотрудник?
            employee = db.get_employee_by_telegram_id(user_id)
            if employee and employee['is_active']:
                state.employee_mode = True
                state.employee_code = employee['code']
                state.employee_club = employee['club']
                state.employee_name = employee['full_name'] or employee['code']
                
                await update.message.reply_text(
                    f"👋 Привет, {state.employee_name}!\n\n"
                    f"🏢 Клуб: {state.employee_club}\n"
                    f"💼 Код: {state.employee_code}\n\n"
                    f"Используйте кнопки меню:",
                    reply_markup=get_employee_menu_keyboard()
                )
                return
            else:
                # Не админ, не сотрудник, не спец код - игнорируем
                return
    
    # УНИВЕРСАЛЬНАЯ КНОПКА ОТМЕНА - работает на ЛЮБОМ этапе!
    # Проверяем ПЕРЕД всеми режимами
    if text_lower == 'отмена' or text_lower == '❌ отмена':
        # Список режимов где ОТМЕНА должна работать
        cancelable_modes = [
            'awaiting_preview_date', 'awaiting_preview_action', 'awaiting_edit_line_number', 'awaiting_edit_line_data',
            'awaiting_edit_params', 'awaiting_edit_data', 'awaiting_delete_choice',
            'awaiting_report_club', 'awaiting_report_period', 'awaiting_duplicate_confirm', 'awaiting_sb_merge_confirm',
            'awaiting_export_club', 'awaiting_export_period',
            'awaiting_merge_confirm', 'awaiting_list_club', 'awaiting_list_date', 'awaiting_payments_input', 'awaiting_salary_input',
            'awaiting_delete_mass_club', 'awaiting_delete_mass_period', 'awaiting_delete_mass_confirm',
            'awaiting_delete_employee_input',
            'awaiting_upload_club', 'awaiting_upload_date', 'awaiting_upload_file', 'awaiting_upload_confirm',
            'awaiting_payments_upload_club', 'awaiting_payments_upload_date', 'awaiting_payments_upload_file',
            'awaiting_stylist_period', 'awaiting_stylist_data', 'awaiting_stylist_confirm', 
            'awaiting_stylist_edit_number', 'awaiting_stylist_edit_data', 'awaiting_stylist_clarification',
            'awaiting_employee_edit_select', 'awaiting_emp_code', 'awaiting_add_employee',
            'awaiting_emp_name', 'awaiting_emp_phone', 'awaiting_emp_tg', 'awaiting_emp_birth',
            'employee_awaiting_date', 'employee_awaiting_period',
            'нал', 'безнал'
        ]
        
        if state.mode in cancelable_modes or state.has_data():
            # Если ограниченный доступ - выходим полностью
            if state.limited_access:
                state.__init__()
                state.limited_access = False
                await update.message.reply_text(
                    "❌ Сессия завершена\n\n"
                    "Для начала работы введите /start"
                )
                return
            
            # Полная очистка (но клуб остаётся!)
            saved_club = state.club  # Сохраняем клуб
            state.reset_input()
            state.mode = None
            state.duplicate_check_data = None
            state.sb_merge_data = None
            state.report_club = None
            state.export_club = None
            state.list_club = None
            state.edit_code = None
            state.edit_date = None
            state.edit_current_data = None
            state.delete_code = None
            state.delete_date = None
            state.delete_records = None
            state.merge_candidates = None
            state.merge_period = None
            state.upload_file_club = None
            state.upload_file_date = None
            state.upload_file_data = None
            state.payments_upload_club = None
            state.payments_upload_date = None
            state.payments_upload_data = None
            state.payments_preview_data = None
            state.payments_name_changes = None
            state.stylist_club = None
            state.stylist_period_from = None
            state.stylist_period_to = None
            state.stylist_expenses = None
            state.stylist_errors = None
            state.club = saved_club  # Восстанавливаем клуб
            
            await update.message.reply_text(
                f"❌ Операция отменена\n\n"
                f"🏢 Клуб: {state.club}\n"
                f"Используйте кнопки меню:",
                reply_markup=get_main_keyboard()
            )
            return
    
    # Проверка ограниченного доступа (пароль 0001)
    # Список команд, доступных ТОЛЬКО при полном доступе
    restricted_commands = [
        'нал', 'безнал', 'готово', 'загрузить файл', 'загрузить зп',
        'отчет', 'список', 'экспорт', 
        'исправить', 'удалить', 'обнулить',
        'сотрудники', 'объединить', 'самозанятые', 'стилисты',
        'помощь', 'старт москвич', 'старт анора'
    ]
    
    # Проверяем режимы ввода данных
    restricted_modes = [
        'нал', 'безнал', 'awaiting_preview_date', 'awaiting_preview_action',
        'awaiting_edit_line_number', 'awaiting_edit_line_data',
        'awaiting_report_club', 'awaiting_report_period',
        'awaiting_list_club', 'awaiting_list_date',
        'awaiting_export_club', 'awaiting_export_period',
        'awaiting_edit_params', 'awaiting_edit_data',
        'awaiting_delete_choice', 'awaiting_delete_mass_club',
        'awaiting_upload_club', 'awaiting_upload_date', 'awaiting_upload_file',
        'awaiting_payments_upload_club', 'awaiting_payments_upload_date', 'awaiting_payments_upload_file',
        'awaiting_stylist_period', 'awaiting_stylist_data',
        'awaiting_merge_confirm', 'awaiting_duplicate_confirm', 'awaiting_sb_merge_confirm',
        'awaiting_salary_input', 'awaiting_employee_edit_select', 'awaiting_emp_code', 'awaiting_add_employee',
        'awaiting_emp_name', 'awaiting_emp_phone', 'awaiting_emp_tg', 'awaiting_emp_birth',
        'employee_awaiting_date', 'employee_awaiting_period'
    ]
    
    if state.limited_access:
        # Проверяем команды
        if text_lower in restricted_commands:
            await update.message.reply_text(
                "❌ Доступ запрещён\n\n"
                "У вас ограниченный доступ.\n"
                "Доступна только функция 'Выплаты'."
            )
            return
        
        # Проверяем режимы (если пользователь пытается что-то ввести в неразрешённом режиме)
        if state.mode in restricted_modes:
            await update.message.reply_text(
                "❌ Доступ запрещён\n\n"
                "У вас ограниченный доступ.\n"
                "Доступна только функция 'Выплаты'."
            )
            state.mode = None  # Сбрасываем режим
            return
    
    if state.mode == 'awaiting_delete_mass_club':
        await handle_delete_mass_club_input(update, state, text, text_lower)
        return
    
    if state.mode == 'awaiting_delete_mass_period':
        await handle_delete_mass_period_input(update, state, text, text_lower)
        return
    
    if state.mode == 'awaiting_delete_mass_confirm':
        await handle_delete_mass_confirm_text(update, state, text_lower)
        return

    if state.mode == 'awaiting_delete_employee_input':
        await handle_delete_employee_input(update, context, state, text)
        return
    
    # Обработка ввода даты для загрузки файла
    if state.mode == 'awaiting_upload_date':
        success, parsed_date, error = parse_short_date(text)
        if success:
            state.upload_file_date = parsed_date
            await update.message.reply_text(
                f"📎 ЗАГРУЗКА ФАЙЛА\n"
                f"🏢 Клуб: {state.upload_file_club}\n"
                f"📅 Дата: {parsed_date}\n\n"
                f"📄 Теперь отправьте Excel файл"
            )
            state.mode = 'awaiting_upload_file'
        else:
            await update.message.reply_text(
                f"❌ {error}\n\n"
                f"Введите дату (формат: 30,10) или напишите: отмена"
            )
        return
    
    # Обработка выбора клуба для загрузки ЗП
    if state.mode == 'awaiting_payments_upload_club':
        club_choice = text_lower
        if club_choice in ['москвич', 'анора']:
            state.payments_upload_club = 'Москвич' if club_choice == 'москвич' else 'Анора'
            await update.message.reply_text(
                f"💰 ЗАГРУЗКА ЗП\n"
                f"🏢 Клуб: {state.payments_upload_club}\n\n"
                f"📅 Введите дату (формат: 30,10):"
            )
            state.mode = 'awaiting_payments_upload_date'
        else:
            await update.message.reply_text("❌ Выберите: москвич или анора")
        return
    
    # Обработка ввода даты для загрузки ЗП
    if state.mode == 'awaiting_payments_upload_date':
        success, parsed_date, error = parse_short_date(text)
        if success:
            state.payments_upload_date = parsed_date
            await update.message.reply_text(
                f"💰 ЗАГРУЗКА ЗП\n"
                f"🏢 Клуб: {state.payments_upload_club}\n"
                f"📅 Дата: {parsed_date}\n\n"
                f"📄 Теперь отправьте Excel файл"
            )
            state.mode = 'awaiting_payments_upload_file'
        else:
            await update.message.reply_text(
                f"❌ {error}\n\n"
                f"Введите дату (формат: 30,10) или напишите: отмена"
            )
        return
    
    # НОВАЯ ЛОГИКА: обработка предпросмотра и ввода даты
    if state.mode == 'awaiting_preview_date':
        # Пытаемся распарсить дату
        success, parsed_date, error = parse_short_date(text)
        if success:
            # Сохраняем дату и показываем финальный предпросмотр
            state.preview_date = parsed_date
            await show_data_preview(update, state, show_duplicates=True)
            
            # Переходим в режим ожидания действия (ЗАПИСАТЬ/ИЗМЕНИТЬ/ОТМЕНА)
            state.mode = 'awaiting_preview_action'
            return
        else:
            await update.message.reply_text(
                f"❌ {error}\n\n"
                f"Введите дату (формат: 30,10) или напишите: отмена"
            )
            return
    
    # Обработка действий в режиме предпросмотра
    if state.mode == 'awaiting_preview_action':
        await handle_preview_action(update, state, text, text_lower)
        return
    
    # Обработка ввода номера строки для редактирования
    if state.mode == 'awaiting_edit_line_number':
        await handle_edit_line_number(update, state, text)
        return
    
    # Обработка ввода новых данных для строки
    if state.mode == 'awaiting_edit_line_data':
        await handle_edit_line_data(update, state, text)
        return
    
    # Команда "обнулить"
    if text_lower == 'обнулить':
        await update.message.reply_text(
            "⚠️ ВНИМАНИЕ! Будут удалены ВСЕ данные из базы!\n\n"
            "Для подтверждения введите пин-код:"
        )
        state.mode = 'awaiting_reset_pin'
        return
    
    # Обработка пина для обнуления
    if state.mode == 'awaiting_reset_pin':
        if text == RESET_PIN_CODE:
            # Удаляем все данные
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM operations")
            cursor.execute("DELETE FROM edit_log")
            conn.commit()
            conn.close()
            
            state.mode = None
            await update.message.reply_text(
                "✅ Все данные удалены из базы.\n"
                "База данных обнулена."
            )
        else:
            state.mode = None
            await update.message.reply_text(
                "❌ Неверный пин-код. Операция отменена."
            )
        return
    
    # Секретная команда для исправления кодов в payments
    if text_lower == 'fix payments':
        await update.message.reply_text("⏳ Исправляю коды в таблице payments...")
        
        fixed_count = db.fix_payment_codes()
        
        await update.message.reply_text(
            f"✅ Исправление завершено!\n"
            f"Обновлено записей: {fixed_count}"
        )
        return
    
    # Команда "завершить" - выход из сессии
    if text_lower == 'завершить' or text_lower == '🚪 завершить':
        # Удаляем сообщения бота (последние сохранённые)
        chat_id = update.effective_chat.id
        deleted_count = 0
        
        for msg_id in state.bot_messages[-50:]:  # Последние 50 сообщений
            try:
                await context.bot.delete_message(chat_id=chat_id, message_id=msg_id)
                deleted_count += 1
            except:
                pass  # Игнорируем ошибки (сообщение уже удалено или старое)
        
        # Очищаем состояние
        state.reset_input()
        state.club = None
        state.bot_messages = []
        state.employee_mode = False
        state.limited_access = False
        
        await update.message.reply_text(
            "👋 Сессия завершена.\n"
            f"Удалено сообщений: {deleted_count}\n\n"
            "Для повторного входа введите пин-код.",
            reply_markup=ReplyKeyboardRemove()
        )
        return
    
    # === ОБРАБОТКА КОМАНД СОТРУДНИКА ===
    if state.employee_mode:
        # Команда "выход"
        if text_lower in ['выход', '❌ выход']:
            state.employee_mode = False
            state.employee_code = None
            state.employee_club = None
            state.employee_name = None
            state.mode = None
            await update.message.reply_text(
                "👋 Сессия завершена\n\n"
                "Для повторного входа используйте /start",
                reply_markup=ReplyKeyboardRemove()
            )
            return
        
        # Команда "Моя ЗП" - последняя начисленная ЗП
        if text_lower in ['моя зп', '💰 моя зп', 'зп', '💵 зп']:
            # Получаем последнюю запись из payments
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT date, stavka, lm_3, percent_5, promo, crz, cons, tips, 
                       fines, total_shift, debt, debt_nal, to_pay
                FROM payments
                WHERE club = ? AND code = ?
                ORDER BY date DESC
                LIMIT 1
            """, (state.employee_club, state.employee_code))
            
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                await update.message.reply_text(
                    "❌ Данные о ЗП не найдены\n\n"
                    "Возможно, ЗП ещё не начислена или файл не загружен."
                )
                return
            
            date, stavka, lm_3, percent_5, promo, crz, cons, tips, fines, total_shift, debt, debt_nal, to_pay = row
            
            # Пересчитываем К выплате
            vychet_10 = round(debt * 0.1) if debt else 0
            k_vyplate = round((debt_nal or 0) + (debt or 0) - vychet_10)
            
            msg = (
                f"💰 ВАША ПОСЛЕДНЯЯ ЗП\n\n"
                f"📅 Дата: {date}\n"
                f"💼 Код: {state.employee_code}\n"
                f"👤 {state.employee_name}\n\n"
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"💵 Ставка: {int(stavka)}\n"
                f"📊 3% ЛМ: {int(lm_3)}\n"
                f"📊 5%: {int(percent_5)}\n"
                f"🎉 Промо: {int(promo)}\n"
                f"🍽 CRZ: {int(crz)}\n"
                f"🥂 Cons: {int(cons)}\n"
                f"💸 Чаевые: {int(tips)}\n"
            )
            
            if fines:
                msg += f"⚠️ Штрафы: {int(fines)}\n"
            
            msg += (
                f"━━━━━━━━━━━━━━━━━━━━━\n"
                f"💰 ИТОГО выплат: {int(total_shift)}\n"
                f"💵 Получила на смене: {int(to_pay or 0)}\n"
                f"📋 Долг БН: {int(debt or 0)}\n"
                f"📋 Долг НАЛ: {int(debt_nal or 0)}\n"
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"💎 К ВЫПЛАТЕ: {k_vyplate} ₽\n"
            )
            
            await update.message.reply_text(msg)
            return
        
        # Команда "История выплат"
        if text_lower in ['история выплат', '💵 история выплат']:
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT date, total_shift, to_pay
                FROM payments
                WHERE club = ? AND code = ?
                ORDER BY date DESC
                LIMIT 10
            """, (state.employee_club, state.employee_code))
            
            rows = cursor.fetchall()
            conn.close()
            
            if not rows:
                await update.message.reply_text("❌ История выплат пуста")
                return
            
            msg = f"💵 ИСТОРИЯ ВЫПЛАТ\n\n"
            msg += f"💼 {state.employee_code} - {state.employee_name}\n\n"
            
            for date, total, paid in rows:
                msg += f"📅 {date}: {int(total)} ₽\n"
            
            msg += f"\n📊 Всего записей: {len(rows)}"
            
            await update.message.reply_text(msg)
            return
        
        # Обработка ввода даты для ЗП
        if state.mode == 'employee_awaiting_date':
            # Парсим дату
            try:
                from datetime import datetime
                # Формат: ДД,ММ или ДД.ММ
                date_str = text.replace(',', '.').strip()
                parts = date_str.split('.')
                
                if len(parts) != 2:
                    raise ValueError
                
                day = int(parts[0])
                month = int(parts[1])
                year = datetime.now().year
                
                date_obj = datetime(year, month, day)
                date_formatted = date_obj.strftime('%Y-%m-%d')
                
            except:
                await update.message.reply_text(
                    "❌ Неверный формат даты\n\n"
                    "Используйте: ДД,ММ или ДД.ММ\n"
                    "Пример: 14,12 или 14.12"
                )
                return
            
            # Получаем ЗП за эту дату
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT date, stavka, lm_3, percent_5, promo, crz, cons, tips, 
                       fines, total_shift, debt, debt_nal, to_pay
                FROM payments
                WHERE club = ? AND code = ? AND date = ?
            """, (state.employee_club, state.employee_code, date_formatted))
            
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                await update.message.reply_text(
                    f"❌ ЗП за {date_str} не найдена\n\n"
                    f"Возможно:\n"
                    f"• В этот день не было смены\n"
                    f"• Файл ещё не загружен\n"
                    f"• Неверная дата"
                )
                state.mode = None
                return
            
            date, stavka, lm_3, percent_5, promo, crz, cons, tips, fines, total_shift, debt, debt_nal, to_pay = row
            
            vychet_10 = round(debt * 0.1) if debt else 0
            k_vyplate = round((debt_nal or 0) + (debt or 0) - vychet_10)
            
            msg = (
                f"💰 ЗП ЗА {date_str}\n\n"
                f"📅 {date}\n"
                f"💼 {state.employee_code}\n"
                f"👤 {state.employee_name}\n\n"
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"💵 Ставка: {int(stavka)}\n"
                f"📊 3% ЛМ: {int(lm_3)}\n"
                f"📊 5%: {int(percent_5)}\n"
                f"🎉 Промо: {int(promo)}\n"
                f"🍽 CRZ: {int(crz)}\n"
                f"🥂 Cons: {int(cons)}\n"
                f"💸 Чаевые: {int(tips)}\n"
            )
            
            if fines:
                msg += f"⚠️ Штрафы: {int(fines)}\n"
            
            msg += (
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"💰 ИТОГО: {int(total_shift)}\n"
                f"💎 К ВЫПЛАТЕ: {k_vyplate} ₽\n"
            )
            
            await update.message.reply_text(msg)
            state.mode = None
            return
        
        # Обработка ввода периода для ЗП
        if state.mode == 'employee_awaiting_period':
            # Парсим период
            try:
                from datetime import datetime
                # Формат: ДД,ММ-ДД,ММ
                period = text.replace('.', ',').strip()
                parts = period.split('-')
                
                if len(parts) != 2:
                    raise ValueError
                
                # Дата от
                date_from_parts = parts[0].split(',')
                day_from = int(date_from_parts[0])
                month_from = int(date_from_parts[1])
                year = datetime.now().year
                date_from = datetime(year, month_from, day_from).strftime('%Y-%m-%d')
                
                # Дата до
                date_to_parts = parts[1].split(',')
                day_to = int(date_to_parts[0])
                month_to = int(date_to_parts[1])
                date_to = datetime(year, month_to, day_to).strftime('%Y-%m-%d')
                
            except:
                await update.message.reply_text(
                    "❌ Неверный формат периода\n\n"
                    "Используйте: ДД,ММ-ДД,ММ\n"
                    "Пример: 14,12-20,12"
                )
                return
            
            # Получаем все ЗП за период
            conn = db.get_connection()
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT date, stavka, lm_3, percent_5, promo, crz, cons, tips, 
                       fines, total_shift, debt, debt_nal, to_pay
                FROM payments
                WHERE club = ? AND code = ? AND date BETWEEN ? AND ?
                ORDER BY date
            """, (state.employee_club, state.employee_code, date_from, date_to))
            
            rows = cursor.fetchall()
            conn.close()
            
            if not rows:
                await update.message.reply_text(
                    f"❌ ЗП за период {parts[0]}-{parts[1]} не найдена"
                )
                state.mode = None
                return
            
            # Суммируем
            total_stavka = sum(r[1] for r in rows)
            total_lm_3 = sum(r[2] for r in rows)
            total_percent_5 = sum(r[3] for r in rows)
            total_promo = sum(r[4] for r in rows)
            total_crz = sum(r[5] for r in rows)
            total_cons = sum(r[6] for r in rows)
            total_tips = sum(r[7] for r in rows)
            total_fines = sum(r[8] for r in rows)
            total_shift = sum(r[9] for r in rows)
            total_debt = sum(r[10] for r in rows)
            total_debt_nal = sum(r[11] for r in rows)
            
            vychet_10 = round(total_debt * 0.1)
            k_vyplate = round(total_debt_nal + total_debt - vychet_10)
            
            msg = (
                f"💰 ЗП ЗА ПЕРИОД\n\n"
                f"📅 {parts[0]} - {parts[1]}\n"
                f"💼 {state.employee_code}\n"
                f"👤 {state.employee_name}\n"
                f"📊 Смен: {len(rows)}\n\n"
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"💵 Ставка: {int(total_stavka)}\n"
                f"📊 3% ЛМ: {int(total_lm_3)}\n"
                f"📊 5%: {int(total_percent_5)}\n"
                f"🎉 Промо: {int(total_promo)}\n"
                f"🍽 CRZ: {int(total_crz)}\n"
                f"🥂 Cons: {int(total_cons)}\n"
                f"💸 Чаевые: {int(total_tips)}\n"
            )
            
            if total_fines:
                msg += f"⚠️ Штрафы: {int(total_fines)}\n"
            
            msg += (
                f"━━━━━━━━━━━━━━━━━━━\n"
                f"💰 ИТОГО: {int(total_shift)}\n"
                f"💎 К ВЫПЛАТЕ: {k_vyplate} ₽\n"
            )
            
            await update.message.reply_text(msg)
            state.mode = None
            return
    
    # Сопоставление кнопок с командами
    button_commands = {
        '🏢 старт москвич': 'старт москвич',
        '🏢 старт анора': 'старт анора',
        '📥 нал': 'нал',
        '📥 безнал': 'безнал',
        '📎 загрузить файл': 'загрузить файл',
        '💰 загрузить зп': 'загрузить зп',
        '✅ готово': 'готово',
        '❌ отмена': 'отмена',
        '📊 отчёт': 'отчет',
        '📊 отчет': 'отчет',
        '💰 выплаты': 'выплаты',
        '💵 зп': 'зп',
        '📋 список': 'список',
        '📤 экспорт': 'экспорт',
        '✏️ исправить': 'исправить',
        '🗑️ удалить': 'удалить',
        '📜 журнал': 'журнал',
        '👔 самозанятые': 'самозанятые',
        '👥 сотрудники': 'сотрудники',
        '💄 стилисты': 'стилисты',
        '❓ помощь': 'помощь',
        '🚪 завершить': 'завершить'
    }
    
    # Если нажата кнопка - преобразуем в команду
    if text_lower in button_commands:
        text_lower = button_commands[text_lower]
    
    # Команда "старт москвич" или "старт анора" - обрабатываем ПЕРВОЙ (после преобразования кнопок!)
    if text_lower.startswith('старт'):
        # Если в режиме ввода данных - предупреждение
        if state.has_data() and state.mode not in ['awaiting_date', 'awaiting_duplicate_confirm', 'awaiting_sb_merge_confirm', 'awaiting_merge_confirm']:
            await update.message.reply_text(
                "⚠️ У вас есть несохранённые данные!\n"
                "Завершите ввод командой: готово\n"
                "Или отмените: отмена"
            )
            return
        await start_command(update, context)
        return
    
    # Команда "кнопки" - показать клавиатуру
    if text_lower == 'кнопки':
        # Блокируем для ограниченного доступа
        if state.limited_access:
            await update.message.reply_text(
                "❌ Доступ запрещён\n\n"
                "У вас ограниченный доступ.\n"
                "Доступна только функция 'Выплаты'."
            )
            return
        
        if state.club:
            await update.message.reply_text(
                "Клавиатура:",
                reply_markup=get_main_keyboard()
            )
        else:
            await update.message.reply_text(
                "Сначала выберите клуб:",
                reply_markup=get_club_keyboard()
            )
        return
    
    # Команда "помощь"
    if text_lower in ['помощь', 'help']:
        await update.message.reply_text(
            "📋 ПОЛНАЯ СПРАВКА ПО КОМАНДАМ\n\n"
            "🏢 НАЧАЛО РАБОТЫ:\n"
            "• Выберите клуб: СТАРТ МОСКВИЧ / СТАРТ АНОРА\n"
            "• После выбора используйте кнопки меню\n\n"
            "💰 ВВОД ДАННЫХ:\n"
            "1️⃣ Нажмите НАЛ или БЕЗНАЛ\n"
            "2️⃣ Вставьте список данных\n"
            "3️⃣ Нажмите ГОТОВО → предпросмотр\n"
            "4️⃣ Укажите дату (например: 3,10)\n"
            "5️⃣ Проверьте данные в предпросмотре\n"
            "6️⃣ ЗАПИСАТЬ - сохранить в базу\n\n"
            "🔍 ПРЕДПРОСМОТР:\n"
            "После ГОТОВО вы увидите все данные с номерами строк\n"
            "• ЗАПИСАТЬ → сохранить данные\n"
            "• ИЗМЕНИТЬ → редактировать строку (укажите номер)\n"
            "• ОТМЕНА → отменить ввод\n"
            "• Если есть дубликаты → команды для объединения\n\n"
            "🔄 ОБЪЕДИНЕНИЕ ДУБЛИКАТОВ:\n"
            "Если найдены записи с одним кодом (с именем и без):\n"
            "• ОК → объединить все\n"
            "• ОК 1 → объединить только пункт 1\n"
            "• ОК 1 2 → объединить пункты 1 и 2\n"
            "• НЕ 1 → НЕ объединять пункт 1 (остальные да)\n"
            "• НЕ 1 2 → НЕ объединять пункты 1 и 2\n\n"
            "📊 ОТЧЁТЫ:\n"
            "• ОТЧЁТ → выбрать клуб → указать период\n"
            "• ВЫПЛАТЫ → код + период (Д7 3,10-5,11)\n"
            "• Получите Excel файл с отчётом\n\n"
            "📝 ПРОСМОТР И РЕДАКТИРОВАНИЕ:\n"
            "• СПИСОК → клуб → дата (посмотреть все записи)\n"
            "• ИСПРАВИТЬ → код + дата (Д7 3,10)\n"
            "• УДАЛИТЬ → код + дата (Д7 3,10)\n"
            "• УДАЛИТЬ ВСЕ → клуб → дата/период (массовое удаление)\n\n"
            "📤 ЭКСПОРТ:\n"
            "• ЭКСПОРТ → клуб → период → Excel файл\n\n"
            "📜 ЖУРНАЛ ИЗМЕНЕНИЙ:\n"
            "• ЖУРНАЛ → последние 20 изменений\n"
            "• ЖУРНАЛ 50 → последние 50 изменений\n"
            "• ЖУРНАЛ Д7 → все изменения по коду Д7\n"
            "• ЖУРНАЛ 3,10 → все изменения за дату\n"
            "Показывает: объединения, исправления, удаления\n\n"
            "🔧 ДОПОЛНИТЕЛЬНО:\n"
            "• ОБНУЛИТЬ → удалить все данные (нужен пин)\n"
            "• ЗАВЕРШИТЬ → выход (очистка истории)\n\n"
            "📖 ФОРМАТЫ ДАТ:\n"
            "• 3,10 = 03.10.2025\n"
            "• 30,10 = 30.10.2025\n"
            "• 3,10-5,11 = период с 3.10 по 5.11\n\n"
            "📝 ФОРМАТЫ ДАННЫХ:\n"
            "• Д7 Надя 6800 или Д7 Надя-6800\n"
            "• Юля Д17 1000\n"
            "• СБ Дмитрий 4000\n"
            "• Уборщица-2000\n"
            "• Суммы: 40,000 или 40.000 → 40000 ✅\n\n"
            "✨ АВТОМАТИЧЕСКАЯ ОЧИСТКА:\n"
            "• Дубли из Excel очищаются автоматически\n"
            "• Разделители тысяч (точки/запятые) удаляются\n"
            "• В предпросмотре видно что было изменено"
        )
        return
    
    # Обработка подтверждения объединения дубликатов
    if state.mode == 'awaiting_duplicate_confirm':
        await handle_duplicate_confirmation(update, context, state, text, text_lower)
        return
    
    if state.mode == 'awaiting_sb_merge_confirm':
        await handle_sb_merge_confirmation(update, context, state, text, text_lower)
        return
    
    # Обработка подтверждения загрузки файла
    if state.mode == 'awaiting_upload_confirm':
        if text_lower == 'отмена' or text_lower == '❌ отмена':
            state.upload_file_club = None
            state.upload_file_date = None
            state.upload_file_data = None
            state.mode = None
            await update.message.reply_text(
                "❌ Загрузка файла отменена\n\n"
                "Используйте кнопки меню:",
                reply_markup=get_main_keyboard()
            )
            return
        elif text_lower.startswith('записать'):
            # Парсим команду
            # Варианты: "записать", "записать 1 2", "записать без 3"
            selected_merges = None  # None = все, [] = без объединений, [1,2] = только указанные
            
            if text_lower == 'записать':
                # Применить все объединения
                selected_merges = None
            elif 'без' in text_lower:
                # ЗАПИСАТЬ БЕЗ 1 2 3 - исключить указанные
                parts = text_lower.replace('записать', '').replace('без', '').strip().split()
                excluded = []
                for part in parts:
                    try:
                        excluded.append(int(part))
                    except ValueError:
                        pass
                
                if excluded:
                    # Получаем все ID объединений
                    data = state.upload_file_data
                    beznal_analysis = data.get('beznal_analysis', {})
                    nal_analysis = data.get('nal_analysis', {})
                    
                    all_merge_ids = []
                    for merge in beznal_analysis.get('merges', []):
                        if 'merge_id' in merge:
                            all_merge_ids.append(merge['merge_id'])
                    for merge in nal_analysis.get('merges', []):
                        if 'merge_id' in merge:
                            all_merge_ids.append(merge['merge_id'])
                    
                    # Все кроме исключенных
                    selected_merges = [mid for mid in all_merge_ids if mid not in excluded]
                else:
                    selected_merges = None  # Нет исключений - все
            else:
                # ЗАПИСАТЬ 1 2 3 - только указанные
                parts = text_lower.replace('записать', '').strip().split()
                selected = []
                for part in parts:
                    try:
                        selected.append(int(part))
                    except ValueError:
                        pass
                
                if selected:
                    selected_merges = selected
                else:
                    selected_merges = None  # Не смогли распарсить - применяем все
            
            # Сохраняем выбор и сохраняем данные
            state.upload_file_data['selected_merges'] = selected_merges
            await save_file_data(update, state)
            return
        else:
            await update.message.reply_text(
                "⚠️ Неверная команда. Введите:\n"
                "  • ЗАПИСАТЬ - применить все\n"
                "  • ЗАПИСАТЬ 1 2 - только [1] и [2]\n"
                "  • ЗАПИСАТЬ БЕЗ 3 - все кроме [3]\n"
                "  • ОТМЕНА - отменить"
            )
            return
    
    # Обработка подтверждения сохранения ЗП
    
    # Команда "нал"
    if text_lower == 'нал':
        if not state.club:
            await update.message.reply_text(
                "❌ Клуб не выбран.\n"
                "Используйте: старт москвич или старт анора"
            )
        else:
            state.mode = 'нал'
            await update.message.reply_text(
                f"📥 РЕЖИМ ВВОДА: НАЛ\n\n"
                f"🏢 Клуб: {state.club}\n\n"
                f"📝 Вставьте список данных:\n"
                f"Примеры форматов:\n"
                f"  • Д7 Юля 1000\n"
                f"  • Д7 Юля-1000\n"
                f"  • Юля Д7 1000\n\n"
                f"⏭️ После ввода всех данных (НАЛ и БЕЗНАЛ)\n"
                f"   нажмите: ГОТОВО"
            )
        return
    
    # Команда "безнал"
    if text_lower == 'безнал':
        if not state.club:
            await update.message.reply_text(
                "❌ Клуб не выбран.\n"
                "Используйте: старт москвич или старт анора"
            )
        else:
            state.mode = 'безнал'
            await update.message.reply_text(
                f"📥 РЕЖИМ ВВОДА: БЕЗНАЛ\n\n"
                f"🏢 Клуб: {state.club}\n\n"
                f"📝 Вставьте список данных:\n"
                f"Примеры форматов:\n"
                f"  • Д7 Юля 1000\n"
                f"  • Д7 Юля-1000\n"
                f"  • Юля Д7 1000\n\n"
                f"⏭️ После ввода всех данных (НАЛ и БЕЗНАЛ)\n"
                f"   нажмите: ГОТОВО"
            )
        return
    
    # Команда "загрузить файл"
    if text_lower == 'загрузить файл':
        if state.has_data():
            await update.message.reply_text(
                "⚠️ У вас есть несохранённые данные!\n"
                "Завершите ввод командой: готово\n"
                "Или отмените: отмена"
            )
            return
        
        await update.message.reply_text(
            "📎 ЗАГРУЗКА EXCEL ФАЙЛА\n\n"
            "Выберите клуб:",
            reply_markup=get_club_keyboard()
        )
        state.mode = 'awaiting_upload_club'
        return
    
    # Команда "загрузить зп"
    if text_lower == 'загрузить зп' or text_lower == '💰 загрузить зп':
        if state.has_data():
            await update.message.reply_text(
                "⚠️ У вас есть несохранённые данные!\n"
                "Завершите ввод командой: готово\n"
                "Или отмените: отмена"
            )
            return
        
        await update.message.reply_text(
            "💰 ЗАГРУЗКА ЛИСТА ВЫПЛАТ\n\n"
            "Выберите клуб:",
            reply_markup=get_club_keyboard()
        )
        state.mode = 'awaiting_payments_upload_club'
        return
    
    # Обработка ввода периода для расходов на стилистов
    if state.mode == 'awaiting_stylist_period':
        # Парсим период
        if '-' in text:
            success, date_from, date_to, error = parse_date_range(text)
            if not success:
                await update.message.reply_text(f"❌ {error}\n\n❌ Для отмены напишите: ОТМЕНА")
                return
        else:
            success, single_date, error = parse_short_date(text)
            if not success:
                await update.message.reply_text(f"❌ {error}\n\n❌ Для отмены напишите: ОТМЕНА")
                return
            date_from = single_date
            date_to = single_date
        
        # Сохраняем период
        state.stylist_period_from = date_from
        state.stylist_period_to = date_to
        state.stylist_expenses = []  # Инициализируем пустой список
        state.stylist_errors = []
        state.mode = 'awaiting_stylist_data'
        
        # Создаем inline кнопку ГОТОВО
        keyboard = [[InlineKeyboardButton("✅ ГОТОВО", callback_data='stylist_done')]]
        
        await update.message.reply_text(
            f"✅ ПЕРИОД: {date_from} - {date_to}\n\n"
            f"💄 Отправьте данные о расходах.\n\n"
            f"Формат:\n"
            f"Д14Бритни 2000\n"
            f"А13Варя 1500\n"
            f"Н3Влада 2500\n\n"
            f"📝 Можете отправлять НЕСКОЛЬКО сообщений.\n"
            f"После завершения нажмите: ГОТОВО",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return
    
    # Обработка ввода данных стилистов (накопление из нескольких сообщений)
    if state.mode == 'awaiting_stylist_data':
        await handle_stylist_data_input(update, state, text, text_lower)
        return
    
    # Обработка подтверждения сохранения расходов на стилистов
    if state.mode == 'awaiting_stylist_confirm':
        await handle_stylist_confirm(update, state, text_lower)
        return
    
    # Обработка ввода номера для редактирования расхода на стилиста
    if state.mode == 'awaiting_stylist_edit_number':
        await handle_stylist_edit_number(update, state, text)
        return
    
    # Обработка ввода новых данных для расхода на стилиста
    if state.mode == 'awaiting_stylist_edit_data':
        await handle_stylist_edit_data(update, state, text)
        return
    
    # Обработка уточнений для расходов на стилистов (выбор имени)
    if state.mode == 'awaiting_stylist_clarification':
        await handle_stylist_clarification(update, state, text)
        return
    
    # Обработка удаления записей стилистов при просмотре
    if state.mode == 'awaiting_stylist_view_delete':
        await handle_stylist_view_delete(update, state, text)
        return
    
    # Обработка редактирования записей стилистов при просмотре
    if state.mode == 'awaiting_stylist_view_edit':
        await handle_stylist_view_edit_number(update, state, text)
        return
    
    # Обработка ввода новых данных при редактировании
    if state.mode == 'awaiting_stylist_view_edit_data':
        await handle_stylist_view_edit_data(update, state, text)
        return
    
    # Команда "готово"
    if text_lower == 'готово':
        if not state.has_data():
            await update.message.reply_text(
                "❌ Нет данных для обработки.\n"
                "Используйте команды: нал и безнал для ввода данных"
            )
            return
        
        # Показываем предпросмотр данных
        await show_data_preview(update, state, show_duplicates=True)
        
        # Переходим в режим ожидания даты (сначала нужно указать дату)
        state.mode = 'awaiting_preview_date'
        return
    
    # Блочный ввод данных (но проверяем сначала - это не команда/кнопка!)
    if state.mode in ['нал', 'безнал']:
        # Проверяем - это команда или кнопка?
        # Если текст начинается с emoji кнопок или это известная команда - НЕ парсим как данные
        emoji_buttons = ['📥', '✅', '❌', '📊', '💰', '📋', '📤', '✏️', '🗑️', '❓', '🚪']
        is_button = any(text.startswith(emoji) for emoji in emoji_buttons)
        
        if is_button or text_lower in ['отмена', 'готово', 'отчет', 'список', 'экспорт', 'помощь']:
            # Это команда/кнопка - НЕ парсим как данные, пропускаем дальше
            pass
        else:
            # Это данные - парсим
            successful, errors = DataParser.parse_block(text)
            
            if successful:
                # Сохраняем в соответствующий список
                if state.mode == 'нал':
                    state.temp_nal_data.extend(successful)
                else:
                    state.temp_beznal_data.extend(successful)
            
            if errors:
                error_msg = "⚠️ Ошибки при парсинге:\n" + '\n'.join(errors[:5])
                if len(errors) > 5:
                    error_msg += f"\n... и ещё {len(errors) - 5} ошибок"
                await update.message.reply_text(error_msg)
            
            return
    
    # Команда "отчет"
    if text_lower == 'отчет':
        await update.message.reply_text(
            "Выберите клуб:",
            reply_markup=get_club_report_keyboard()
        )
        state.mode = 'awaiting_report_club'
        return
    
    # Обработка выбора клуба для отчета
    if state.mode == 'awaiting_report_club':
        if text_lower in ['москвич', 'анора', 'оба']:
            state.report_club = text_lower
            await update.message.reply_text(
                "Укажите дату или период:\n"
                "• Одна дата: 12,12\n"
                "• Период: 10,06-11,08"
            )
            state.mode = 'awaiting_report_period'
        else:
            await update.message.reply_text(
                "❌ Неверный выбор. Выберите: москвич, анора или оба"
            )
        return
    
    # Обработка периода для отчета
    if state.mode == 'awaiting_report_period':
        # Проверяем, это одна дата или диапазон
        if '-' in text:
            # Диапазон дат: 10,06-11,08
            success, date_from, date_to, error = parse_date_range(text)
            if not success:
                await update.message.reply_text(f"❌ {error}")
                return
        else:
            # Одна дата: 12,12
            success, single_date, error = parse_short_date(text)
            if not success:
                await update.message.reply_text(f"❌ {error}")
                return
            date_from = single_date
            date_to = single_date
        
        # Генерируем отчет
        if state.report_club == 'оба':
            # Инициализируем отслеживание обработанных клубов
            state.processed_clubs_for_report = set()
            
            # Сохраняем период для дальнейшего использования
            state.pending_report_period = (date_from, date_to)
            
            # Сначала отчеты по каждому клубу
            for club in ['Москвич', 'Анора']:
                await generate_and_send_report(update, club, date_from, date_to, state)
                # Если generate_and_send_report установил режим awaiting_duplicate_confirm или awaiting_sb_merge_confirm - выходим
                # НО НЕ ПРЕРЫВАЕМ ВЕСЬ ПРОЦЕСС! Ждём подтверждения пользователя
                if state.mode in ['awaiting_duplicate_confirm', 'awaiting_sb_merge_confirm']:
                    return
            
            # Затем проверяем возможность сводного отчета
            await prepare_merged_report(update, state, date_from, date_to)
            
            # НЕ сбрасываем режим если ждём подтверждения объединения!
            if state.mode != 'awaiting_merge_confirm':
                state.mode = None
                state.report_club = None
                state.pending_report_period = None
        else:
            club = 'Москвич' if state.report_club == 'москвич' else 'Анора'
            await generate_and_send_report(update, club, date_from, date_to, state)
            
            # НЕ сбрасываем режим если ждём подтверждения дубликатов!
            if state.mode not in ['awaiting_duplicate_confirm', 'awaiting_sb_merge_confirm']:
                state.mode = None
                state.report_club = None
        return
    
    # Обработка подтверждения объединения для сводного отчета
    if state.mode == 'awaiting_merge_confirm':
        await handle_merge_confirmation(update, state, text_lower)
        return
    
    # Команда "выплаты"
    if text_lower.startswith('выплаты') or text_lower == 'выплаты':
        if text_lower == 'выплаты':
            # Нажата кнопка - переходим в режим ожидания
            await update.message.reply_text(
                "Выплаты за период\n\n"
                "Введите сотрудника и период:\n\n"
                "Примеры:\n"
                "• Д7 12,12\n"
                "• Д7 10,06-11,08"
            )
            state.mode = 'awaiting_payments_input'
        else:
            await handle_payments_command(update, context, state, text)
        return
    
    # Команда "зп" (новый расширенный отчёт из таблицы payments)
    if text_lower.startswith('зп') or text_lower == 'зп':
        if text_lower == 'зп':
            # Нажата кнопка - переходим в режим ожидания
            await update.message.reply_text(
                "💵 ОТЧЁТ ЗП\n\n"
                "Введите:\n"
                "• Код + период (Д7 3,10-5,11)\n"
                "• Или клуб + период (москвич 3,10-5,11)\n"
                "• Или оба + период (оба 3,10-5,11)"
            )
            state.mode = 'awaiting_salary_input'
        else:
            await handle_salary_command(update, context, state, text)
        return
    
    # Обработка ввода для ЗП (после кнопки)
    if state.mode == 'awaiting_salary_input':
        await handle_salary_command(update, context, state, text)
        state.mode = None
        return
    
    # Обработка ввода для выплат (после кнопки)
    if state.mode == 'awaiting_payments_input':
        await handle_payments_command(update, context, state, text)
        # Для ограниченного доступа НЕ сбрасываем режим - оставляем в цикле
        if not state.limited_access:
            state.mode = None
        return
    
    # Команда "список"
    if text_lower.startswith('список') or text_lower == 'список':
        if text_lower == 'список':
            await update.message.reply_text(
                "📋 Выберите клуб для просмотра записей:",
                reply_markup=get_club_report_keyboard()
            )
            state.mode = 'awaiting_list_club'
        else:
            await handle_list_command(update, context, state, text)
        return
    
    # Обработка выбора клуба для списка
    if state.mode == 'awaiting_list_club':
        club_choice = text_lower
        if club_choice in ['москвич', 'анора', 'оба']:
            state.list_club = club_choice
            await update.message.reply_text(
                "📅 Введите дату:\n\n"
                "Примеры:\n"
                "• 3,11\n"
                "• 30,10"
            )
            state.mode = 'awaiting_list_date'
        else:
            await update.message.reply_text("❌ Выберите: москвич, анора или оба")
        return
    
    # Обработка ввода даты для списка
    if state.mode == 'awaiting_list_date':
        success, parsed_date, error = parse_short_date(text)
        if success:
            if state.list_club == 'оба':
                # Показываем списки для обоих клубов
                for club in ['Москвич', 'Анора']:
                    operations = db.get_operations_by_date(club, parsed_date)
                    response = format_operations_list(operations, parsed_date, club)
                    await update.message.reply_text(response)
            else:
                # Показываем список для одного клуба
                club = 'Москвич' if state.list_club == 'москвич' else 'Анора'
                operations = db.get_operations_by_date(club, parsed_date)
                response = format_operations_list(operations, parsed_date, club)
                await update.message.reply_text(response)
            
            state.mode = None
            state.list_club = None
        else:
            await update.message.reply_text(f"❌ {error}")
        return
    
    # Команда "исправить"
    if text_lower.startswith('исправить') or text_lower == 'исправить':
        if text_lower == 'исправить':
            await update.message.reply_text(
                "📝 Введите код и дату:\n\n"
                "Примеры:\n"
                "• Д7 3,11\n"
                "• Д1 30,10"
            )
            state.mode = 'awaiting_edit_params'
        else:
            await handle_edit_command_new(update, context, state, text)
        return
    
    # Обработка ввода параметров для исправления
    if state.mode == 'awaiting_edit_params':
        await handle_edit_command_new(update, context, state, text)
        return
    
    # Обработка ввода новых данных для исправления
    if state.mode == 'awaiting_edit_data':
        await handle_edit_input(update, context, state, text, text_lower)
        return
    
    # Команда "удалить"
    if text_lower.startswith('удалить') or text_lower == 'удалить':
        if text_lower == 'удалить':
            await update.message.reply_text(
                "Формат: удалить КОД дата\n\n"
                "Примеры:\n"
                "• удалить Д7 12,12\n"
                "• удалить Д1 30,10\n\n"
                "Массовое удаление:\n"
                "• удалить все"
            , reply_markup=get_delete_mode_keyboard())
        else:
            await handle_delete_command_new(update, context, state, text)
        return
    
    # Обработка выбора что удалить
    if state.mode == 'awaiting_delete_choice':
        await handle_delete_choice(update, context, state, text_lower)
        return
    
    # Команда "журнал"
    if text_lower.startswith('журнал') or text_lower == 'журнал' or text_lower == '📜 журнал':
        await handle_journal_command(update, context, state, text)
        return
    
    # Команда "самозанятые"
    if text_lower in ['самозанятые', '👔 самозанятые']:
        await handle_self_employed_command(update, context, state)
        return
    
    # Команда "сотрудники"
    if text_lower in ['сотрудники', '👥 сотрудники']:
        await update.message.reply_text(
            "👥 УПРАВЛЕНИЕ СОТРУДНИКАМИ\n\n"
            "Выберите действие:",
            reply_markup=get_employees_menu_keyboard()
        )
        return
    
    # Команда "стилисты"
    if text_lower in ['стилисты', '💄 стилисты']:
        print(f"DEBUG: Обработка команды СТИЛИСТЫ")
        keyboard = [
            [InlineKeyboardButton("💄 Загрузить расходы", callback_data='stylist_load')],
            [InlineKeyboardButton("📋 Показать расходы", callback_data='stylist_view')]
        ]
        await update.message.reply_text(
            "💄 УПРАВЛЕНИЕ РАСХОДАМИ НА СТИЛИСТОВ\n\n"
            "Выберите действие:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return
    
    # Обработка режима добавления самозанятого
    if state.mode == 'awaiting_self_employed_add':
        await handle_self_employed_add(update, state, text)
        return
    
    # Обработка режима удаления самозанятого
    if state.mode == 'awaiting_self_employed_remove':
        await handle_self_employed_remove(update, state, text)
        return
    
    # Обработка добавления доступа
    # === ОБРАБОТКА ВВОДА ПРИ РЕДАКТИРОВАНИИ СОТРУДНИКОВ ===
    
    if state.mode == 'awaiting_emp_name':
        emp = state.edit_employee_selected
        new_name = text.strip()
        
        if not new_name:
            await update.message.reply_text("❌ Имя не может быть пустым")
            return
        
        # Обновляем в БД
        conn = db.get_connection()
        cursor = conn.cursor()
        
        from datetime import datetime
        cursor.execute("""
            UPDATE employees
            SET full_name = ?, updated_at = ?
            WHERE code = ? AND club = ?
        """, (new_name, datetime.now().isoformat(), emp['code'], state.edit_employees_club))
        
        conn.commit()
        conn.close()
        
        await update.message.reply_text(
            f"✅ ИМЯ ИЗМЕНЕНО\n\n"
            f"Код: {emp['code']}\n"
            f"Было: {emp['name']}\n"
            f"Стало: {new_name}"
        )
        
        state.mode = None
        state.edit_employee_selected = None
        return
    
    if state.mode == 'awaiting_emp_phone':
        emp = state.edit_employee_selected
        
        if text_lower == 'удалить':
            new_phone = None
            action = "удалён"
        else:
            new_phone = text.strip()
            action = "изменён"
        
        # Обновляем в БД
        conn = db.get_connection()
        cursor = conn.cursor()
        
        from datetime import datetime
        cursor.execute("""
            UPDATE employees
            SET phone = ?, updated_at = ?
            WHERE code = ? AND club = ?
        """, (new_phone, datetime.now().isoformat(), emp['code'], state.edit_employees_club))
        
        conn.commit()
        conn.close()
        
        await update.message.reply_text(
            f"✅ ТЕЛЕФОН {action.upper()}\n\n"
            f"Код: {emp['code']}\n"
            f"Имя: {emp['name']}\n"
            f"Телефон: {new_phone or 'удалён'}"
        )
        
        state.mode = None
        state.edit_employee_selected = None
        return
    
    if state.mode == 'awaiting_emp_tg':
        emp = state.edit_employee_selected
        
        if text_lower == 'удалить':
            new_tg = None
            action = "удалён (доступ отключён)"
        else:
            try:
                new_tg = int(text.strip())
                action = "изменён (доступ активен)"
            except:
                await update.message.reply_text("❌ Telegram ID должен быть числом")
                return
        
        # Обновляем в БД
        conn = db.get_connection()
        cursor = conn.cursor()
        
        from datetime import datetime
        cursor.execute("""
            UPDATE employees
            SET telegram_user_id = ?, updated_at = ?
            WHERE code = ? AND club = ?
        """, (new_tg, datetime.now().isoformat(), emp['code'], state.edit_employees_club))
        
        conn.commit()
        conn.close()
        
        await update.message.reply_text(
            f"✅ TELEGRAM ID {action.upper()}\n\n"
            f"Код: {emp['code']}\n"
            f"Имя: {emp['name']}\n"
            f"Telegram ID: {new_tg or 'удалён'}"
        )
        
        state.mode = None
        state.edit_employee_selected = None
        return
    
    if state.mode == 'awaiting_emp_birth':
        emp = state.edit_employee_selected
        
        if text_lower == 'удалить':
            new_birth = None
            action = "удалена"
        else:
            # Парсим дату
            try:
                from datetime import datetime
                birth_date = datetime.strptime(text.strip(), '%d.%m.%Y')
                new_birth = birth_date.strftime('%Y-%m-%d')
                action = "изменена"
            except:
                await update.message.reply_text(
                    "❌ Неверный формат даты\n\n"
                    "Используйте: ДД.ММ.ГГГГ\n"
                    "Пример: 15.03.1998"
                )
                return
        
        # Обновляем в БД
        conn = db.get_connection()
        cursor = conn.cursor()
        
        from datetime import datetime
        cursor.execute("""
            UPDATE employees
            SET birth_date = ?, updated_at = ?
            WHERE code = ? AND club = ?
        """, (new_birth, datetime.now().isoformat(), emp['code'], state.edit_employees_club))
        
        conn.commit()
        conn.close()
        
        display_birth = datetime.strptime(new_birth, '%Y-%m-%d').strftime('%d.%m.%Y') if new_birth else 'удалена'
        
        await update.message.reply_text(
            f"✅ ДАТА РОЖДЕНИЯ {action.upper()}\n\n"
            f"Код: {emp['code']}\n"
            f"Имя: {emp['name']}\n"
            f"Дата рождения: {display_birth}"
        )
        
        state.mode = None
        state.edit_employee_selected = None
        return
    
    if state.mode == 'awaiting_emp_code':
        emp = state.edit_employee_selected
        new_code = DataParser.normalize_code(text.strip())
        
        if not new_code:
            await update.message.reply_text("❌ Код не может быть пустым")
            return
        
        # Проверяем: нет ли уже такого кода
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT code FROM employees
            WHERE code = ? AND club = ? AND code != ?
        """, (new_code, state.edit_employees_club, emp['code']))
        
        if cursor.fetchone():
            conn.close()
            await update.message.reply_text(
                f"❌ Код {new_code} уже используется другим сотрудником в клубе {state.edit_employees_club}"
            )
            return
        
        old_code = emp['code']
        
        from datetime import datetime
        now = datetime.now().isoformat()
        
        # Обновляем в employees
        cursor.execute("""
            UPDATE employees
            SET code = ?, updated_at = ?
            WHERE code = ? AND club = ?
        """, (new_code, now, old_code, state.edit_employees_club))
        
        # Обновляем в operations
        cursor.execute("""
            UPDATE operations
            SET code = ?
            WHERE code = ? AND club = ?
        """, (new_code, old_code, state.edit_employees_club))
        
        # Обновляем в payments
        cursor.execute("""
            UPDATE payments
            SET code = ?
            WHERE code = ? AND club = ?
        """, (new_code, old_code, state.edit_employees_club))
        
        # Обновляем в employee_merges
        cursor.execute("""
            UPDATE employee_merges
            SET merged_code = ?
            WHERE merged_code = ? AND club = ?
        """, (new_code, old_code, state.edit_employees_club))
        
        cursor.execute("""
            UPDATE employee_merges
            SET original_code = ?
            WHERE original_code = ? AND club = ?
        """, (new_code, old_code, state.edit_employees_club))
        
        conn.commit()
        conn.close()
        
        await update.message.reply_text(
            f"✅ КОД ИЗМЕНЁН\n\n"
            f"Было: {old_code}\n"
            f"Стало: {new_code}\n\n"
            f"Обновлены все связанные записи в БД"
        )
        
        state.mode = None
        state.edit_employee_selected = None
        return
    
    if state.mode == 'awaiting_add_employee':
        parts = text.split(maxsplit=2)
        
        if len(parts) < 2:
            await update.message.reply_text(
                "❌ Неверный формат\n\n"
                "Минимум: КОД ИМЯ\n"
                "Пример: Д1 Юлия"
            )
            return
        
        code = DataParser.normalize_code(parts[0])
        name = parts[1]
        
        # Парсим дату найма
        if len(parts) > 2:
            try:
                from datetime import datetime
                hired = datetime.strptime(parts[2], '%d.%m.%Y').strftime('%Y-%m-%d')
            except:
                await update.message.reply_text(
                    "❌ Неверный формат даты\n\n"
                    "Используйте: ДД.ММ.ГГГГ"
                )
                return
        else:
            from datetime import datetime
            hired = datetime.now().strftime('%Y-%m-%d')
        
        # Проверяем существование
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT code FROM employees
            WHERE code = ? AND club = ?
        """, (code, state.add_employee_club))
        
        if cursor.fetchone():
            conn.close()
            await update.message.reply_text(
                f"❌ Сотрудник {code} уже существует в клубе {state.add_employee_club}"
            )
            return
        
        # Добавляем
        from datetime import datetime
        now = datetime.now().isoformat()
        
        cursor.execute("""
            INSERT INTO employees 
            (code, club, full_name, hired_date, is_active, created_at)
            VALUES (?, ?, ?, ?, 1, ?)
        """, (code, state.add_employee_club, name, hired, now))
        
        conn.commit()
        conn.close()
        
        await update.message.reply_text(
            f"✅ СОТРУДНИК ДОБАВЛЕН\n\n"
            f"🏢 Клуб: {state.add_employee_club}\n"
            f"Код: {code}\n"
            f"Имя: {name}\n"
            f"Дата найма: {hired}"
        )
        
        state.mode = None
        state.add_employee_club = None
        return
    
    # Обработка выбора сотрудника для редактирования
    if state.mode == 'awaiting_employee_edit_select':
        if text_lower == 'отмена':
            await update.message.reply_text("❌ Редактирование отменено")
            state.mode = None
            state.edit_employees_list = None
            state.edit_employees_club = None
            return
        
        # Парсим номер
        try:
            emp_index = int(text)
        except:
            await update.message.reply_text("❌ Введите номер сотрудника")
            return
        
        # Проверка индекса
        if emp_index < 1 or emp_index > len(state.edit_employees_list):
            await update.message.reply_text(
                f"❌ Неверный номер\n"
                f"Доступны номера от 1 до {len(state.edit_employees_list)}"
            )
            return
        
        # Получаем сотрудника
        employee = state.edit_employees_list[emp_index - 1]
        state.edit_employee_selected = employee
        
        # Формируем карточку
        status = "✅ Работает" if employee['is_active'] else "🗂️ Уволен"
        access = "🔐 Есть доступ" if employee['telegram_user_id'] else "❌ Нет доступа"
        
        card_text = (
            f"👤 КАРТОЧКА СОТРУДНИКА\n\n"
            f"🏢 Клуб: {state.edit_employees_club}\n"
            f"💼 Код: {employee['code']}\n"
            f"👤 Имя: {employee['name']}\n"
            f"📱 Телефон: {employee['phone'] or 'не указан'}\n"
            f"🆔 Telegram ID: {employee['telegram_user_id'] or 'не указан'}\n"
            f"📊 Статус: {status}\n"
            f"🔐 Доступ: {access}\n"
        )
        
        # Кнопки редактирования
        buttons = []
        
        if employee['is_active']:
            # Для действующих сотрудников
            buttons.append([InlineKeyboardButton("🔢 Изменить код", callback_data='emp_edit_code')])
            buttons.append([InlineKeyboardButton("✏️ Изменить имя", callback_data='emp_edit_name')])
            buttons.append([InlineKeyboardButton("📱 Изменить телефон", callback_data='emp_edit_phone')])
            
            if employee['telegram_user_id']:
                buttons.append([InlineKeyboardButton("🔐 Изменить TG ID", callback_data='emp_edit_tg')])
                buttons.append([InlineKeyboardButton("🚫 Удалить TG ID (убрать доступ)", callback_data='emp_remove_tg')])
            else:
                buttons.append([InlineKeyboardButton("➕ Добавить TG ID (дать доступ)", callback_data='emp_edit_tg')])
            
            buttons.append([InlineKeyboardButton("🎂 Дата рождения", callback_data='emp_edit_birth')])
            buttons.append([InlineKeyboardButton("🚫 Уволить", callback_data='emp_fire')])
        else:
            # Для уволенных
            buttons.append([InlineKeyboardButton("👀 Просмотр", callback_data='emp_view')])
            buttons.append([InlineKeyboardButton("🔄 Вернуть на работу", callback_data='emp_restore')])
        
        buttons.append([InlineKeyboardButton("❌ Назад", callback_data='emp_edit_cancel')])
        
        await update.message.reply_text(
            card_text,
            reply_markup=InlineKeyboardMarkup(buttons)
        )
        
        state.mode = None  # Ждём callback от кнопок
        return
    
    # Обработка ввода номеров для объединения сотрудников
    if state.mode == 'awaiting_merge_employees':
        await handle_merge_employees_input(update, state, text)
        return
    
    # Обработка подтверждения объединения сотрудников
    if state.mode == 'awaiting_merge_employees_confirm':
        # Обрабатывается через inline кнопки, текст игнорируем
        await update.message.reply_text("Используйте кнопки выше для подтверждения")
        return
    
    # Команда "экспорт"
    if text_lower == 'экспорт':
        await update.message.reply_text(
            "Выберите клуб для экспорта:",
            reply_markup=get_club_report_keyboard()  # Используем ту же клавиатуру
        )
        state.mode = 'awaiting_export_club'
        return
    
    # Обработка выбора клуба для экспорта
    if state.mode == 'awaiting_export_club':
        if text_lower in ['москвич', 'анора', 'оба']:
            state.export_club = text_lower
            await update.message.reply_text(
                "Укажите дату или период:\n"
                "• 12,12\n"
                "• 10,06-11,08"
            )
            state.mode = 'awaiting_export_period'
        else:
            await update.message.reply_text(
                "❌ Неверный выбор. Выберите: москвич, анора или оба"
            )
        return
    
    # Обработка периода для экспорта
    if state.mode == 'awaiting_export_period':
        # Парсим период
        if '-' in text:
            success, date_from, date_to, error = parse_date_range(text)
            if not success:
                await update.message.reply_text(f"❌ {error}")
                return
        else:
            success, single_date, error = parse_short_date(text)
            if not success:
                await update.message.reply_text(f"❌ {error}")
                return
            date_from = single_date
            date_to = single_date
        
        # Экспортируем
        if state.export_club == 'оба':
            for club in ['Москвич', 'Анора']:
                await export_report(update, club, date_from, date_to)
        else:
            club = 'Москвич' if state.export_club == 'москвич' else 'Анора'
            await export_report(update, club, date_from, date_to)
        
        state.mode = None
        state.export_club = None
        return
    
    # Если дошли сюда - либо данные в режиме ввода, либо неизвестная команда
    # В режиме ввода данные уже обработаны выше, поэтому просто игнорируем
    if state.mode in ['нал', 'безнал']:
        return
    
    # Неизвестная команда (только если не в режиме ввода)
    await update.message.reply_text(
        "❓ Команда не распознана\n\n"
        "📋 ОСНОВНЫЕ КОМАНДЫ:\n\n"
        "💰 Ввод данных:\n"
        "  • НАЛ / БЕЗНАЛ → ввод данных\n"
        "  • ГОТОВО → завершить и сохранить\n\n"
        "📊 Просмотр:\n"
        "  • ОТЧЁТ → отчёт по периоду\n"
        "  • ВЫПЛАТЫ → выплаты сотруднику\n"
        "  • СПИСОК → записи за дату\n\n"
        "📤 Другое:\n"
        "  • ЭКСПОРТ → экспорт в Excel\n"
        "  • ИСПРАВИТЬ → редактировать запись\n"
        "  • УДАЛИТЬ → удалить запись\n\n"
        "💡 Используйте кнопки меню ⬇️\n"
        "📖 Полная справка: ПОМОЩЬ"
    )


async def handle_save_command(update: Update, context: ContextTypes.DEFAULT_TYPE, state: UserState):
    """Обработка команды дата/записать"""
    if not state.club:
        await update.message.reply_text(
            "❌ Клуб не выбран.\n"
            "Используйте: старт москвич или старт анора"
        )
        return
    
    text = update.message.text.strip()
    parts = text.split()
    
    # Определяем дату
    target_date = state.current_date
    if len(parts) >= 2:
        potential_date = parts[-1]
        success, parsed_date, error = parse_date(potential_date)
        if success:
            target_date = parsed_date
    
    # Проверяем наличие данных для записи
    ready_nal = context.user_data.get('ready_нал', [])
    ready_beznal = context.user_data.get('ready_безнал', [])
    
    if not ready_nal and not ready_beznal:
        await update.message.reply_text(
            "❌ Нет данных для записи.\n"
            "Используйте: нал / безнал → вставьте данные → готово"
        )
        return
    
    # Записываем в БД
    saved_count = 0
    
    for item in ready_nal:
        db.add_or_update_operation(
            club=state.club,
            date=target_date,
            code=item['code'],
            name=item['name'],
            channel='нал',
            amount=item['amount'],
            original_line=item['original_line'],
            aggregate=True
        )
        saved_count += 1
    
    for item in ready_beznal:
        db.add_or_update_operation(
            club=state.club,
            date=target_date,
            code=item['code'],
            name=item['name'],
            channel='безнал',
            amount=item['amount'],
            original_line=item['original_line'],
            aggregate=True
        )
        saved_count += 1
    
    # Очищаем временные данные
    context.user_data['ready_нал'] = []
    context.user_data['ready_безнал'] = []
    
    await update.message.reply_text(
        f"✅ Сохранено: клуб {state.club}, дата {target_date}\n"
        f"Записей: {saved_count}\n\n"
        f"Можно вводить следующий день или запросить: прошу отчёт"
    )


async def handle_report_command(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                               state: UserState, text: str):
    """Обработка команды прошу отчёт"""
    # Определяем клуб
    club = extract_club_from_text(text)
    if not club:
        club = state.club
    
    if not club:
        await update.message.reply_text(
            "❌ Клуб не указан.\n"
            "Используйте: прошу отчёт москвич | прошу отчёт анора\n"
            "Или сначала выберите клуб: старт москвич"
        )
        return
    
    # Определяем период
    date_from, date_to = None, None
    
    # Поиск явного диапазона
    period_match = re.search(r'(\d{4}-\d{2}-\d{2})\.\.(\d{4}-\d{2}-\d{2})', text)
    if period_match:
        date_from = period_match.group(1)
        date_to = period_match.group(2)
    elif 'неделя' in text or 'недел' in text:
        date_from, date_to = get_week_range()
    else:
        # По умолчанию - текущая неделя
        date_from, date_to = get_week_range()
    
    # Получаем данные
    operations = db.get_operations_by_period(club, date_from, date_to)
    
    if not operations:
        await update.message.reply_text(
            f"📊 Отчет по клубу {club}\n"
            f"Период: {date_from} .. {date_to}\n\n"
            f"Данных нет."
        )
        return
    
    # Загружаем расходы на стилистов для этого периода
    stylist_expenses = db.get_stylist_expenses_for_period(club, date_from, date_to)
    
    # Генерируем отчет
    report_rows, totals, totals_recalc, check_ok = ReportGenerator.calculate_report(
        operations, 
        stylist_expenses=stylist_expenses
    )
    
    report_text = ReportGenerator.format_report_text(
        report_rows, totals, check_ok, totals_recalc, club, f"{date_from} .. {date_to}"
    )
    
    await update.message.reply_text(report_text, parse_mode='Markdown')


async def handle_list_command(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                              state: UserState, text: str):
    """Обработка команды список"""
    if not state.club:
        await update.message.reply_text(
            "❌ Клуб не выбран.\n"
            "Используйте: старт москвич или старт анора"
        )
        return
    
    parts = text.split()
    if len(parts) < 2:
        await update.message.reply_text(
            "❌ Укажите дату.\n"
            "Пример: список 2025-11-03"
        )
        return
    
    date_str = parts[1]
    success, parsed_date, error = parse_date(date_str)
    
    if not success:
        await update.message.reply_text(f"❌ {error}")
        return
    
    operations = db.get_operations_by_date(state.club, parsed_date)
    
    response = format_operations_list(operations, parsed_date, state.club)
    await update.message.reply_text(response)


async def handle_edit_command_new(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                                  state: UserState, text: str):
    """Новая интерактивная команда исправить"""
    if not state.club:
        await update.message.reply_text(
            "❌ Клуб не выбран.\n"
            "Используйте: старт москвич или старт анора"
        )
        return
    
    # Формат: Д1 30,10 (без слова "исправить")
    # Убираем "исправить" если есть (для обратной совместимости)
    parts = text.split()
    if parts[0].lower() in ['исправить', 'ispravit']:
        parts = parts[1:]  # Убираем первое слово
    
    if len(parts) < 2:
        await update.message.reply_text(
            "❌ Неверный формат.\n"
            "Пример: Д1 30,10"
        )
        return
    
    code = DataParser.normalize_code(parts[0])
    date_str = parts[1]
    
    # Парсим дату
    success, parsed_date, error = parse_short_date(date_str)
    if not success:
        await update.message.reply_text(f"❌ {error}")
        return
    
    # Получаем текущие данные
    operations = db.get_operations_by_date(state.club, parsed_date)
    
    # Фильтруем по коду
    code_ops = [op for op in operations if op['code'] == code]
    
    if not code_ops:
        await update.message.reply_text(
            f"❌ Записи для {code} за {parsed_date} не найдены."
        )
        return
    
    # Показываем текущие данные
    response = [f"📊 Текущие данные {code} за {parsed_date}:"]
    current_data = {}
    
    for op in code_ops:
        response.append(f"• {op['channel'].upper()}: {op['amount']:.0f}")
        current_data[op['channel']] = op['amount']
    
    response.append("\nВведите новые значения:")
    response.append("Примеры:")
    response.append("• нал 1100")
    response.append("• безнал 2500")
    response.append("• нал 1100 безнал 2500")
    
    await update.message.reply_text('\n'.join(response))
    
    # Сохраняем состояние
    state.edit_code = code
    state.edit_date = parsed_date
    state.edit_current_data = current_data
    state.mode = 'awaiting_edit_data'


async def handle_edit_input(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                            state: UserState, text: str, text_lower: str):
    """Обработка ввода новых значений для исправления"""
    # Парсим ввод: нал 1100 или безнал 2500 или нал 1100 безнал 2500
    parts = text_lower.split()
    
    # Ищем пары: канал + сумма
    updates = []
    i = 0
    while i < len(parts):
        if parts[i] in ['нал', 'безнал']:
            if i + 1 < len(parts):
                channel = parts[i]
                success, amount, error = DataParser.parse_amount(parts[i + 1])
                
                if success:
                    updates.append((channel, amount))
                    i += 2
                else:
                    await update.message.reply_text(f"❌ {error}")
                    return
            else:
                await update.message.reply_text(f"❌ Не указана сумма для {parts[i]}")
                return
        else:
            await update.message.reply_text(
                f"❌ Неверный формат.\n\n"
                f"Примеры:\n"
                f"• нал 1100\n"
                f"• безнал 2500\n"
                f"• нал 1100 безнал 2500"
            )
            return
    
    if not updates:
        await update.message.reply_text(
            "❌ Не указаны данные для обновления.\n\n"
            "Примеры:\n"
            "• нал 1100\n"
            "• безнал 2500\n"
            "• нал 1100 безнал 2500"
        )
        return
    
    # Сохраняем изменения СРАЗУ
    updated_channels = []
    for channel, amount in updates:
        success, msg = db.update_operation(state.club, state.edit_date, state.edit_code, channel, amount)
        if success:
            updated_channels.append(f"{channel.upper()}: {amount:.0f}")
        else:
            await update.message.reply_text(f"❌ Ошибка обновления {channel}: {msg}")
            return
    
    # Показываем результат
    await update.message.reply_text(
        f"✅ Данные {state.edit_code} за {state.edit_date} обновлены:\n" +
        "\n".join(f"• {ch}" for ch in updated_channels)
    )
    
    # Очищаем состояние
    state.mode = None
    state.edit_code = None
    state.edit_date = None
    state.edit_current_data = None


async def handle_delete_command_new(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                                    state: UserState, text: str):
    """Новая интерактивная команда удалить"""
    if not state.club:
        await update.message.reply_text(
            "❌ Клуб не выбран.\n"
            "Используйте: старт москвич или старт анора"
        )
        return
    
    text_lower = normalize_command(text)
    if text_lower in ['удалить все', 'удалить всё']:
        await update.message.reply_text(
            "🏢 Выберите клуб для удаления данных:",
            reply_markup=get_club_report_keyboard()
        )
        state.mode = 'awaiting_delete_mass_club'
        return
    
    # Формат: удалить Д1 30,10
    parts = text.split()
    if len(parts) < 3:
        await update.message.reply_text(
            "❌ Неверный формат.\n"
            "Пример: удалить Д1 30,10\n"
            "Или напишите: удалить все"
        )
        return
    
    code = DataParser.normalize_code(parts[1])
    date_str = parts[2]
    
    # Парсим дату
    success, parsed_date, error = parse_short_date(date_str)
    if not success:
        await update.message.reply_text(f"❌ {error}")
        return
    
    # Получаем данные
    operations = db.get_operations_by_date(state.club, parsed_date)
    code_ops = [op for op in operations if op['code'] == code]
    
    if not code_ops:
        await update.message.reply_text(
            f"❌ Записи для {code} за {parsed_date} не найдены."
        )
        return
    
    # Показываем записи
    response = [f"📊 Записи {code} за {parsed_date}:"]
    delete_records = {}
    
    for op in code_ops:
        response.append(f"• {op['channel'].upper()}: {op['amount']:.0f}")
        delete_records[op['channel']] = op['amount']
    
    response.append("\nЧто удалить?")
    
    await update.message.reply_text('\n'.join(response), reply_markup=get_delete_keyboard())
    
    # Сохраняем состояние
    state.delete_code = code
    state.delete_date = parsed_date
    state.delete_records = delete_records
    state.mode = 'awaiting_delete_choice'


async def handle_delete_choice(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                               state: UserState, choice: str):
    """Обработка выбора что удалить"""
    if choice in ['нал', 'безнал']:
        # Удаляем один канал
        if choice in state.delete_records:
            db.delete_operation(state.club, state.delete_date, state.delete_code, choice)
            await update.message.reply_text(
                f"✅ Удалено: {state.delete_code} {choice.upper()} за {state.delete_date}"
            )
        else:
            await update.message.reply_text(f"❌ Записи {choice.upper()} нет")
    
    elif choice in ['обе', 'все']:
        # Удаляем оба канала
        deleted = []
        for channel in ['нал', 'безнал']:
            if channel in state.delete_records:
                db.delete_operation(state.club, state.delete_date, state.delete_code, channel)
                deleted.append(channel.upper())
        
        if deleted:
            await update.message.reply_text(
                f"✅ Удалено: {state.delete_code} {', '.join(deleted)} за {state.delete_date}"
            )
        else:
            await update.message.reply_text("❌ Нет записей для удаления")
    
    else:
        await update.message.reply_text(
            "❌ Неверный выбор. Выберите: нал, безнал или обе"
        )
        return
    
    # Очищаем состояние
    state.mode = None


async def handle_delete_employee_input(update: Update, context: ContextTypes.DEFAULT_TYPE,
                                       state: UserState, text: str):
    """Обработка ввода для удаления конкретного сотрудника"""
    cleaned = text.strip()
    if not cleaned:
        await update.message.reply_text(
            "❌ Введите код и дату.\n"
            "Пример: Д1 30,10"
        )
        return
    
    # Используем существующий обработчик, добавляя ключевое слово
    await handle_delete_command_new(update, context, state, f"удалить {cleaned}")


def _summarize_operations_for_delete(operations: list) -> Dict:
    """Возвращает агрегаты по списку операций"""
    total_nal = sum(op['amount'] for op in operations if op['channel'] == 'нал')
    total_beznal = sum(op['amount'] for op in operations if op['channel'] == 'безнал')
    return {
        'count': len(operations),
        'total_nal': total_nal,
        'total_beznal': total_beznal
    }


def _format_delete_preview_lines(club_label: str, date_from: str, date_to: str,
                                 operations: list) -> Tuple[str, Dict]:
    """Формирует текст предпросмотра удаления и возвращает агрегаты"""
    summary = _summarize_operations_for_delete(operations)
    lines = []
    lines.append(f"🏢 {club_label}")
    lines.append(f"📅 Период: {date_from} .. {date_to}")
    lines.append(f"🧾 Записей: {summary['count']}")
    lines.append(f"💵 НАЛ: {summary['total_nal']:,.0f}".replace(',', ' '))
    lines.append(f"💳 БЕЗНАЛ: {summary['total_beznal']:,.0f}".replace(',', ' '))
    
    if operations:
        lines.append("Первые записи:")
        for op in operations[:5]:
            code = op['code']
            name = op['name'] or "(без имени)"
            channel = op['channel'].upper()
            amount = f"{op['amount']:,.0f}".replace(',', ' ')
            lines.append(f" • {op['date']} | {code} {name} | {channel} {amount}")
        if len(operations) > 5:
            lines.append(f" • ... и ещё {len(operations) - 5} записей")
    else:
        lines.append("Нет записей за выбранный период.")
    
    return '\n'.join(lines), summary


def create_delete_preview_excel(preview_data: List[Dict], filename: str):
    """Создаёт Excel-файл с данными для удаления"""
    wb = Workbook()
    ws = wb.active
    ws.title = "К удалению"
    ws.append(["Клуб", "Дата", "Код", "Имя", "Канал", "Сумма"])
    
    for item in preview_data:
        club = item['club']
        for op in item.get('operations', []):
            ws.append([
                club,
                op['date'],
                op['code'],
                op['name'] or "",
                op['channel'],
                op['amount']
            ])
    
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    wb.save(filename)


async def handle_delete_mass_club_input(update: Update, state: UserState,
                                        text: str, text_lower: str):
    """Выбор клуба для массового удаления"""
    club_map = {
        'москвич': 'москвич',
        'анора': 'анора',
        'оба': 'оба'
    }
    
    normalized = text_lower
    if normalized in ['🏢 москвич', 'москвич']:
        selection = 'москвич'
    elif normalized in ['🏢 анора', 'анора', 'anora']:
        selection = 'анора'
    elif normalized in ['🏢🏢 оба', 'оба']:
        selection = 'оба'
    else:
        await update.message.reply_text(
            "❌ Неверный выбор. Напишите: москвич, анора или оба"
        )
        return
    
    state.delete_mass_club = selection
    state.delete_mass_date_from = None
    state.delete_mass_date_to = None
    state.delete_mass_preview = None
    await update.message.reply_text(
        "📅 Укажите дату или период для удаления:\n\n"
        "Примеры:\n"
        "• 5,11\n"
        "• 2,11-5,11"
    )
    state.mode = 'awaiting_delete_mass_period'


async def handle_delete_mass_period_input(update: Update, state: UserState,
                                          text: str, text_lower: str):
    """Обработка ввода даты/периода для массового удаления"""
    if '-' in text:
        success, date_from, date_to, error = parse_date_range(text)
        if not success:
            await update.message.reply_text(f"❌ {error}")
            return
    else:
        success, single_date, error = parse_short_date(text)
        if not success:
            await update.message.reply_text(f"❌ {error}")
            return
        date_from = single_date
        date_to = single_date
    
    selection = state.delete_mass_club
    club_labels = []
    if selection == 'оба':
        club_labels = [('Москвич', 'Москвич'), ('Анора', 'Анора')]
    else:
        label = 'Москвич' if selection == 'москвич' else 'Анора'
        club_labels = [(label, label)]
    
    preview_sections = []
    preview_data = []
    total_records = 0
    
    for club_key, club_label in club_labels:
        operations = db.get_operations_by_period(club_label, date_from, date_to)
        if operations:
            section_text, summary = _format_delete_preview_lines(club_label, date_from, date_to, operations)
            preview_sections.append(section_text)
            preview_data.append({
                'club': club_label,
                'summary': summary,
                'operations': operations
            })
            total_records += summary['count']
        else:
            preview_sections.append(
                f"🏢 {club_label}\n"
                f"📅 Период: {date_from} .. {date_to}\n"
                "Нет записей за выбранный период."
            )
    
    if total_records == 0:
        await update.message.reply_text(
            "ℹ️ За указанный период данных нет.\n"
            "Удаление не требуется."
        )
        # Сброс
        state.mode = None
        state.delete_mass_club = None
        return
    
    # Сохраняем параметры
    state.delete_mass_date_from = date_from
    state.delete_mass_date_to = date_to
    state.delete_mass_preview = {
        'clubs': preview_data,
        'total_records': total_records
    }
    
    await update.message.reply_text(
        "📊 Предпросмотр удаления:\n\n" + '\n\n'.join(preview_sections)
    )
    
    # Отправляем Excel с деталями
    filename = f"delete_preview_{uuid.uuid4().hex}.xlsx"
    create_delete_preview_excel(preview_data, filename)
    with open(filename, 'rb') as f:
        await update.message.reply_document(
            document=f,
            filename=f"preview_delete_{date_from}_{date_to}.xlsx",
            caption="📄 Excel с данными для удаления"
        )
    os.remove(filename)
    
    await update.message.reply_text(
        "❗ Подтвердите удаление всех записей за этот период.",
        reply_markup=get_delete_mass_confirm_keyboard()
    )
    state.mode = 'awaiting_delete_mass_confirm'


async def handle_delete_mass_confirm_message(message, state: UserState, confirmed: bool):
    """Подтверждение массового удаления (message может быть update.message или query.message)"""
    if confirmed:
        selection = state.delete_mass_club
        date_from = state.delete_mass_date_from
        date_to = state.delete_mass_date_to
        preview = state.delete_mass_preview or {}
        
        results = []
        total_deleted = 0
        
        clubs_to_process = []
        if selection == 'оба':
            clubs_to_process = ['Москвич', 'Анора']
        else:
            clubs_to_process = ['Москвич' if selection == 'москвич' else 'Анора']
        
        for club in clubs_to_process:
            deleted = db.delete_operations_by_period(club, date_from, date_to)
            total_deleted += deleted
            summary = None
            if preview:
                for item in preview.get('clubs', []):
                    if item['club'] == club:
                        summary = item['summary']
                        break
            results.append({
                'club': club,
                'deleted': deleted,
                'summary': summary
            })
        
        lines = []
        if total_deleted == 0:
            lines.append("ℹ️ Записей для удаления не найдено.")
        else:
            lines.append("🗑️ Удаление завершено!")
            lines.append(f"📅 Период: {date_from} .. {date_to}")
            lines.append(f"🧾 Удалено записей: {total_deleted}")
            lines.append("")
            for item in results:
                summary = item['summary']
                lines.append(f"🏢 {item['club']}")
                lines.append(f"Удалено записей: {item['deleted']}")
                if summary:
                    lines.append(f"НАЛ: {summary['total_nal']:,.0f}".replace(',', ' '))
                    lines.append(f"БЕЗНАЛ: {summary['total_beznal']:,.0f}".replace(',', ' '))
                lines.append("")
            lines.append("📜 История доступна в ЖУРНАЛ.")
        
        await message.reply_text('\n'.join(line for line in lines if line))
        
        # Сброс
        state.mode = None
        state.delete_mass_club = None
        state.delete_mass_date_from = None
        state.delete_mass_date_to = None
        state.delete_mass_preview = None
        return
    
    await message.reply_text("✅ Удаление отменено.")
    state.mode = None
    state.delete_mass_club = None
    state.delete_mass_date_from = None
    state.delete_mass_date_to = None
    state.delete_mass_preview = None


async def handle_delete_mass_confirm_text(update: Update, state: UserState, text_lower: str):
    """Фолбэк на текстовое подтверждение"""
    if text_lower in ['да', 'ok', 'ок', 'yes', 'y']:
        await handle_delete_mass_confirm_message(update.message, state, True)
    elif text_lower in ['нет', 'no', 'n', 'отмена', 'cancel']:
        await handle_delete_mass_confirm_message(update.message, state, False)
    else:
        await update.message.reply_text(
            "❓ Не понял. Напишите: да / нет\n"
            "Для отмены напишите: отмена"
        )
        return

        state.mode = None
        state.delete_mass_club = None
        state.delete_mass_date_from = None
        state.delete_mass_date_to = None
        state.delete_mass_preview = None
        return
    
    await update.message.reply_text(
        "❓ Не понял. Напишите: да / нет\n"
        "Для отмены напишите: отмена"
    )


async def export_report(update: Update, club: str, date_from: str, date_to: str):
    """Экспорт отчёта в XLSX"""
    operations = db.get_operations_by_period(club, date_from, date_to)
    
    if not operations:
        await update.message.reply_text(
            f"📊 Нет данных для экспорта\n"
            f"Клуб: {club}\n"
            f"Период: {date_from} .. {date_to}"
        )
        return
    
    # Загружаем расходы на стилистов для этого периода
    stylist_expenses = db.get_stylist_expenses_for_period(club, date_from, date_to)
    
    # Генерируем отчет
    report_rows, totals, totals_recalc, check_ok = ReportGenerator.calculate_report(
        operations, 
        stylist_expenses=stylist_expenses
    )
    
    # Создаем XLSX
    club_translit = 'moskvich' if club == 'Москвич' else 'anora'
    filename = f"otchet_{club_translit}_{date_from}_{date_to}.xlsx"
    
    ReportGenerator.generate_xlsx(
        report_rows, totals, club, f"{date_from} .. {date_to}", filename, db
    )
    
    # Отправляем файл
    with open(filename, 'rb') as f:
        await update.message.reply_document(
            document=f,
            filename=filename,
            caption=f"📊 Экспорт: {club}\nПериод: {date_from} .. {date_to}"
        )
    
    # Удаляем временный файл
    os.remove(filename)


async def prepare_merged_report(update: Update, state: UserState, date_from: str, date_to: str):
    """Подготовка сводного отчета с проверкой совпадений"""
    # Получаем данные по обоим клубам
    ops_moskvich = db.get_operations_by_period('Москвич', date_from, date_to)
    ops_anora = db.get_operations_by_period('Анора', date_from, date_to)
    
    # Группируем по сотрудникам (код)
    from collections import defaultdict
    
    employees_m = defaultdict(lambda: {'names': set(), 'nal': 0, 'beznal': 0})
    employees_a = defaultdict(lambda: {'names': set(), 'nal': 0, 'beznal': 0})
    
    for op in ops_moskvich:
        code = op['code']
        employees_m[code]['names'].add(op['name'])
        if op['channel'] == 'нал':
            employees_m[code]['nal'] += op['amount']
        else:
            employees_m[code]['beznal'] += op['amount']
    
    for op in ops_anora:
        code = op['code']
        employees_a[code]['names'].add(op['name'])
        if op['channel'] == 'нал':
            employees_a[code]['nal'] += op['amount']
        else:
            employees_a[code]['beznal'] += op['amount']
    
    # Ищем совпадения по КОД+ИМЯ
    merge_candidates = []
    all_codes = set(employees_m.keys()) | set(employees_a.keys())
    
    for code in all_codes:
        if code in employees_m and code in employees_a:
            # Код есть в обоих клубах
            names_m = employees_m[code]['names']
            names_a = employees_a[code]['names']
            
            # Проверяем совпадение имён
            common_names = names_m & names_a
            
            if common_names:
                # Есть полное совпадение КОД+ИМЯ
                name = list(common_names)[0]
                merge_candidates.append({
                    'code': code,
                    'name': name,
                    'moskvich': {'nal': employees_m[code]['nal'], 'beznal': employees_m[code]['beznal']},
                    'anora': {'nal': employees_a[code]['nal'], 'beznal': employees_a[code]['beznal']},
                    'names_m': list(names_m),
                    'names_a': list(names_a)
                })
    
    # Проверяем СБ с похожими именами между клубами
    sb_cross_club_matches = []
    
    # Получаем всех СБ из обоих клубов
    sb_moskvich = [op for op in ops_moskvich if op['code'] == 'СБ' and op.get('name')]
    sb_anora = [op for op in ops_anora if op['code'] == 'СБ' and op.get('name')]
    
    if sb_moskvich and sb_anora:
        # Группируем по именам С ПРИМЕНЕНИЕМ ОБЪЕДИНЕНИЙ ВНУТРИ КЛУБА
        from collections import defaultdict
        sb_names_m = defaultdict(lambda: {'nal': 0, 'beznal': 0})
        sb_names_a = defaultdict(lambda: {'nal': 0, 'beznal': 0})
        
        # Получаем словари объединений СБ из state (если есть)
        sb_merges_m = getattr(state, 'sb_merges_moskvich', {}) or {}
        sb_merges_a = getattr(state, 'sb_merges_anora', {}) or {}
        
        for op in sb_moskvich:
            name = op['name'].strip()
            # Применяем объединение Москвича
            if name in sb_merges_m:
                name = sb_merges_m[name]
            
            if op['channel'] == 'нал':
                sb_names_m[name]['nal'] += op['amount']
            else:
                sb_names_m[name]['beznal'] += op['amount']
        
        for op in sb_anora:
            name = op['name'].strip()
            # Применяем объединение Аноры
            if name in sb_merges_a:
                name = sb_merges_a[name]
            
            if op['channel'] == 'нал':
                sb_names_a[name]['nal'] += op['amount']
            else:
                sb_names_a[name]['beznal'] += op['amount']
        
        # Ищем похожие имена СБ между клубами
        for name_m in sb_names_m.keys():
            surnames_m = extract_surname_candidates(name_m)
            
            for name_a in sb_names_a.keys():
                surnames_a = extract_surname_candidates(name_a)
                
                # Проверяем фамилии
                has_matching_surname = False
                for s_m in surnames_m:
                    for s_a in surnames_a:
                        surname_similarity = SequenceMatcher(None, s_m, s_a).ratio()
                        if surname_similarity >= 0.90:
                            has_matching_surname = True
                            break
                    if has_matching_surname:
                        break
                
                if has_matching_surname:
                    # Нашли СБ с похожими именами в разных клубах
                    sb_cross_club_matches.append({
                        'name_moskvich': name_m,
                        'name_anora': name_a,
                        'similarity': surname_similarity,
                        'moskvich': sb_names_m[name_m],
                        'anora': sb_names_a[name_a]
                    })
    
    # Если нет совпадений по коду И нет СБ между клубами - генерируем простой сводный
    if not merge_candidates and not sb_cross_club_matches:
        # Совпадений нет - генерируем сводный без объединения (просто все записи)
        await update.message.reply_text(
            "ℹ️ Совпадений не найдено\n"
            "Генерируется сводный отчёт из всех записей..."
        )
        
        # Создаём сводный из всех операций
        all_ops = ops_moskvich + ops_anora
        
        if all_ops:
            # Загружаем расходы на стилистов для обоих клубов
            stylist_expenses_m = db.get_stylist_expenses_for_period('Москвич', date_from, date_to)
            stylist_expenses_a = db.get_stylist_expenses_for_period('Анора', date_from, date_to)
            stylist_expenses_merged = stylist_expenses_m + stylist_expenses_a
            
            report_rows, totals, totals_recalc, check_ok = ReportGenerator.calculate_report(
                all_ops,
                stylist_expenses=stylist_expenses_merged
            )
            
            # Определяем куда отправлять
            msg = update.message if update.message else (update.callback_query.message if update.callback_query else None)
            
            # Краткая сводка
            summary = format_report_summary(
                totals, 
                "СВОДНЫЙ (Москвич + Анора)", 
                f"{date_from} .. {date_to}",
                len(report_rows)
            )
            await msg.reply_text(summary)
            
            # Экспорт
            filename = f"otchet_svodny_{date_from}_{date_to}.xlsx"
            ReportGenerator.generate_xlsx(
                report_rows, totals, "СВОДНЫЙ (Москвич + Анора)", f"{date_from} .. {date_to}", filename, db
            )
            with open(filename, 'rb') as f:
                await msg.reply_document(
                    document=f, filename=filename,
                    caption=f"📊 СВОДНЫЙ ОТЧЁТ (Оба клуба)\nПериод: {date_from} .. {date_to}"
                )
            os.remove(filename)
        
        state.mode = None
        state.report_club = None
        return
    
    # Создаем текстовый файл со списком совпадений
    file_content = ["📋 НАЙДЕНЫ СОВПАДЕНИЯ ДЛЯ ОБЪЕДИНЕНИЯ\n"]
    file_content.append(f"Период: {date_from} .. {date_to}\n")
    file_content.append("=" * 50 + "\n\n")
    
    # Совпадения по коду+имени
    if merge_candidates:
        file_content.append("🔸 СОВПАДЕНИЯ ПО КОДУ И ИМЕНИ:\n\n")
        for i, candidate in enumerate(merge_candidates, 1):
            file_content.append(f"{i}. {candidate['name']} {candidate['code']}\n")
            file_content.append(f"   • Москвич: НАЛ {candidate['moskvich']['nal']:.0f}, БЕЗНАЛ {candidate['moskvich']['beznal']:.0f}\n")
            file_content.append(f"   • Анора: НАЛ {candidate['anora']['nal']:.0f}, БЕЗНАЛ {candidate['anora']['beznal']:.0f}\n")
            file_content.append("\n")
    
    # Совпадения СБ между клубами
    if sb_cross_club_matches:
        file_content.append("\n🔸 СБ С ПОХОЖИМИ ИМЕНАМИ (разные клубы):\n\n")
        start_idx = len(merge_candidates) + 1
        for i, match in enumerate(sb_cross_club_matches, start_idx):
            similarity_pct = int(match['similarity'] * 100)
            file_content.append(f"{i}. СБ (Похожесть фамилий: {similarity_pct}%)\n")
            file_content.append(f"   • Москвич: {match['name_moskvich']} - НАЛ {match['moskvich']['nal']:.0f}, БЕЗНАЛ {match['moskvich']['beznal']:.0f}\n")
            file_content.append(f"   • Анора: {match['name_anora']} - НАЛ {match['anora']['nal']:.0f}, БЕЗНАЛ {match['anora']['beznal']:.0f}\n")
            file_content.append("\n")
    
    file_content.append("=" * 50 + "\n")
    file_content.append("\n🔄 ОБЪЕДИНЕНИЕ ДЛЯ СВОДНОГО ОТЧЁТА:\n")
    file_content.append("• ОК → объединить все\n")
    file_content.append("• ОК 1 → объединить только пункт 1\n")
    file_content.append("• ОК 1 2 → объединить пункты 1 и 2\n")
    file_content.append("• НЕ 1 → НЕ объединять пункт 1 (остальные да)\n")
    file_content.append("• НЕ 1 2 → НЕ объединять пункты 1 и 2\n")
    file_content.append("\nℹ️ Примечание: объединение ТОЛЬКО для отчёта\n")
    file_content.append("(данные в БД не изменяются)\n")
    
    # Сохраняем во временный файл
    temp_file = tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', suffix='.txt', delete=False)
    temp_file.write(''.join(file_content))
    temp_file.close()
    
    # Отправляем короткое сообщение с кнопками
    total_count = len(merge_candidates) + len(sb_cross_club_matches)
    short_message = (
        f"📋 Найдено совпадений: {total_count}\n"
        f"   • По коду+имени: {len(merge_candidates)}\n"
        f"   • СБ между клубами: {len(sb_cross_club_matches)}\n\n"
        f"🔄 Объединение для сводного отчёта:\n"
        f"• Используйте кнопки ниже\n"
        f"• Или введите: ОК / ОК 1 / НЕ 1\n\n"
        f"📄 Детальный список в файле ⬇️\n\n"
        f"ℹ️ Объединение ТОЛЬКО для отчёта\n"
        f"(данные в БД не изменяются)"
    )
    
    # Отправляем файл и сообщение с кнопками
    # Определяем куда отправлять (может быть callback query или обычное сообщение)
    msg = update.message if update.message else (update.callback_query.message if update.callback_query else None)
    
    with open(temp_file.name, 'rb') as f:
        await msg.reply_document(
            document=f,
            filename=f"sovpadeniya_{date_from}_{date_to}.txt",
            caption=short_message,
            reply_markup=get_merge_confirmation_keyboard()
        )
    
    # Удаляем временный файл
    os.remove(temp_file.name)
    
    # Сохраняем кандидатов (включая СБ)
    state.merge_candidates = merge_candidates
    state.sb_cross_club_matches = sb_cross_club_matches  # Новое поле для СБ
    state.merge_period = (date_from, date_to)
    state.mode = 'awaiting_merge_confirm'


async def handle_merge_confirmation(update: Update, state: UserState, choice: str, message=None):
    """Обработка подтверждения объединения для сводного отчёта"""
    # Используем message если передан, иначе update.message
    msg = message if message else update.message
    
    # Обработка ответа с новой логикой
    indices_to_merge = set()
    
    # Убираем знаки препинания для удобства парсинга: "не1,2" -> "не 1 2"
    normalized_text = choice.replace(',', ' ').replace('.', ' ')
    parts = normalized_text.split()
    
    if not parts:
        await msg.reply_text("❌ Неверный формат. Используйте: ок, ок 1, ок 1 2, не 1, не 1 2")
        return
    
    command = parts[0]
    
    # Общее количество совпадений (обычные + СБ)
    sb_matches = getattr(state, 'sb_cross_club_matches', [])
    total_candidates = len(state.merge_candidates) + len(sb_matches)
    
    if command in ['ок', 'ok']:
        # "ок" без номеров -> объединить ВСЕ
        if len(parts) == 1:
            indices_to_merge = set(range(total_candidates))
        else:
            # "ок 1 2" -> объединить ТОЛЬКО указанные
            try:
                indices_to_merge = set(int(x) - 1 for x in parts[1:] if x.isdigit())
            except:
                await msg.reply_text("❌ Неверный формат номеров. Используйте: ок 1 2")
                return
    elif command in ['не', 'net', 'нет']:
        # "не 1 2" -> НЕ объединять указанные (объединить остальные)
        try:
            exclude_indices = set(int(x) - 1 for x in parts[1:] if x.isdigit())
            indices_to_merge = set(range(total_candidates)) - exclude_indices
        except:
            await msg.reply_text("❌ Неверный формат номеров. Используйте: не 1 2")
            return
    else:
        await msg.reply_text(
            "❌ Неверная команда.\n\n"
            "Используйте:\n"
            "• ок - объединить все\n"
            "• ок 1 - объединить только пункт 1\n"
            "• ок 1 2 - объединить пункты 1 и 2\n"
            "• не 1 - НЕ объединять пункт 1 (остальные объединить)\n"
            "• не 1 2 - НЕ объединять пункты 1 и 2"
        )
        return
    
    # Преобразуем в формат excluded (для совместимости с generate_merged_report)
    excluded = set(range(total_candidates)) - indices_to_merge
    
    # Разделяем excluded на обычные и СБ
    excluded_regular = excluded & set(range(len(state.merge_candidates)))
    excluded_sb = excluded - excluded_regular
    
    # Уведомление о начале генерации
    merged_count = len(indices_to_merge)
    await msg.reply_text(
        f"⏳ Генерация сводного отчёта...\n"
        f"Объединяется: {merged_count} из {total_candidates}"
    )
    
    # Генерируем сводный отчет (используем message если передан)
    await generate_merged_report(update, state, excluded_regular, excluded_sb, message)
    
    # Очищаем
    state.mode = None
    state.report_club = None
    state.merge_candidates = None
    state.sb_cross_club_matches = None
    state.merge_period = None


async def generate_merged_report(update: Update, state: UserState, excluded_regular: set, excluded_sb: set, message=None):
    """
    Генерация сводного отчета из ОБОИХ клубов
    excluded_regular: индексы обычных совпадений которые НЕ объединяем
    excluded_sb: индексы СБ совпадений которые НЕ объединяем
    """
    # Используем message если передан, иначе update.message
    msg = message if message else update.message
    
    try:
        date_from, date_to = state.merge_period
        
        # Получаем ВСЕ данные обоих клубов
        ops_m = db.get_operations_by_period('Москвич', date_from, date_to)
        ops_a = db.get_operations_by_period('Анора', date_from, date_to)
    except Exception as e:
        await msg.reply_text(f"❌ Ошибка получения данных: {str(e)}")
        return
    
    # Создаём объединённый список операций для СВОДНОГО отчёта
    merged_ops = []
    
    # Множество обработанных пар (код, имя)
    processed = set()
    
    # 1. Добавляем ОБЪЕДИНЁННЫЕ записи по коду+имени (которые пользователь подтвердил)
    for i, candidate in enumerate(state.merge_candidates):
        code = candidate['code']
        name = candidate['name']
        names_m = candidate.get('names_m', [])
        names_a = candidate.get('names_a', [])
        name_variants = set(names_m + names_a)
        name_variants.add(name)
        
        if i not in excluded_regular:
            # ОБЪЕДИНЯЕМ - суммируем из обоих клубов
            total_nal = candidate['moskvich']['nal'] + candidate['anora']['nal']
            total_beznal = candidate['moskvich']['beznal'] + candidate['anora']['beznal']
            
            if total_nal > 0:
                merged_ops.append({
                    'code': code, 'name': name, 'channel': 'нал', 
                    'amount': total_nal, 'date': date_from
                })
            if total_beznal > 0:
                merged_ops.append({
                    'code': code, 'name': name, 'channel': 'безнал', 
                    'amount': total_beznal, 'date': date_from
                })
            
            for variant in name_variants:
                processed.add(make_processed_key(code, variant))
        else:
            # НЕ объединяем - добавляем раздельно с пометкой клуба
            if candidate['moskvich']['nal'] > 0:
                merged_ops.append({
                    'code': code, 'name': f"{name} (Москвич)", 'channel': 'нал',
                    'amount': candidate['moskvich']['nal'], 'date': date_from
                })
            if candidate['moskvich']['beznal'] > 0:
                merged_ops.append({
                    'code': code, 'name': f"{name} (Москвич)", 'channel': 'безнал',
                    'amount': candidate['moskvich']['beznal'], 'date': date_from
                })
            if candidate['anora']['nal'] > 0:
                merged_ops.append({
                    'code': code, 'name': f"{name} (Анора)", 'channel': 'нал',
                    'amount': candidate['anora']['nal'], 'date': date_from
                })
            if candidate['anora']['beznal'] > 0:
                merged_ops.append({
                    'code': code, 'name': f"{name} (Анора)", 'channel': 'безнал',
                    'amount': candidate['anora']['beznal'], 'date': date_from
                })
            
            for variant in name_variants:
                processed.add(make_processed_key(code, variant))
    
    # 1.5. Добавляем ОБЪЕДИНЁННЫЕ СБ между клубами
    sb_matches = getattr(state, 'sb_cross_club_matches', [])
    for i, match in enumerate(sb_matches):
        sb_idx = len(state.merge_candidates) + i  # Индекс в общем списке
        name_m = match['name_moskvich']
        name_a = match['name_anora']
        
        if sb_idx not in excluded_sb:
            # ОБЪЕДИНЯЕМ СБ - берем более полное имя
            united_name = max(name_m, name_a, key=len)
            total_nal = match['moskvich']['nal'] + match['anora']['nal']
            total_beznal = match['moskvich']['beznal'] + match['anora']['beznal']
            
            if total_nal > 0:
                merged_ops.append({
                    'code': 'СБ', 'name': united_name, 'channel': 'нал',
                    'amount': total_nal, 'date': date_from
                })
            if total_beznal > 0:
                merged_ops.append({
                    'code': 'СБ', 'name': united_name, 'channel': 'безнал',
                    'amount': total_beznal, 'date': date_from
                })
            
            processed.add(make_processed_key('СБ', name_m))
            processed.add(make_processed_key('СБ', name_a))
        else:
            # НЕ объединяем - добавляем раздельно
            if match['moskvich']['nal'] > 0:
                merged_ops.append({
                    'code': 'СБ', 'name': f"{name_m} (Москвич)", 'channel': 'нал',
                    'amount': match['moskvich']['nal'], 'date': date_from
                })
            if match['moskvich']['beznal'] > 0:
                merged_ops.append({
                    'code': 'СБ', 'name': f"{name_m} (Москвич)", 'channel': 'безнал',
                    'amount': match['moskvich']['beznal'], 'date': date_from
                })
            if match['anora']['nal'] > 0:
                merged_ops.append({
                    'code': 'СБ', 'name': f"{name_a} (Анора)", 'channel': 'нал',
                    'amount': match['anora']['nal'], 'date': date_from
                })
            if match['anora']['beznal'] > 0:
                merged_ops.append({
                    'code': 'СБ', 'name': f"{name_a} (Анора)", 'channel': 'безнал',
                    'amount': match['anora']['beznal'], 'date': date_from
                })
            
            processed.add(make_processed_key('СБ', name_m))
            processed.add(make_processed_key('СБ', name_a))
    
    # Объединяем словари СБ из обоих клубов
    combined_sb_merges = {}
    if hasattr(state, 'sb_merges_moskvich') and state.sb_merges_moskvich:
        combined_sb_merges.update(state.sb_merges_moskvich)
    if hasattr(state, 'sb_merges_anora') and state.sb_merges_anora:
        combined_sb_merges.update(state.sb_merges_anora)
    
    # 2. Добавляем СБ из каждого клуба с применением объединений ВНУТРИ клуба
    # Сначала группируем СБ по клубам и применяем их внутренние объединения
    from collections import defaultdict
    
    # Группируем СБ из Москвича
    sb_moskvich_grouped = defaultdict(lambda: {'nal': 0, 'beznal': 0})
    for op in ops_m:
        if op['code'] == 'СБ':
            name = op['name']
            # Применяем объединение Москвича
            if state.sb_merges_moskvich and name in state.sb_merges_moskvich:
                name = state.sb_merges_moskvich[name]
            
            if op['channel'] == 'нал':
                sb_moskvich_grouped[name]['nal'] += op['amount']
            else:
                sb_moskvich_grouped[name]['beznal'] += op['amount']
    
    # Группируем СБ из Аноры
    sb_anora_grouped = defaultdict(lambda: {'nal': 0, 'beznal': 0})
    for op in ops_a:
        if op['code'] == 'СБ':
            name = op['name']
            # Применяем объединение Аноры
            if state.sb_merges_anora and name in state.sb_merges_anora:
                name = state.sb_merges_anora[name]
            
            if op['channel'] == 'нал':
                sb_anora_grouped[name]['nal'] += op['amount']
            else:
                sb_anora_grouped[name]['beznal'] += op['amount']
    
    # Теперь добавляем СБ из обоих клубов
    # Сначала те что УЖЕ в sb_cross_club_matches (они уже добавлены выше)
    # Остальные добавляем отдельно
    for name, amounts in sb_moskvich_grouped.items():
        # Проверяем: не был ли этот СБ уже добавлен как часть sb_cross_club_matches
        if make_processed_key('СБ', name) not in processed:
            if amounts['nal'] > 0:
                merged_ops.append({'code': 'СБ', 'name': name, 'channel': 'нал', 'amount': amounts['nal'], 'date': date_from})
            if amounts['beznal'] > 0:
                merged_ops.append({'code': 'СБ', 'name': name, 'channel': 'безнал', 'amount': amounts['beznal'], 'date': date_from})
            processed.add(make_processed_key('СБ', name))
    
    for name, amounts in sb_anora_grouped.items():
        if make_processed_key('СБ', name) not in processed:
            if amounts['nal'] > 0:
                merged_ops.append({'code': 'СБ', 'name': name, 'channel': 'нал', 'amount': amounts['nal'], 'date': date_from})
            if amounts['beznal'] > 0:
                merged_ops.append({'code': 'СБ', 'name': name, 'channel': 'безнал', 'amount': amounts['beznal'], 'date': date_from})
            processed.add(make_processed_key('СБ', name))
    
    # 3. Добавляем ВСЕ ОСТАЛЬНЫЕ записи (НЕ СБ)
    for op in ops_m + ops_a:
        if op['code'] != 'СБ' and make_processed_key(op['code'], op['name']) not in processed:
            merged_ops.append(op)
    
    # Генерируем СВОДНЫЙ отчет
    # ДЛЯ СВОДНОГО НЕ передаём sb_name_merges, т.к. уже применили выше!
    if merged_ops:
        try:
            # Загружаем расходы на стилистов для каждого клуба
            stylist_expenses_m = db.get_stylist_expenses_for_period('Москвич', date_from, date_to)
            stylist_expenses_a = db.get_stylist_expenses_for_period('Анора', date_from, date_to)
            
            # Объединяем расходы на стилистов для сводного отчета
            stylist_expenses_merged = stylist_expenses_m + stylist_expenses_a
            
            # Генерируем сводный отчет
            report_rows_merged, totals_merged, totals_recalc, check_ok = ReportGenerator.calculate_report(
                merged_ops,
                sb_name_merges=None,  # УЖЕ применили объединения!
                stylist_expenses=stylist_expenses_merged
            )
            
            # Генерируем отчеты для каждого клуба отдельно
            report_rows_m, totals_m, _, _ = ReportGenerator.calculate_report(
                ops_m,
                sb_name_merges=state.sb_merges_moskvich if hasattr(state, 'sb_merges_moskvich') else None,
                stylist_expenses=stylist_expenses_m
            )
            report_rows_a, totals_a, _, _ = ReportGenerator.calculate_report(
                ops_a,
                sb_name_merges=state.sb_merges_anora if hasattr(state, 'sb_merges_anora') else None,
                stylist_expenses=stylist_expenses_a
            )
            
            # НОВАЯ ЛОГИКА: формируем сводный отчет складывая готовые отчеты
            from collections import defaultdict
            
            # Индексируем строки по (код, имя)
            merged_dict = defaultdict(lambda: {
                'name': '', 'code': '', 'nal': 0, 'beznal': 0, 'minus10': 0, 'stylist': 0, 'itog': 0
            })
            
            # Добавляем строки из Москвича
            for row in report_rows_m:
                key = (row['code'], row['name'])
                merged_dict[key]['name'] = row['name']
                merged_dict[key]['code'] = row['code']
                merged_dict[key]['nal'] += row['nal']
                merged_dict[key]['beznal'] += row['beznal']
                merged_dict[key]['stylist'] += row['stylist']
            
            # Добавляем строки из Аноры
            for row in report_rows_a:
                key = (row['code'], row['name'])
                merged_dict[key]['name'] = row['name']
                merged_dict[key]['code'] = row['code']
                merged_dict[key]['nal'] += row['nal']
                merged_dict[key]['beznal'] += row['beznal']
                merged_dict[key]['stylist'] += row['stylist']
            
            # Пересчитываем 10% и итого для каждой строки
            report_rows_merged = []
            for key, data in sorted(merged_dict.items()):
                data['minus10'] = round(data['beznal'] * 0.10, 2)
                data['itog'] = round(data['nal'] + (data['beznal'] - data['minus10']) - data['stylist'], 2)
                report_rows_merged.append(data)
            
            # Пересчитываем итоги
            totals_merged = {
                'nal': sum(row['nal'] for row in report_rows_merged),
                'beznal': sum(row['beznal'] for row in report_rows_merged),
                'minus10': sum(row['minus10'] for row in report_rows_merged),
                'stylist': sum(row['stylist'] for row in report_rows_merged),
                'itog': sum(row['itog'] for row in report_rows_merged)
            }
            
            # Краткая сводка вместо полного отчёта
            merged_regular = len(state.merge_candidates) - len(excluded_regular) if state.merge_candidates else 0
            merged_sb = len(sb_matches) - len(excluded_sb) if sb_matches else 0
            merged_count = merged_regular + merged_sb
            
            summary = format_report_summary(
                totals_merged, 
                "СВОДНЫЙ (Москвич + Анора)", 
                f"{date_from} .. {date_to}",
                len(report_rows_merged),
                merged_count
            )
            await msg.reply_text(summary)
        except Exception as e:
            await msg.reply_text(f"❌ Ошибка генерации отчёта: {str(e)}")
            return
        
        # Экспорт сводного с тремя листами
        try:
            filename = f"otchet_svodny_{date_from}_{date_to}.xlsx"
            ReportGenerator.generate_merged_xlsx(
                report_moskvich=(report_rows_m, totals_m),
                report_anora=(report_rows_a, totals_a),
                report_merged=(report_rows_merged, totals_merged),
                period=f"{date_from} .. {date_to}",
                filename=filename,
                db=db
            )
            with open(filename, 'rb') as f:
                await msg.reply_document(
                    document=f, filename=filename,
                    caption=f"📊 СВОДНЫЙ ОТЧЁТ (Оба клуба)\nПериод: {date_from} .. {date_to}\n\n📄 Файл содержит 3 листа:\n• Москвич\n• Анора\n• Сводный"
                )
            os.remove(filename)
        except Exception as e:
            await msg.reply_text(f"⚠️ Ошибка создания Excel: {str(e)}")
    else:
        await msg.reply_text("ℹ️ Нет данных для сводного отчета")


def find_code_duplicates(operations: list) -> list:
    """
    Поиск дубликатов: один код, но одна запись с именем, другая без
    """
    from collections import defaultdict
    
    by_code = defaultdict(lambda: {'with_name': [], 'without_name': []})
    
    for op in operations:
        code = op['code']
        if op['name']:
            by_code[code]['with_name'].append(op)
        else:
            by_code[code]['without_name'].append(op)
    
    # Ищем коды где есть И с именем И без имени
    duplicates = []
    for code, data in by_code.items():
        if data['with_name'] and data['without_name']:
            duplicates.append({
                'code': code,
                'with_name': data['with_name'],
                'without_name': data['without_name']
            })
    
    return duplicates


def name_similarity(name1: str, name2: str) -> float:
    """
    Вычисление похожести двух имен с приоритетом фамилии (0.0 - 1.0)
    Фамилия (последнее слово) имеет вес 70%, имя - 30%
    """
    if not name1 or not name2:
        return 0.0
    
    name1_clean = name1.lower().strip()
    name2_clean = name2.lower().strip()
    
    # Разбиваем на части
    parts1 = name1_clean.split()
    parts2 = name2_clean.split()
    
    if not parts1 or not parts2:
        return 0.0
    
    # Если одно из имен содержит только одно слово - обычное сравнение
    if len(parts1) == 1 or len(parts2) == 1:
        return SequenceMatcher(None, name1_clean, name2_clean).ratio()
    
    # Извлекаем фамилию (последнее слово) и имя (остальное)
    surname1 = parts1[-1]
    surname2 = parts2[-1]
    firstname1 = ' '.join(parts1[:-1])
    firstname2 = ' '.join(parts2[:-1])
    
    # Словарь сокращений имен
    name_abbreviations = {
        'дима': 'дмитрий',
        'дмитр': 'дмитрий',
        'саша': 'александр',
        'алекс': 'александр',
        'лёша': 'алексей',
        'леша': 'алексей',
        'макс': 'максим',
        'максимка': 'максим',
        'миша': 'михаил',
        'паша': 'павел',
        'женя': 'евгений',
        'вова': 'владимир',
        'володя': 'владимир',
        'коля': 'николай',
        'серёга': 'сергей',
        'серега': 'сергей',
        'андрюха': 'андрей',
        'влад': 'владислав',
        'юра': 'юрий',
        'катя': 'екатерина',
        'настя': 'анастасия',
        'маша': 'мария',
        'лена': 'елена',
        'оля': 'ольга',
        'таня': 'татьяна',
        'вика': 'виктория',
        'даша': 'дарья'
    }
    
    # Нормализуем имена через словарь сокращений
    firstname1_normalized = name_abbreviations.get(firstname1.lower(), firstname1.lower())
    firstname2_normalized = name_abbreviations.get(firstname2.lower(), firstname2.lower())
    
    # Сравниваем фамилии
    surname_similarity = SequenceMatcher(None, surname1, surname2).ratio()
    
    # Сравниваем имена (с учетом нормализации)
    firstname_similarity = SequenceMatcher(None, firstname1_normalized, firstname2_normalized).ratio()
    
    # Взвешенная сумма: фамилия 70%, имя 30%
    weighted_similarity = surname_similarity * 0.7 + firstname_similarity * 0.3
    
    return weighted_similarity


def normalize_name_variants(name: str) -> List[str]:
    """Создание вариантов имени в разных порядках (ФИО, ИОФ, ОИФ)"""
    if not name:
        return ['']
    
    name = name.strip()
    parts = name.split()
    
    if len(parts) <= 1:
        return [name]
    
    # Убираем лишние пробелы и приводим к единому формату
    normalized = ' '.join(part.strip() for part in parts if part.strip())
    
    # Если 2 части - только 2 варианта
    if len(parts) == 2:
        return [normalized, f"{parts[1]} {parts[0]}"]
    
    # Если 3+ части - создаем основные варианты
    variants = set()
    variants.add(normalized)  # Оригинальный порядок
    
    if len(parts) >= 2:
        variants.add(f"{parts[1]} {parts[0]}")  # ИОФ
        if len(parts) >= 3:
            variants.add(f"{parts[0]} {parts[2]} {parts[1]}")  # ФОИ
            variants.add(f"{parts[1]} {parts[0]} {parts[2]}")  # ИОФ (полный)
    
    return list(variants)


def extract_surname_candidates(name: str) -> set:
    """
    Извлечение возможных вариантов фамилии из имени
    Возвращает set из возможных фамилий (учитывает разный порядок слов)
    """
    if not name:
        return set()
    
    parts = name.strip().split()
    if len(parts) == 0:
        return set()
    
    # Словарь популярных имен (не фамилий) - расширенный
    common_first_names = {
        'дима', 'дмитрий', 'дмитр', 'димон', 'митя',
        'александр', 'саша', 'алекс', 'сан', 'шура',
        'максим', 'макс', 'максимка',
        'иван', 'ваня', 'ванька',
        'петр', 'петя', 'пётр',
        'сергей', 'серёга', 'серега', 'серёжа', 'сережа',
        'андрей', 'андрюха', 'дрей',
        'алексей', 'лёша', 'леша', 'лёха', 'леха', 'алёша', 'алеша',
        'михаил', 'миша', 'мишка', 'михась',
        'павел', 'паша', 'пашка',
        'николай', 'коля', 'колька', 'николя',
        'владимир', 'вова', 'володя', 'вован', 'влад',
        'евгений', 'женя', 'жека',
        'юрий', 'юра', 'юрка',
        'владислав', 'влад', 'владик',
        'артем', 'артём', 'тёма', 'тема',
        'денис', 'ден', 'дэн',
        'роман', 'рома', 'ромка',
        'игорь', 'гарик',
        'олег', 'олежка',
        'виктор', 'витя', 'витек',
        'анатолий', 'толя', 'толик',
        'екатерина', 'катя', 'катюша', 'катерина',
        'анастасия', 'настя', 'настюша',
        'мария', 'маша', 'машка', 'марья',
        'ольга', 'оля', 'олюшка',
        'татьяна', 'таня', 'танюша',
        'елена', 'лена', 'ленка', 'алёна', 'алена',
        'наталья', 'наташа', 'ната',
        'светлана', 'света', 'светик',
        'ирина', 'ира', 'ирка',
        'виктория', 'вика', 'викуля',
        'дарья', 'даша', 'дашка',
        'анна', 'аня', 'анька', 'аннушка'
    }
    
    # Если одно слово - возвращаем его
    if len(parts) == 1:
        return {parts[0].lower()}
    
    # Если два+ слова - определяем что фамилия
    surnames = set()
    
    for part in parts:
        part_lower = part.lower()
        # Если слово НЕ является популярным именем - это кандидат на фамилию
        if part_lower not in common_first_names:
            surnames.add(part_lower)
    
    # Если не нашли фамилию (все слова - имена или неизвестны) - берем все
    if not surnames:
        surnames = {p.lower() for p in parts}
    
    return surnames


def find_sb_name_duplicates(operations: list, similarity_threshold: float = 0.75) -> list:
    """
    Поиск СБ сотрудников с похожими именами для объединения
    Использует СТРОГУЮ кластеризацию по фамилии для точности
    similarity_threshold: порог похожести (0.75 = 75%)
    """
    from collections import defaultdict
    
    # Фильтруем только СБ
    sb_operations = [op for op in operations if op['code'] == 'СБ' and op.get('name')]
    
    if len(sb_operations) < 2:
        return []
    
    # Группируем по именам
    by_name = defaultdict(list)
    for op in sb_operations:
        name = op['name'].strip()
        if name:
            by_name[name].append(op)
    
    names_list = list(by_name.keys())
    
    # ШАГ 1: Строгая кластеризация по фамилии
    # Используем Union-Find для группировки
    parent = {name: name for name in names_list}
    
    def find(x):
        if parent[x] != x:
            parent[x] = find(parent[x])
        return parent[x]
    
    def union(x, y):
        px, py = find(x), find(y)
        if px != py:
            parent[px] = py
    
    # Объединяем имена с общей фамилией
    for i, name1 in enumerate(names_list):
        surnames1 = extract_surname_candidates(name1)
        
        for j, name2 in enumerate(names_list[i+1:], i+1):
            surnames2 = extract_surname_candidates(name2)
            
            # Проверяем: есть ли общая фамилия с похожестью >= 90%
            has_matching_surname = False
            
            for s1 in surnames1:
                for s2 in surnames2:
                    # СТРОГОЕ сравнение фамилий
                    surname_similarity = SequenceMatcher(None, s1, s2).ratio()
                    if surname_similarity >= 0.90:
                        has_matching_surname = True
                        break
                if has_matching_surname:
                    break
            
            if has_matching_surname:
                union(name1, name2)
    
    # Собираем кластеры
    clusters = defaultdict(list)
    for name in names_list:
        root = find(name)
        clusters[root].append(name)
    
    # ШАГ 2: Формируем группы только если в кластере > 1 имени
    name_groups = []
    
    for root, cluster_names in clusters.items():
        if len(cluster_names) > 1:
            # Собираем все операции для этого кластера
            group_operations = []
            for name in cluster_names:
                group_operations.extend(by_name[name])
            
            # Вычисляем суммы
            total_nal = sum(op['amount'] for op in group_operations if op['channel'] == 'нал')
            total_beznal = sum(op['amount'] for op in group_operations if op['channel'] == 'безнал')
            
            # Определяем основное имя (самое полное/длинное)
            main_name = max(cluster_names, key=lambda n: (len(n.split()), len(n)))
            
            # Вычисляем максимальную похожесть в группе
            max_similarity = 0.0
            for i, n1 in enumerate(cluster_names):
                for n2 in cluster_names[i+1:]:
                    variants1 = normalize_name_variants(n1)
                    variants2 = normalize_name_variants(n2)
                    for v1 in variants1:
                        for v2 in variants2:
                            sim = name_similarity(v1, v2)
                            max_similarity = max(max_similarity, sim)
            
            name_groups.append({
                'names': sorted(cluster_names),  # Сортируем для консистентности
                'main_name': main_name,
                'operations': group_operations,
                'total_nal': total_nal,
                'total_beznal': total_beznal,
                'similarity': max_similarity if max_similarity > 0 else 1.0
            })
    
    return name_groups


async def handle_duplicate_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                                       state: UserState, text: str, text_lower: str):
    """Обработка подтверждения объединения дубликатов"""
    if not state.duplicate_check_data:
        await update.message.reply_text("❌ Ошибка: данные не найдены")
        state.mode = None
        return
    
    data = state.duplicate_check_data
    duplicates = data['duplicates']
    operations = data['operations']
    
    # Обработка ответа с новой логикой
    indices_to_merge = set()
    
    # Убираем знаки препинания для удобства парсинга: "не1,2" -> "не 1 2"
    normalized_text = text_lower.replace(',', ' ').replace('.', ' ')
    parts = normalized_text.split()
    
    if not parts:
        await update.message.reply_text("❌ Неверный формат. Используйте: ок, ок 1, ок 1 2, не 1, не 1 2")
        return
    
    command = parts[0]
    
    if command in ['ок', 'ok']:
        # "ок" без номеров -> объединить ВСЕ
        if len(parts) == 1:
            indices_to_merge = set(range(len(duplicates)))
        else:
            # "ок 1 2" -> объединить ТОЛЬКО указанные
            try:
                indices_to_merge = set(int(x) - 1 for x in parts[1:] if x.isdigit())
            except:
                await update.message.reply_text("❌ Неверный формат номеров. Используйте: ок 1 2")
                return
    elif command in ['не', 'net', 'нет']:
        # "не 1 2" -> НЕ объединять указанные (объединить остальные)
        try:
            exclude_indices = set(int(x) - 1 for x in parts[1:] if x.isdigit())
            indices_to_merge = set(range(len(duplicates))) - exclude_indices
        except:
            await update.message.reply_text("❌ Неверный формат номеров. Используйте: не 1 2")
            return
    else:
        await update.message.reply_text(
            "❌ Неверная команда.\n\n"
            "Используйте:\n"
            "• ок - объединить все\n"
            "• ок 1 - объединить только пункт 1\n"
            "• ок 1 2 - объединить пункты 1 и 2\n"
            "• не 1 - НЕ объединять пункт 1 (остальные объединить)\n"
            "• не 1 2 - НЕ объединять пункты 1 и 2"
        )
        return
    
    # СОХРАНЯЕМ ОБЪЕДИНЕНИЕ В БД!
    updated_count = 0
    
    for i, dup in enumerate(duplicates):
        if i in indices_to_merge:
            code = dup['code']
            
            # Берём имя из записи с именем
            if dup['with_name']:
                merged_name = dup['with_name'][0]['name']
                
                # Обновляем ВСЕ записи БЕЗ имени для этого кода в БД
                for op_without_name in dup['without_name']:
                    # Обновляем в БД
                    success, msg = db.update_operation_name(
                        club=data['club'],
                        date=op_without_name['date'],
                        code=code,
                        channel=op_without_name['channel'],
                        new_name=merged_name
                    )
                    if success:
                        updated_count += 1
    
    # Получаем ОБНОВЛЁННЫЕ данные из БД
    updated_operations = db.get_operations_by_period(data['club'], data['date_from'], data['date_to'])
    
    # Проверяем СБ с похожими именами после обработки дубликатов кода
    sb_duplicates = find_sb_name_duplicates(updated_operations)
    if sb_duplicates:
        # Показываем запрос на объединение СБ
        await prepare_sb_merge(update, state, data['club'], data['date_from'], data['date_to'], updated_operations, sb_duplicates)
        return
    
    # Загружаем расходы на стилистов для этого периода
    stylist_expenses = db.get_stylist_expenses_for_period(data['club'], data['date_from'], data['date_to'])
    
    # Генерируем отчёт с объединёнными данными
    report_rows, totals, totals_recalc, check_ok = ReportGenerator.calculate_report(
        updated_operations,
        stylist_expenses=stylist_expenses
    )
    
    # Краткая сводка с информацией об объединении
    summary = format_report_summary(
        totals, 
        data['club'], 
        f"{data['date_from']} .. {data['date_to']}",
        len(report_rows),
        updated_count
    )
    
    await update.message.reply_text(summary)
    
    # Создаем XLSX
    club_translit = 'moskvich' if data['club'] == 'Москвич' else 'anora'
    filename = f"otchet_{club_translit}_{data['date_from']}_{data['date_to']}.xlsx"
    
    ReportGenerator.generate_xlsx(report_rows, totals, data['club'], 
                                  f"{data['date_from']} .. {data['date_to']}", filename, db)
    
    with open(filename, 'rb') as f:
        await update.message.reply_document(
            document=f,
            filename=filename,
            caption=f"📊 Отчет {data['club']} ({data['date_from']} .. {data['date_to']})"
        )
    
    os.remove(filename)
    
    # Проверяем, был ли выбран "оба" клуба - если да, продолжаем обработку
    if state.report_club == 'оба':
        # Отслеживаем обработанные клубы чтобы избежать зацикленности
        if not hasattr(state, 'processed_clubs_for_report'):
            state.processed_clubs_for_report = set()
        
        processed_club = data['club']
        state.processed_clubs_for_report.add(processed_club)
        
        # Определяем оставшиеся клубы
        all_clubs = {'Москвич', 'Анора'}
        remaining_clubs = all_clubs - state.processed_clubs_for_report
        
        # Если есть необработанные клубы - обрабатываем
        if remaining_clubs:
            for club in remaining_clubs:
                await generate_and_send_report(update, club, data['date_from'], data['date_to'], state, check_duplicates=True)
                # Если установлен режим ожидания - прерываем цикл и ждём подтверждения
                if state.mode in ['awaiting_duplicate_confirm', 'awaiting_sb_merge_confirm']:
                    return
        
        # Если ВСЕ клубы обработаны И нет активных режимов ожидания - генерируем сводный отчет
        if len(state.processed_clubs_for_report) == 2 and state.mode not in ['awaiting_duplicate_confirm', 'awaiting_sb_merge_confirm']:
            await prepare_merged_report(update, state, data['date_from'], data['date_to'])
            
            # НЕ сбрасываем режим если ждём подтверждения объединения!
            if state.mode != 'awaiting_merge_confirm':
                state.mode = None
                state.report_club = None
                state.processed_clubs_for_report = set()
                state.pending_report_period = None
    else:
        # Очищаем состояние
        state.mode = None
        state.duplicate_check_data = None
        state.sb_merge_data = None
        state.report_club = None


async def handle_sb_merge_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE,
                                      state: UserState, text: str, text_lower: str, message=None):
    """Обработка подтверждения объединения СБ с похожими именами"""
    # Используем message если передан, иначе update.message
    msg = message if message else update.message
    
    
    if not state.sb_merge_data:
        await msg.reply_text("❌ Ошибка: данные не найдены")
        state.mode = None
        return
    
    data = state.sb_merge_data
    sb_duplicates = data['sb_duplicates']
    
    # Обработка ответа
    indices_to_merge = set()
    
    # Убираем знаки препинания
    normalized_text = text_lower.replace(',', ' ').replace('.', ' ')
    parts = normalized_text.split()
    
    if not parts:
        await msg.reply_text("❌ Неверный формат. Используйте: ок, ок 1, ок 1 2, не 1, не 1 2")
        return
    
    command = parts[0]
    
    if command in ['ок', 'ok']:
        if len(parts) == 1:
            indices_to_merge = set(range(len(sb_duplicates)))
        else:
            try:
                indices_to_merge = set(int(x) - 1 for x in parts[1:] if x.isdigit())
            except:
                await msg.reply_text("❌ Неверный формат номеров. Используйте: ок 1 2")
                return
    elif command in ['не', 'net', 'нет']:
        try:
            exclude_indices = set(int(x) - 1 for x in parts[1:] if x.isdigit())
            indices_to_merge = set(range(len(sb_duplicates))) - exclude_indices
        except:
            await msg.reply_text("❌ Неверный формат номеров. Используйте: не 1 2")
            return
    else:
        await msg.reply_text(
            "❌ Неверная команда.\n\n"
            "Используйте:\n"
            "• ок - объединить все\n"
            "• ок 1 - объединить только пункт 1\n"
            "• ок 1 2 - объединить пункты 1 и 2\n"
            "• не 1 - НЕ объединять пункт 1 (остальные объединить)\n"
            "• не 1 2 - НЕ объединять пункты 1 и 2"
        )
        return
    
    # Создаем словарь объединений (ТОЛЬКО для отчета, БД не изменяем!)
    sb_name_merges = {}
    merged_sb_count = 0  # Счетчик объединенных СБ
    
    for i, group in enumerate(sb_duplicates):
        if i in indices_to_merge:
            main_name = group['main_name']
            
            # Для всех похожих имен указываем основное имя
            for name in group['names']:
                if name != main_name:
                    # Нормализуем имена (Ё→Е) для корректного сопоставления
                    normalized_name = name.replace('ё', 'е').replace('Ё', 'Е')
                    normalized_main = main_name.replace('ё', 'е').replace('Ё', 'Е')
                    
                    # Добавляем оба варианта (с Ё и без) для надёжности
                    sb_name_merges[name] = main_name
                    if normalized_name != name:
                        sb_name_merges[normalized_name] = main_name
                    
                    merged_sb_count += 1  # Считаем объединенные имена
    
    # Получаем данные из БД (БЕЗ изменений!)
    operations = db.get_operations_by_period(data['club'], data['date_from'], data['date_to'])
    
    # Загружаем расходы на стилистов для этого периода
    stylist_expenses = db.get_stylist_expenses_for_period(data['club'], data['date_from'], data['date_to'])
    
    # Генерируем отчёт с объединёнными данными (только для отчета)
    report_rows, totals, totals_recalc, check_ok = ReportGenerator.calculate_report(
        operations, 
        sb_name_merges=sb_name_merges if sb_name_merges else None,
        stylist_expenses=stylist_expenses
    )
    
    # Краткая сводка с информацией об объединенных СБ
    summary_lines = []
    summary_lines.append("✅ ОТЧЁТ ГОТОВ!\n")
    summary_lines.append(f"🏢 Клуб: {data['club']}")
    summary_lines.append(f"📅 Период: {data['date_from']} .. {data['date_to']}")
    summary_lines.append(f"👥 Сотрудников: {len(report_rows)}")
    
    if merged_sb_count > 0:
        summary_lines.append(f"🔄 Объединено СБ имён: {merged_sb_count} (только в отчете)")
    
    summary_lines.append("\n💰 ИТОГО:")
    summary_lines.append(f"   НАЛ:      {totals['nal']:,.0f}".replace(',', ' '))
    summary_lines.append(f"   БЕЗНАЛ:   {totals['beznal']:,.0f}".replace(',', ' '))
    summary_lines.append(f"   10%:      {totals['minus10']:,.0f}".replace(',', ' '))
    summary_lines.append(f"   {'─' * 25}")
    summary_lines.append(f"   ИТОГО:    {totals['itog']:,.0f}".replace(',', ' '))
    summary_lines.append("\n📄 Детальный отчёт в Excel файле ⬇️")
    
    summary = '\n'.join(summary_lines)
    
    await msg.reply_text(summary)
    
    # Создаем XLSX
    club_translit = 'moskvich' if data['club'] == 'Москвич' else 'anora'
    filename = f"otchet_{club_translit}_{data['date_from']}_{data['date_to']}.xlsx"
    
    ReportGenerator.generate_xlsx(report_rows, totals, data['club'], 
                                  f"{data['date_from']} .. {data['date_to']}", filename, db)
    
    with open(filename, 'rb') as f:
        await msg.reply_document(
            document=f,
            filename=filename,
            caption=f"📊 Отчет {data['club']} ({data['date_from']} .. {data['date_to']})"
        )
    
    os.remove(filename)
    
    # СОХРАНЯЕМ словарь объединений СБ в state для сводного отчёта
    if sb_name_merges:
        if data['club'] == 'Москвич':
            state.sb_merges_moskvich = sb_name_merges
        elif data['club'] == 'Анора':
            state.sb_merges_anora = sb_name_merges
    
    # Проверяем, был ли выбран "оба" клуба - если да, продолжаем обработку
    if state.report_club == 'оба':
        # Отслеживаем обработанные клубы чтобы избежать зацикленности
        if not hasattr(state, 'processed_clubs_for_report'):
            state.processed_clubs_for_report = set()
        
        processed_club = data['club']
        state.processed_clubs_for_report.add(processed_club)
        
        # Определяем оставшиеся клубы
        all_clubs = {'Москвич', 'Анора'}
        remaining_clubs = all_clubs - state.processed_clubs_for_report
        
        # Если есть необработанные клубы - обрабатываем
        if remaining_clubs:
            new_update = update
            
            # Обрабатываем оставшийся клуб через generate_and_send_report
            for club in remaining_clubs:
                await generate_and_send_report(new_update, club, data['date_from'], data['date_to'], state, check_duplicates=True, message=msg)
                # Если установлен режим ожидания - выходим
                if state.mode in ['awaiting_duplicate_confirm', 'awaiting_sb_merge_confirm']:
                    return
        
        # Проверяем - все ли клубы обработаны?
        if len(state.processed_clubs_for_report) == 2:
            new_update = update
            
            await prepare_merged_report(new_update, state, data['date_from'], data['date_to'])
            
            # НЕ сбрасываем режим если ждём подтверждения объединения!
            if state.mode != 'awaiting_merge_confirm':
                state.mode = None
                state.report_club = None
                state.processed_clubs_for_report = set()
                state.pending_report_period = None
    else:
        # Очищаем состояние
        state.mode = None
        state.sb_merge_data = None
        state.report_club = None


async def prepare_sb_merge_with_message(msg, state: UserState, club: str, date_from: str,
                           date_to: str, operations: list, sb_duplicates: list):
    """Подготовка объединения СБ с похожими именами (принимает message напрямую)"""
    await prepare_sb_merge_internal(msg, state, club, date_from, date_to, operations, sb_duplicates)


async def prepare_sb_merge(update: Update, state: UserState, club: str, date_from: str,
                           date_to: str, operations: list, sb_duplicates: list):
    """Подготовка объединения СБ с похожими именами (обертка для update)"""
    msg = update.message
    await prepare_sb_merge_internal(msg, state, club, date_from, date_to, operations, sb_duplicates)


async def prepare_sb_merge_internal(msg, state: UserState, club: str, date_from: str,
                           date_to: str, operations: list, sb_duplicates: list):
    """Внутренняя функция подготовки объединения СБ"""
    # Создаем текстовый файл со списком СБ кандидатов
    file_content = ["📋 НАЙДЕНЫ СБ С ПОХОЖИМИ ИМЕНАМИ\n"]
    file_content.append(f"Клуб: {club}\n")
    file_content.append(f"Период: {date_from} .. {date_to}\n")
    file_content.append("=" * 50 + "\n\n")
    
    for i, group in enumerate(sb_duplicates, 1):
        similarity_pct = int(group['similarity'] * 100)
        file_content.append(f"{i}. Группа: {group['main_name']}\n")
        file_content.append(f"   Похожесть: {similarity_pct}%\n")
        
        # Группируем операции по именам для отображения
        by_name = {}
        for op in group['operations']:
            name = op['name']
            if name not in by_name:
                by_name[name] = {'nal': 0, 'beznal': 0}
            if op['channel'] == 'нал':
                by_name[name]['nal'] += op['amount']
            else:
                by_name[name]['beznal'] += op['amount']
        
        for name in group['names']:
            if name in by_name:
                file_content.append(f"   • {name}: НАЛ {by_name[name]['nal']:.0f}, БЕЗНАЛ {by_name[name]['beznal']:.0f}\n")
        
        file_content.append(f"   ИТОГО: НАЛ {group['total_nal']:.0f}, БЕЗНАЛ {group['total_beznal']:.0f}\n")
        file_content.append("\n")
    
    file_content.append("=" * 50 + "\n")
    file_content.append("\n🔄 ОБЪЕДИНЕНИЕ СБ:\n")
    file_content.append("• ОК → объединить все\n")
    file_content.append("• ОК 1 → объединить только пункт 1\n")
    file_content.append("• ОК 1 2 → объединить пункты 1 и 2\n")
    file_content.append("• НЕ 1 → НЕ объединять пункт 1 (остальные да)\n")
    file_content.append("• НЕ 1 2 → НЕ объединять пункты 1 и 2\n")
    file_content.append("\n⚠️ ВАЖНО: объединение применяется ТОЛЬКО для отчета\n")
    file_content.append("          (база данных НЕ изменяется)\n")
    
    # Сохраняем во временный файл
    temp_file = tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', suffix='.txt', delete=False)
    temp_file.write(''.join(file_content))
    temp_file.close()
    
    # Отправляем короткое сообщение с кнопками
    count = len(sb_duplicates)
    short_message = (
        f"📋 Найдено групп СБ: {count}\n\n"
        f"🔄 Объединение СБ с похожими именами:\n"
        f"• Используйте кнопки ниже\n"
        f"• Или введите: ОК / ОК 1 / НЕ 1\n\n"
        f"📄 Детальный список в файле ⬇️\n\n"
        f"⚠️ ВАЖНО: объединение применяется ТОЛЬКО для отчета\n"
        f"          (база данных НЕ изменяется)"
    )
    
    # Отправляем файл и сообщение с кнопками
    with open(temp_file.name, 'rb') as f:
        await msg.reply_document(
            document=f,
            filename=f"sb_merge_{club}_{date_from}_{date_to}.txt",
            caption=short_message,
            reply_markup=get_merge_confirmation_keyboard()
        )
    
    # Удаляем временный файл
    os.remove(temp_file.name)
    
    # Сохраняем данные для обработки
    state.sb_merge_data = {
        'club': club,
        'date_from': date_from,
        'date_to': date_to,
        'operations': operations,
        'sb_duplicates': sb_duplicates
    }
    state.mode = 'awaiting_sb_merge_confirm'


async def generate_and_send_report(update: Update, club: str, date_from: str, date_to: str, 
                                  state: UserState = None, check_duplicates: bool = True, message=None, sb_name_merges: dict = None):
    """Генерация и отправка отчета"""
    # Определяем куда отправлять сообщения
    msg = message if message else update.message
    
    operations = db.get_operations_by_period(club, date_from, date_to)
    
    if not operations:
        await msg.reply_text(
            f"📊 Отчет по клубу {club}\n"
            f"Период: {date_from} .. {date_to}\n\n"
            f"Данных нет."
        )
        return
    
    # Проверка на дубликаты (одинаковый код, но с именем и без)
    if check_duplicates and state:
        duplicates = find_code_duplicates(operations)
        
        if duplicates:
            # Показываем запрос на объединение
            response = [f"⚠️ Найдены записи с одинаковым кодом:\n"]
            
            for i, dup in enumerate(duplicates, 1):
                response.append(f"{i}. Код: {dup['code']}")
                
                # С именем
                names_with = set(op['name'] for op in dup['with_name'])
                for name in names_with:
                    ops = [op for op in dup['with_name'] if op['name'] == name]
                    total_nal = sum(op['amount'] for op in ops if op['channel'] == 'нал')
                    total_bez = sum(op['amount'] for op in ops if op['channel'] == 'безнал')
                    response.append(f"   • {name}: НАЛ {total_nal:.0f}, БЕЗНАЛ {total_bez:.0f}")
                
                # Без имени
                total_nal_no = sum(op['amount'] for op in dup['without_name'] if op['channel'] == 'нал')
                total_bez_no = sum(op['amount'] for op in dup['without_name'] if op['channel'] == 'безнал')
                response.append(f"   • (без имени): НАЛ {total_nal_no:.0f}, БЕЗНАЛ {total_bez_no:.0f}")
                response.append("")
            
            response.append("─" * 35)
            response.append("\n🔄 ОБЪЕДИНЕНИЕ ДУБЛИКАТОВ:\n")
            response.append("• ОК → объединить все")
            response.append("• ОК 1 → объединить только пункт 1")
            response.append("• ОК 1 2 → объединить пункты 1 и 2")
            response.append("• НЕ 1 → НЕ объединять пункт 1 (остальные да)")
            response.append("• НЕ 1 2 → НЕ объединять пункты 1 и 2")
            
            await msg.reply_text('\n'.join(response))
            
            # Сохраняем данные для обработки
            state.duplicate_check_data = {
                'club': club,
                'date_from': date_from,
                'date_to': date_to,
                'operations': operations,
                'duplicates': duplicates
            }
            state.mode = 'awaiting_duplicate_confirm'
            return
    
    # Проверка на СБ с похожими именами (после проверки дубликатов кода)
    if check_duplicates and state:
        sb_duplicates = find_sb_name_duplicates(operations)
        
        if sb_duplicates:
            # Показываем запрос на объединение СБ с инлайн-кнопками и файлом
            await prepare_sb_merge_with_message(msg, state, club, date_from, date_to, operations, sb_duplicates)
            return
    
    # Генерируем отчет (без дубликатов или после подтверждения)
    # Загружаем расходы на стилистов для этого периода
    stylist_expenses = db.get_stylist_expenses_for_period(club, date_from, date_to)
    
    report_rows, totals, totals_recalc, check_ok = ReportGenerator.calculate_report(
        operations,
        sb_name_merges=sb_name_merges if sb_name_merges else None,
        stylist_expenses=stylist_expenses
    )
    
    # Краткая сводка вместо полного отчёта
    summary = format_report_summary(
        totals, 
        club, 
        f"{date_from} .. {date_to}",
        len(report_rows)
    )
    await msg.reply_text(summary)
    
    # Создаем XLSX
    club_translit = 'moskvich' if club == 'Москвич' else 'anora'
    filename = f"otchet_{club_translit}_{date_from}_{date_to}.xlsx"
    
    ReportGenerator.generate_xlsx(
        report_rows, totals, club, f"{date_from} .. {date_to}", filename, db
    )
    
    # Отправляем файл
    with open(filename, 'rb') as f:
        await msg.reply_document(
            document=f,
            filename=filename,
            caption=f"📊 Отчет по клубу {club}\nПериод: {date_from} .. {date_to}"
        )
    
    # Удаляем временный файл
    os.remove(filename)
    
    # Если это часть обработки "оба" клуба - отмечаем клуб как обработанный
    if state and state.report_club == 'оба':
        if not hasattr(state, 'processed_clubs_for_report'):
            state.processed_clubs_for_report = set()
        state.processed_clubs_for_report.add(club)


async def handle_payments_command(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                                  state: UserState, text: str):
    """Обработка команды выплаты"""
    parts = text.split()
    
    # Определяем формат ввода
    if parts[0].lower() == 'выплаты':
        # Формат: выплаты Д1 30,10-1,11
        if len(parts) < 3:
            await update.message.reply_text(
                "❌ Неверный формат.\n"
                "Пример: выплаты Д1 30,10-1,11"
            )
            return
        code = DataParser.normalize_code(parts[1])
        period_str = parts[2]
    else:
        # Упрощённый формат (после кнопки): Д1 30,10-1,11
        if len(parts) < 2:
            await update.message.reply_text(
                "❌ Неверный формат.\n"
                "Пример: Д1 30,10-1,11"
            )
            return
        code = DataParser.normalize_code(parts[0])
        period_str = parts[1]
    
    # Парсим период (одна дата или диапазон)
    if '-' in period_str:
        success, date_from, date_to, error = parse_date_range(period_str)
        if not success:
            await update.message.reply_text(f"❌ {error}")
            return
    else:
        success, single_date, error = parse_short_date(period_str)
        if not success:
            await update.message.reply_text(f"❌ {error}")
            return
        date_from = single_date
        date_to = single_date
    
    # Получаем выплаты ПО ВСЕМ КЛУБАМ
    payments = db.get_employee_payments(code, date_from, date_to, None)
    
    if not payments:
        await update.message.reply_text(
            f"📊 Выплаты сотруднику {code}\n"
            f"Период: {date_from} .. {date_to}\n\n"
            f"Данных нет."
        )
        return
    
    # Группируем по клубам и датам
    from collections import defaultdict
    by_club = defaultdict(lambda: {'nal': 0, 'beznal': 0, 'by_date': defaultdict(lambda: {'nal': 0, 'beznal': 0})})
    
    for payment in payments:
        club = payment['club']
        date = payment['date']
        amount = payment['amount']
        
        # Группируем по дате
        if payment['channel'] == 'нал':
            by_club[club]['by_date'][date]['nal'] += amount
            by_club[club]['nal'] += amount
        else:
            by_club[club]['by_date'][date]['beznal'] += amount
            by_club[club]['beznal'] += amount
    
    # Создаем Excel файл
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Выплаты"
    
    # Стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    ws['A1'] = f"Выплаты сотруднику {code}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Период: {date_from} .. {date_to}"
    ws['A2'].font = Font(size=11)
    
    row_num = 4
    
    # Общие итоги
    total_nal = 0
    total_beznal = 0
    
    # Выводим по каждому клубу
    for club in sorted(by_club.keys()):
        data = by_club[club]
        
        # Заголовок клуба
        ws.cell(row=row_num, column=1, value=f"Клуб: {club}")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        row_num += 1
        
        # Шапка таблицы
        headers = ['Дата', 'НАЛ', 'БЕЗНАЛ', '10%', 'ИТОГО']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        row_num += 1
        
        # Данные по датам
        for date in sorted(data['by_date'].keys()):
            date_data = data['by_date'][date]
            nal_sum = date_data['nal']
            beznal_sum = date_data['beznal']
            minus10 = beznal_sum * 0.1
            itog = nal_sum + (beznal_sum - minus10)
            
            # Преобразуем дату из 2024-10-30 в 30.10.24
            try:
                year, month, day = date.split('-')
                date_short = f"{day}.{month}.{year[2:]}"
            except:
                date_short = date
            
            # Дата
            cell = ws.cell(row=row_num, column=1, value=date_short)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # НАЛ
            cell = ws.cell(row=row_num, column=2, value=nal_sum)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border
            
            # БЕЗНАЛ
            cell = ws.cell(row=row_num, column=3, value=beznal_sum)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border
            
            # 10%
            cell = ws.cell(row=row_num, column=4, value=minus10)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border
            
            # ИТОГО
            cell = ws.cell(row=row_num, column=5, value=itog)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border
            
            row_num += 1
        
        # Итог по клубу
        club_nal = data['nal']
        club_beznal = data['beznal']
        club_minus10 = club_beznal * 0.1
        club_total = club_nal + (club_beznal - club_minus10)
        
        cell = ws.cell(row=row_num, column=1, value='ИТОГО ПО КЛУБУ')
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = border
        
        cell = ws.cell(row=row_num, column=2, value=club_nal)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        
        cell = ws.cell(row=row_num, column=3, value=club_beznal)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        
        cell = ws.cell(row=row_num, column=4, value=club_minus10)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        
        cell = ws.cell(row=row_num, column=5, value=club_total)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.border = border
        
        row_num += 2  # Пропускаем строку
        
        total_nal += club_nal
        total_beznal += club_beznal
    
    # Общий итог
    total_minus10 = total_beznal * 0.1
    total_itog = total_nal + (total_beznal - total_minus10)
    
    cell = ws.cell(row=row_num, column=1, value='ИТОГО ПО ВСЕМ КЛУБАМ')
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = border
    
    cell = ws.cell(row=row_num, column=2, value=total_nal)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = border
    
    cell = ws.cell(row=row_num, column=3, value=total_beznal)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = border
    
    cell = ws.cell(row=row_num, column=4, value=total_minus10)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = border
    
    cell = ws.cell(row=row_num, column=5, value=total_itog)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='right', vertical='center')
    cell.border = border
    
    # Автоподгонка ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохраняем и отправляем
    filename = f"vyplaty_{code}_{date_from}_{date_to}.xlsx"
    wb.save(filename)
    
    with open(filename, 'rb') as f:
        await update.message.reply_document(
            document=f,
            filename=filename,
            caption=f"💰 Выплаты сотруднику {code}\nПериод: {date_from} .. {date_to}"
        )
    
    import os
    os.remove(filename)
    
    # Если ограниченный доступ - предлагаем повторить
    if state.limited_access:
        keyboard = [[InlineKeyboardButton("❌ Выход", callback_data="quick_exit")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            "✅ Готово!\n\n"
            "💡 Введите данные для нового запроса:\n"
            "Пример: Д7 12,12",
            reply_markup=reply_markup
        )
        state.mode = 'awaiting_payments_input'


async def handle_salary_command(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                                state: UserState, text: str):
    """Обработка команды ЗП - расширенный отчёт из таблицы payments"""
    parts = text.split()
    
    if len(parts) < 2:
        await update.message.reply_text(
            "❌ Неверный формат.\n\n"
            "Примеры:\n"
            "• Д7 3,10-5,11 (один сотрудник)\n"
            "• москвич 3,10-5,11 (весь клуб)\n"
            "• оба 3,10-5,11 (оба клуба)"
        )
        return
    
    # Определяем режим: код сотрудника или клуб
    first_param = parts[0].lower()
    period_str = parts[1]
    
    # Парсим период
    if '-' in period_str:
        success, date_from, date_to, error = parse_date_range(period_str)
        if not success:
            await update.message.reply_text(f"❌ {error}")
            return
    else:
        success, single_date, error = parse_short_date(period_str)
        if not success:
            await update.message.reply_text(f"❌ {error}")
            return
        date_from = single_date
        date_to = single_date
    
    # Определяем клуб и код
    if first_param in ['москвич', 'анора', 'оба']:
        # Режим: весь клуб
        mode = 'club'
        if first_param == 'оба':
            clubs = ['Москвич', 'Анора']
        else:
            clubs = ['Москвич' if first_param == 'москвич' else 'Анора']
        code = None
    else:
        # Режим: один сотрудник
        mode = 'employee'
        code = DataParser.normalize_code(first_param)
        clubs = None
    
    await update.message.reply_text("⏳ Генерирую отчёт...")
    
    # Генерируем Excel
    if mode == 'employee':
        await generate_salary_excel_by_employee(update, code, date_from, date_to)
    else:
        await generate_salary_excel_by_club(update, clubs, date_from, date_to)


async def generate_salary_excel_by_employee(update: Update, code: str, date_from: str, date_to: str):
    """
    Генерация Excel отчёта ЗП для одного сотрудника из таблицы payments
    
    Колонки в Excel:
    Дата | Код | Имя | Ставка | 3% ЛМ | 5% | Промо | CRZ | Cons | Чаевые | 
    ИТОГО выплат | Получила на смене | Долг БН | 10% (вычет) | Долг НАЛ | К выплате
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from collections import defaultdict
    
    # Получаем данные из БД по всем клубам
    all_payments = []
    for club in ['Москвич', 'Анора']:
        club_payments = db.get_payments(club, date_from, date_to)
        for row in club_payments:
            # row это tuple из БД, преобразуем в dict
            payment_dict = {
                'id': row[0],
                'club': row[1],
                'date': row[2],
                'code': row[3],
                'name': row[4],
                'stavka': row[5],
                'lm_3': row[6],
                'percent_5': row[7],
                'promo': row[8],
                'crz': row[9],
                'cons': row[10],
                'tips': row[11],
                'fines': row[12],
                'total_shift': row[13],
                'debt': row[14],
                'debt_nal': row[15],
                'to_pay': row[16],
                'created_at': row[17]
            }
            # Фильтруем по коду
            if payment_dict['code'] == code:
                all_payments.append(payment_dict)
    
    if not all_payments:
        await update.message.reply_text(
            f"📊 Отчёт ЗП для {code}\n"
            f"Период: {date_from} .. {date_to}\n\n"
            f"❌ Данных нет в таблице payments.\n"
            f"Загрузите данные через кнопку 'ЗАГРУЗИТЬ ЗП'"
        )
        return
    
    # Сортируем по дате и клубу
    all_payments.sort(key=lambda x: (x['date'], x['club']))
    
    # Создаём Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "ЗП"
    
    # Стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    ws['A1'] = f"Отчёт ЗП: {code} - {all_payments[0]['name']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Период: {date_from} .. {date_to}"
    ws['A2'].font = Font(size=11)
    
    row_num = 4
    
    # Шапка таблицы
    headers = [
        'Дата', 'Клуб', 'Код', 'Имя', 'Ставка', '3% ЛМ', '5%', 'Промо', 
        'CRZ', 'Cons', 'Чаевые', 'ИТОГО выплат', 'Получила на смене',
        'Долг БН', '10% (вычет)', 'Долг НАЛ', 'К выплате'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    row_num += 1
    
    # Итоговые суммы
    totals = {
        'stavka': 0, 'lm_3': 0, 'percent_5': 0, 'promo': 0,
        'crz': 0, 'cons': 0, 'tips': 0, 'total_shift': 0,
        'to_pay': 0, 'debt': 0, 'debt_nal': 0, 'final_pay': 0
    }
    
    # Данные
    for payment in all_payments:
        # Преобразуем дату
        try:
            year, month, day = payment['date'].split('-')
            date_short = f"{day}.{month}.{year[2:]}"
        except:
            date_short = payment['date']
        
        # Рассчитываем 10% и к выплате
        vychet_10 = round(payment['debt'] * 0.1)  # Округление до целого
        k_vyplate = round(payment['debt_nal'] + payment['debt'] - vychet_10)  # Без стилистов
        
        # Обработка кода для отображения
        display_code = payment['code']
        if display_code.startswith('СБ-'):
            display_code = 'СБ'  # Убираем имя из кода для отображения
        elif display_code.startswith('Уборщица'):
            display_code = 'Уборщица'  # Убираем "Москвич/Анора" из кода для отображения
        
        # Записываем строку
        row_data = [
            date_short,
            payment['club'],
            display_code,  # Используем обработанный код
            payment['name'],
            payment['stavka'],
            payment['lm_3'],
            payment['percent_5'],
            payment['promo'],
            payment['crz'],
            payment['cons'],
            payment['tips'],
            payment['total_shift'],
            payment['to_pay'],
            payment['debt'],
            vychet_10,
            payment['debt_nal'],
            k_vyplate  # БЕЗ stylist_amount
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = border
            if col > 4:  # Числовые столбцы
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Обновляем итоги
        totals['stavka'] += payment['stavka']
        totals['lm_3'] += payment['lm_3']
        totals['percent_5'] += payment['percent_5']
        totals['promo'] += payment['promo']
        totals['crz'] += payment['crz']
        totals['cons'] += payment['cons']
        totals['tips'] += payment['tips']
        totals['total_shift'] += payment['total_shift']
        totals['to_pay'] += payment['to_pay']
        totals['debt'] += payment['debt']
        totals['debt_nal'] += payment['debt_nal']
        totals['final_pay'] += k_vyplate
        
        row_num += 1
    
    # Строка ИТОГО
    vychet_10_total = round(totals['debt'] * 0.1)  # Округление до целого
    
    itogo_data = [
        'ИТОГО', '', '', '',
        totals['stavka'],
        totals['lm_3'],
        totals['percent_5'],
        totals['promo'],
        totals['crz'],
        totals['cons'],
        totals['tips'],
        totals['total_shift'],
        totals['to_pay'],
        totals['debt'],
        vychet_10_total,
        totals['debt_nal'],
        round(totals['final_pay'])  # Округление до целого
    ]
    
    for col, value in enumerate(itogo_data, 1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.font = Font(bold=True)
        cell.border = border
        if col > 4:
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Автоподгонка ширины
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 20)
    
    # Сохраняем и отправляем
    filename = f"zp_{code}_{date_from}_{date_to}.xlsx"
    wb.save(filename)
    
    with open(filename, 'rb') as f:
        await update.message.reply_document(
            document=f,
            filename=filename,
            caption=f"💵 Отчёт ЗП: {code}\nПериод: {date_from} .. {date_to}"
        )
    
    import os
    os.remove(filename)
    
    # === ВТОРОЙ ФАЙЛ: СТИЛИСТЫ (только для одного сотрудника) ===
    
    # Получаем расходы на стилистов для этого сотрудника
    stylist_records = []
    for club in ['Москвич', 'Анора']:
        conn = db.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("""
                SELECT period_from, period_to, amount, club
                FROM stylist_expenses
                WHERE club = ? AND code = ?
                  AND NOT (period_to < ? OR period_from > ?)
                ORDER BY period_from
            """, (club, code, date_from, date_to))
            
            rows = cursor.fetchall()
            for row in rows:
                stylist_records.append({
                    'period_from': row[0],
                    'period_to': row[1],
                    'amount': row[2],
                    'club': row[3]
                })
        except Exception as e:
            print(f"Ошибка получения стилистов: {e}")
        finally:
            conn.close()
    
    # Если есть данные по стилистам - создаём второй файл
    if stylist_records:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Стилисты"
        
        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        ws2['A1'] = f"Расходы на стилистов: {code}"
        ws2['A1'].font = Font(bold=True, size=14)
        ws2['A2'] = f"Период: {date_from} .. {date_to}"
        ws2['A2'].font = Font(size=11)
        
        row_num = 4
        
        # Шапка таблицы
        headers = ['Клуб', 'Период с', 'Период по', 'Сумма']
        for col, header in enumerate(headers, 1):
            cell = ws2.cell(row=row_num, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        row_num += 1
        
        # Данные
        total_stylist = 0
        for record in stylist_records:
            # Преобразуем даты
            try:
                year, month, day = record['period_from'].split('-')
                date_from_short = f"{day}.{month}.{year[2:]}"
            except:
                date_from_short = record['period_from']
            
            try:
                year, month, day = record['period_to'].split('-')
                date_to_short = f"{day}.{month}.{year[2:]}"
            except:
                date_to_short = record['period_to']
            
            row_data = [
                record['club'],
                date_from_short,
                date_to_short,
                record['amount']
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_num, column=col, value=value)
                cell.border = border
                if col == 4:  # Сумма
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            total_stylist += record['amount']
            row_num += 1
        
        # Строка ИТОГО
        itogo_data = ['ИТОГО', '', '', total_stylist]
        for col, value in enumerate(itogo_data, 1):
            cell = ws2.cell(row=row_num, column=col, value=value)
            cell.font = Font(bold=True)
            cell.border = border
            if col == 4:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Автоподгонка ширины
        for column in ws2.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws2.column_dimensions[column_letter].width = min(max_length + 2, 20)
        
        # Сохраняем и отправляем
        filename2 = f"stilisty_{code}_{date_from}_{date_to}.xlsx"
        wb2.save(filename2)
        
        with open(filename2, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=filename2,
                caption=f"💄 Стилисты: {code}\nПериод: {date_from} .. {date_to}"
            )
        
        import os
        os.remove(filename2)


async def generate_salary_excel_by_club(update: Update, clubs: List[str], date_from: str, date_to: str):
    """
    Генерация Excel отчёта ЗП для всех сотрудников клуба(ов)
    
    Колонки те же что и для одного сотрудника
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Получаем все данные из БД
    all_payments = []
    for club in clubs:
        club_payments = db.get_payments(club, date_from, date_to)
        for row in club_payments:
            payment_dict = {
                'id': row[0],
                'club': row[1],
                'date': row[2],
                'code': row[3],
                'name': row[4],
                'stavka': row[5],
                'lm_3': row[6],
                'percent_5': row[7],
                'promo': row[8],
                'crz': row[9],
                'cons': row[10],
                'tips': row[11],
                'fines': row[12],
                'total_shift': row[13],
                'debt': row[14],
                'debt_nal': row[15],
                'to_pay': row[16],
                'created_at': row[17]
            }
            all_payments.append(payment_dict)
    
    if not all_payments:
        club_names = ', '.join(clubs)
        await update.message.reply_text(
            f"📊 Отчёт ЗП для клуба: {club_names}\n"
            f"Период: {date_from} .. {date_to}\n\n"
            f"❌ Данных нет в таблице payments.\n"
            f"Загрузите данные через кнопку 'ЗАГРУЗИТЬ ЗП'"
        )
        return
    
    # Сортируем по дате, клубу и коду
    all_payments.sort(key=lambda x: (x['date'], x['club'], x['code']))
    
    # Группируем по дате
    payments_by_date = {}
    for payment in all_payments:
        date = payment['date']
        if date not in payments_by_date:
            payments_by_date[date] = []
        payments_by_date[date].append(payment)
    
    # Создаём Excel
    wb = Workbook()
    wb.remove(wb.active)  # Удаляем дефолтный лист
    
    club_names = ', '.join(clubs)  # Определяем для использования в create_sheet
    
    # Стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Шапка таблицы
    headers = [
        'Дата', 'Клуб', 'Код', 'Имя', 'Ставка', '3% ЛМ', '5%', 'Промо', 
        'CRZ', 'Cons', 'Чаевые', 'ИТОГО выплат', 'Получила на смене',
        'Долг БН', '10% (вычет)', 'Долг НАЛ', 'К выплате'
    ]
    
    # Функция для создания листа с данными
    def create_sheet(ws, title, payments_list, show_date_col=True):
        ws.title = title
        ws['A1'] = f"Отчёт ЗП: {club_names}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = f"Период: {date_from} .. {date_to}"
        ws['A2'].font = Font(size=11)
        
        row_num = 4
        
        # Шапка таблицы
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        row_num += 1
        
        # Итоговые суммы
        totals = {
            'stavka': 0, 'lm_3': 0, 'percent_5': 0, 'promo': 0,
            'crz': 0, 'cons': 0, 'tips': 0, 'total_shift': 0,
            'to_pay': 0, 'debt': 0, 'debt_nal': 0, 'final_pay': 0
        }
        
        # Данные
        for payment in payments_list:
            # Преобразуем дату
            try:
                year, month, day = payment['date'].split('-')
                date_short = f"{day}.{month}.{year[2:]}"
            except:
                date_short = payment['date']
            
            # Рассчитываем 10% и к выплате
            vychet_10 = round(payment['debt'] * 0.1)  # Округление до целого
            k_vyplate = round(payment['debt_nal'] + payment['debt'] - vychet_10)  # Без стилистов
            
            # Обработка кода для отображения
            display_code = payment['code']
            if display_code.startswith('СБ-'):
                display_code = 'СБ'  # Убираем имя из кода для отображения
            elif display_code.startswith('Уборщица'):
                display_code = 'Уборщица'  # Убираем "Москвич/Анора" из кода для отображения
            
            # Записываем строку
            row_data = [
                date_short if show_date_col else '',
                payment['club'],
                display_code,  # Используем обработанный код
                payment['name'],
                payment['stavka'],
                payment['lm_3'],
                payment['percent_5'],
                payment['promo'],
                payment['crz'],
                payment['cons'],
                payment['tips'],
                payment['total_shift'],
                payment['to_pay'],
                payment['debt'],
                vychet_10,
                payment['debt_nal'],
                k_vyplate  # БЕЗ stylist_amount
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = border
                if col > 4:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Обновляем итоги
            totals['stavka'] += payment['stavka']
            totals['lm_3'] += payment['lm_3']
            totals['percent_5'] += payment['percent_5']
            totals['promo'] += payment['promo']
            totals['crz'] += payment['crz']
            totals['cons'] += payment['cons']
            totals['tips'] += payment['tips']
            totals['total_shift'] += payment['total_shift']
            totals['to_pay'] += payment['to_pay']
            totals['debt'] += payment['debt']
            totals['debt_nal'] += payment['debt_nal']
            totals['final_pay'] += k_vyplate
            
            row_num += 1
        
        # Строка ИТОГО
        vychet_10_total = round(totals['debt'] * 0.1)  # Округление до целого
        
        itogo_data = [
            'ИТОГО', '', '', '',
            totals['stavka'],
            totals['lm_3'],
            totals['percent_5'],
            totals['promo'],
            totals['crz'],
            totals['cons'],
            totals['tips'],
            totals['total_shift'],
            totals['to_pay'],
            totals['debt'],
            vychet_10_total,
            totals['debt_nal'],
            round(totals['final_pay'])  # Округление до целого
        ]
        
        for col, value in enumerate(itogo_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.font = Font(bold=True)
            cell.border = border
            if col > 4:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Автоподгонка ширины
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 20)
        
        return totals
    
    # Создаём лист для каждой даты
    for date in sorted(payments_by_date.keys()):
        try:
            year, month, day = date.split('-')
            sheet_name = f"{day}.{month}.{year[2:]}"
        except:
            sheet_name = date
        
        ws = wb.create_sheet(title=sheet_name)
        create_sheet(ws, sheet_name, payments_by_date[date], show_date_col=True)
    
    # Создаём лист ИТОГО с группировкой по (код, имя)
    employee_totals = {}
    for payment in all_payments:
        # Обработка кода для отображения
        display_code = payment['code']
        if display_code.startswith('СБ-'):
            display_code = 'СБ'
        elif display_code.startswith('Уборщица'):
            display_code = 'Уборщица'
        
        key = (display_code, payment['name'])
        if key not in employee_totals:
            employee_totals[key] = {
                'code': display_code,
                'name': payment['name'],
                'stavka': 0, 'lm_3': 0, 'percent_5': 0, 'promo': 0,
                'crz': 0, 'cons': 0, 'tips': 0, 'total_shift': 0,
                'to_pay': 0, 'debt': 0, 'debt_nal': 0
            }
        
        employee_totals[key]['stavka'] += payment['stavka']
        employee_totals[key]['lm_3'] += payment['lm_3']
        employee_totals[key]['percent_5'] += payment['percent_5']
        employee_totals[key]['promo'] += payment['promo']
        employee_totals[key]['crz'] += payment['crz']
        employee_totals[key]['cons'] += payment['cons']
        employee_totals[key]['tips'] += payment['tips']
        employee_totals[key]['total_shift'] += payment['total_shift']
        employee_totals[key]['to_pay'] += payment['to_pay']
        employee_totals[key]['debt'] += payment['debt']
        employee_totals[key]['debt_nal'] += payment['debt_nal']
    
    # Создаём лист ИТОГО
    ws_itogo = wb.create_sheet(title="ИТОГО")
    club_names = ', '.join(clubs)
    ws_itogo['A1'] = f"Отчёт ЗП: {club_names}"
    ws_itogo['A1'].font = Font(bold=True, size=14)
    ws_itogo['A2'] = f"Период: {date_from} .. {date_to}"
    ws_itogo['A2'].font = Font(size=11)
    
    row_num = 4
    
    # Шапка таблицы
    for col, header in enumerate(headers, 1):
        cell = ws_itogo.cell(row=row_num, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    row_num += 1
    
    # Итоговые суммы для листа ИТОГО
    grand_totals = {
        'stavka': 0, 'lm_3': 0, 'percent_5': 0, 'promo': 0,
        'crz': 0, 'cons': 0, 'tips': 0, 'total_shift': 0,
        'to_pay': 0, 'debt': 0, 'debt_nal': 0, 'final_pay': 0
    }
    
    # Данные по сотрудникам
    for key in sorted(employee_totals.keys()):
        emp = employee_totals[key]
        
        # Рассчитываем 10% и к выплате
        vychet_10 = round(emp['debt'] * 0.1)  # Округление до целого
        k_vyplate = round(emp['debt_nal'] + emp['debt'] - vychet_10)
        
        row_data = [
            '',  # Дата пустая в ИТОГО
            '',  # Клуб пустой в ИТОГО
            emp['code'],
            emp['name'],
            emp['stavka'],
            emp['lm_3'],
            emp['percent_5'],
            emp['promo'],
            emp['crz'],
            emp['cons'],
            emp['tips'],
            emp['total_shift'],
            emp['to_pay'],
            emp['debt'],
            vychet_10,
            emp['debt_nal'],
            k_vyplate
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws_itogo.cell(row=row_num, column=col, value=value)
            cell.border = border
            if col > 4:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Обновляем итоги
        grand_totals['stavka'] += emp['stavka']
        grand_totals['lm_3'] += emp['lm_3']
        grand_totals['percent_5'] += emp['percent_5']
        grand_totals['promo'] += emp['promo']
        grand_totals['crz'] += emp['crz']
        grand_totals['cons'] += emp['cons']
        grand_totals['tips'] += emp['tips']
        grand_totals['total_shift'] += emp['total_shift']
        grand_totals['to_pay'] += emp['to_pay']
        grand_totals['debt'] += emp['debt']
        grand_totals['debt_nal'] += emp['debt_nal']
        grand_totals['final_pay'] += k_vyplate
        
        row_num += 1
    
    # Строка ИТОГО в листе ИТОГО
    vychet_10_grand = round(grand_totals['debt'] * 0.1)  # Округление до целого
    
    itogo_data = [
        'ИТОГО', '', '', '',
        grand_totals['stavka'],
        grand_totals['lm_3'],
        grand_totals['percent_5'],
        grand_totals['promo'],
        grand_totals['crz'],
        grand_totals['cons'],
        grand_totals['tips'],
        grand_totals['total_shift'],
        grand_totals['to_pay'],
        grand_totals['debt'],
        vychet_10_grand,
        grand_totals['debt_nal'],
        round(grand_totals['final_pay'])  # Округление до целого
    ]
    
    for col, value in enumerate(itogo_data, 1):
        cell = ws_itogo.cell(row=row_num, column=col, value=value)
        cell.font = Font(bold=True)
        cell.border = border
        if col > 4:
            cell.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Автоподгонка ширины для листа ИТОГО
    for column in ws_itogo.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_itogo.column_dimensions[column_letter].width = min(max_length + 2, 20)
    
    # Сохраняем и отправляем
    club_str = '_'.join([c.lower() for c in clubs])
    filename = f"zp_{club_str}_{date_from}_{date_to}.xlsx"
    wb.save(filename)
    
    with open(filename, 'rb') as f:
        await update.message.reply_document(
            document=f,
            filename=filename,
            caption=f"💵 Отчёт ЗП: {club_names}\nПериод: {date_from} .. {date_to}"
        )
    
    import os
    os.remove(filename)


async def handle_callback_query(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка нажатий на inline кнопки"""
    query = update.callback_query
    
    user_id = update.effective_user.id
    state = get_user_state(user_id)
    
    # Проверка авторизации
    if not db.is_admin(user_id) and not state.employee_mode and not state.limited_access:
        await query.answer("🔒 Доступ запрещён", show_alert=True)
        return
    
    await query.answer()
    
    # Кнопка "Выход" из быстрого доступа
    if query.data == 'quick_exit':
        # Полная очистка состояния - сбрасываем все поля
        state.__init__()
        state.limited_access = False
        await query.edit_message_text(
            "❌ Сессия завершена\n\n"
            "Для начала работы введите /start"
        )
        return
    
    # Выбор клуба при старте
    if query.data == 'club_moskvich':
        # Блокируем для ограниченного доступа
        if state.limited_access:
            await query.answer("❌ Доступ запрещён", show_alert=True)
            return
        
        # Проверяем режим - загрузка файла, загрузка ЗП или обычный старт
        if state.mode == 'awaiting_upload_club':
            state.upload_file_club = 'Москвич'
            await query.edit_message_text(
                f"📎 ЗАГРУЗКА ФАЙЛА\n"
                f"🏢 Клуб: Москвич\n\n"
                f"📅 Введите дату для этих данных:\n"
                f"Формат: 3,11 или 30,10"
            )
            state.mode = 'awaiting_upload_date'
        elif state.mode == 'awaiting_payments_upload_club':
            state.payments_upload_club = 'Москвич'
            await query.edit_message_text(
                f"💰 ЗАГРУЗКА ЗП\n"
                f"🏢 Клуб: Москвич\n\n"
                f"📅 Введите дату (формат: 30,10):"
            )
            state.mode = 'awaiting_payments_upload_date'
        else:
            state.club = 'Москвич'
            state.current_date = get_current_date()
            state.reset_input()
            
            await query.edit_message_text(
                f"✅ Выбран клуб: Москвич\n"
                f"📅 Дата: {state.current_date}"
            )
            await query.message.reply_text(
                "Используйте кнопки ниже для работы:",
                reply_markup=get_main_keyboard()
            )
    
    elif query.data == 'club_anora':
        # Блокируем для ограниченного доступа
        if state.limited_access:
            await query.answer("❌ Доступ запрещён", show_alert=True)
            return
        
        # Проверяем режим - загрузка файла, загрузка ЗП или обычный старт
        if state.mode == 'awaiting_upload_club':
            state.upload_file_club = 'Анора'
            await query.edit_message_text(
                f"📎 ЗАГРУЗКА ФАЙЛА\n"
                f"🏢 Клуб: Анора\n\n"
                f"📅 Введите дату для этих данных:\n"
                f"Формат: 3,11 или 30,10"
            )
            state.mode = 'awaiting_upload_date'
        elif state.mode == 'awaiting_payments_upload_club':
            state.payments_upload_club = 'Анора'
            await query.edit_message_text(
                f"💰 ЗАГРУЗКА ЗП\n"
                f"🏢 Клуб: Анора\n\n"
                f"📅 Введите дату (формат: 30,10):"
            )
            state.mode = 'awaiting_payments_upload_date'
        else:
            state.club = 'Анора'
            state.current_date = get_current_date()
            state.reset_input()
            
            await query.edit_message_text(
                f"✅ Выбран клуб: Анора\n"
                f"📅 Дата: {state.current_date}"
            )
            await query.message.reply_text(
                "Используйте кнопки ниже для работы:",
                reply_markup=get_main_keyboard()
            )
    
    # Выбор режима удаления
    elif query.data == 'delete_mode_employee':
        state.mode = 'awaiting_delete_employee_input'
        await query.edit_message_text(
            "✏️ Введите код и дату сотрудника для удаления:\n"
            "Пример: Д1 30,10"
        )
    elif query.data == 'delete_mode_mass':
        state.delete_mass_club = None
        state.delete_mass_date_from = None
        state.delete_mass_date_to = None
        state.delete_mass_preview = None
        state.mode = 'awaiting_delete_mass_club'
        await query.edit_message_text(
            "🏢 Выберите клуб для удаления:",
            reply_markup=get_club_report_keyboard()
        )
    
    # Меню управления сотрудниками
    elif query.data == 'employees_merge':
        # Старая логика - показываем выбор клуба для объединения
        await query.edit_message_text(
            "🔗 ОБЪЕДИНЕНИЕ СОТРУДНИКОВ\n\n"
            "Выберите клуб:",
            reply_markup=get_club_employees_keyboard()
        )
    
    elif query.data == 'employees_add':
        await query.edit_message_text(
            "➕ ДОБАВИТЬ СОТРУДНИКА\n\n"
            "Выберите клуб:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🏢 Москвич", callback_data='add_emp_club_moskvich')],
                [InlineKeyboardButton("🏢 Анора", callback_data='add_emp_club_anora')],
                [InlineKeyboardButton("❌ Назад", callback_data='employees_menu')]
            ])
        )
    
    elif query.data in ['add_emp_club_moskvich', 'add_emp_club_anora']:
        club = 'Москвич' if query.data == 'add_emp_club_moskvich' else 'Анора'
        
        await query.edit_message_text(
            f"➕ ДОБАВИТЬ СОТРУДНИКА\n"
            f"🏢 Клуб: {club}\n\n"
            f"Введите данные в формате:\n"
            f"КОД ИМЯ [ДАТА_НАЙМА]\n\n"
            f"📝 Примеры:\n"
            f"• Д1 Юлия\n"
            f"• Д7 Марина 15.03.2024\n"
            f"• СБ-Иван Петров\n\n"
            f"Дата найма по умолчанию = сегодня"
        )
        
        state.add_employee_club = club
        state.mode = 'awaiting_add_employee'
    
    elif query.data == 'employees_edit':
        await query.edit_message_text(
            "✏️ РЕДАКТИРОВАНИЕ СОТРУДНИКОВ\n\n"
            "Выберите клуб:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🏢 Москвич", callback_data='edit_club_moskvich')],
                [InlineKeyboardButton("🏢 Анора", callback_data='edit_club_anora')],
                [InlineKeyboardButton("❌ Назад", callback_data='employees_menu')]
            ])
        )
    
    elif query.data == 'employees_menu':
        # Возврат в главное меню сотрудников
        await query.edit_message_text(
            "👥 УПРАВЛЕНИЕ СОТРУДНИКАМИ\n\n"
            "Выберите действие:",
            reply_markup=get_employees_menu_keyboard()
        )
    
    elif query.data == 'employees_cancel':
        await query.edit_message_text("👥 Управление сотрудниками отменено")
    
    # === РЕДАКТИРОВАНИЕ СОТРУДНИКОВ ===
    
    elif query.data == 'emp_edit_code':
        await query.edit_message_text(
            f"🔢 ИЗМЕНИТЬ КОД\n\n"
            f"⚠️ ВНИМАНИЕ! Изменение кода затронет:\n"
            f"• Все записи в таблице operations\n"
            f"• Все записи в таблице payments\n"
            f"• Историю объединений\n\n"
            f"Текущий код: {state.edit_employee_selected['code']}\n\n"
            f"Введите новый код (например: Д7, СБ-Иван Петров):"
        )
        state.mode = 'awaiting_emp_code'
    
    elif query.data == 'emp_edit_name':
        await query.edit_message_text(
            f"✏️ ИЗМЕНИТЬ ИМЯ\n\n"
            f"Текущее имя: {state.edit_employee_selected['name']}\n\n"
            f"Введите новое имя:"
        )
        state.mode = 'awaiting_emp_name'
    
    elif query.data == 'emp_edit_phone':
        await query.edit_message_text(
            f"📱 ИЗМЕНИТЬ ТЕЛЕФОН\n\n"
            f"Текущий: {state.edit_employee_selected['phone'] or 'не указан'}\n\n"
            f"Введите новый телефон (или 'удалить' для удаления):"
        )
        state.mode = 'awaiting_emp_phone'
    
    elif query.data == 'emp_edit_tg':
        current_tg = state.edit_employee_selected['telegram_user_id']
        await query.edit_message_text(
            f"🔐 ИЗМЕНИТЬ TELEGRAM ID\n\n"
            f"Текущий: {current_tg or 'не указан'}\n\n"
            f"Введите новый Telegram User ID (или 'удалить' для удаления):"
        )
        state.mode = 'awaiting_emp_tg'
    
    elif query.data == 'emp_remove_tg':
        emp = state.edit_employee_selected
        
        # Удаляем TG ID
        conn = db.get_connection()
        cursor = conn.cursor()
        
        from datetime import datetime
        cursor.execute("""
            UPDATE employees
            SET telegram_user_id = NULL, updated_at = ?
            WHERE code = ? AND club = ?
        """, (datetime.now().isoformat(), emp['code'], state.edit_employees_club))
        
        conn.commit()
        conn.close()
        
        await query.edit_message_text(
            f"✅ ДОСТУП УДАЛЁН\n\n"
            f"Код: {emp['code']}\n"
            f"Имя: {emp['name']}\n"
            f"Telegram ID: удалён\n\n"
            f"Доступ к боту отключён."
        )
        
        state.edit_employee_selected = None
    
    elif query.data == 'emp_edit_birth':
        await query.edit_message_text(
            f"🎂 ДАТА РОЖДЕНИЯ\n\n"
            f"Введите дату в формате ДД.ММ.ГГГГ\n"
            f"Пример: 15.03.1998\n\n"
            f"Или 'удалить' для удаления:"
        )
        state.mode = 'awaiting_emp_birth'
    
    elif query.data == 'emp_fire':
        await query.edit_message_text(
            f"🚫 УВОЛИТЬ СОТРУДНИКА\n\n"
            f"Сотрудник: {state.edit_employee_selected['code']} - {state.edit_employee_selected['name']}\n"
            f"Клуб: {state.edit_employees_club}\n\n"
            f"⚠️ После увольнения:\n"
            f"• Статус → Уволен\n"
            f"• Доступ к боту → Отключён\n"
            f"• Дата увольнения → Сегодня\n\n"
            f"Подтвердите увольнение:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ УВОЛИТЬ", callback_data='emp_fire_confirm')],
                [InlineKeyboardButton("❌ Отмена", callback_data='emp_fire_cancel')]
            ])
        )
    
    elif query.data == 'emp_fire_confirm':
        emp = state.edit_employee_selected
        
        # Обновляем в БД
        conn = db.get_connection()
        cursor = conn.cursor()
        
        from datetime import datetime
        today = datetime.now().strftime('%Y-%m-%d')
        
        cursor.execute("""
            UPDATE employees
            SET is_active = 0, fired_date = ?, updated_at = ?
            WHERE code = ? AND club = ?
        """, (today, datetime.now().isoformat(), emp['code'], state.edit_employees_club))
        
        conn.commit()
        conn.close()
        
        await query.edit_message_text(
            f"✅ СОТРУДНИК УВОЛЕН\n\n"
            f"Код: {emp['code']}\n"
            f"Имя: {emp['name']}\n"
            f"Клуб: {state.edit_employees_club}\n"
            f"Дата увольнения: {today}\n\n"
            f"Доступ к боту отключён."
        )
        
        state.edit_employee_selected = None
    
    elif query.data == 'emp_fire_cancel':
        await query.answer("Отменено")
        # Возвращаемся к карточке
        await query.message.reply_text("Используйте кнопки меню для продолжения")
        state.edit_employee_selected = None
    
    elif query.data == 'emp_restore':
        emp = state.edit_employee_selected
        
        # Обновляем в БД
        conn = db.get_connection()
        cursor = conn.cursor()
        
        from datetime import datetime
        cursor.execute("""
            UPDATE employees
            SET is_active = 1, fired_date = NULL, updated_at = ?
            WHERE code = ? AND club = ?
        """, (datetime.now().isoformat(), emp['code'], state.edit_employees_club))
        
        conn.commit()
        conn.close()
        
        await query.edit_message_text(
            f"✅ СОТРУДНИК ВОЗВРАЩЁН\n\n"
            f"Код: {emp['code']}\n"
            f"Имя: {emp['name']}\n"
            f"Клуб: {state.edit_employees_club}\n\n"
            f"Статус: Действующий"
        )
        
        state.edit_employee_selected = None
    
    elif query.data == 'emp_view':
        # Просто показываем информацию (уже показана в карточке)
        await query.answer("Информация отображена выше")
    
    elif query.data == 'emp_edit_cancel':
        await query.edit_message_text("❌ Редактирование отменено")
        state.edit_employee_selected = None
        state.edit_employees_list = None
        state.edit_employees_club = None
    
    # === ОБРАБОТКА ЗАГРУЗКИ ВЫПЛАТ ===
    
    elif query.data == 'payments_save_confirm':
        # Берём данные из state
        if not state.payments_preview_data:
            await query.edit_message_text("❌ Данные не найдены")
            state.payments_upload_club = None
            state.payments_upload_date = None
            state.payments_preview_data = None
            return
        
        await query.edit_message_text("⏳ Сохраняю данные в базу...")
        
        # СНАЧАЛА УДАЛЯЕМ ВСЕ СТАРЫЕ ЗАПИСИ ДЛЯ ЭТОЙ ДАТЫ И КЛУБА
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            DELETE FROM payments 
            WHERE club = ? AND date = ?
        """, (state.payments_upload_club, state.payments_upload_date))
        conn.commit()
        conn.close()
        
        print(f"DEBUG: Deleted old payments for {state.payments_upload_club} {state.payments_upload_date}")
        
        # ПОТОМ ВСТАВЛЯЕМ НОВЫЕ
        saved_count = 0
        for payment in state.payments_preview_data:
            db.add_payment(
                club=state.payments_upload_club,
                date=state.payments_upload_date,
                code=payment['code'],
                name=payment['name'],
                stavka=payment['stavka'],
                lm_3=payment['lm_3'],
                percent_5=payment['percent_5'],
                promo=payment['promo'],
                crz=payment['crz'],
                cons=payment['cons'],
                tips=payment['tips'],
                fines=payment['fines'],
                total_shift=payment['total_shift'],
                debt=payment['debt'],
                debt_nal=payment['debt_nal'],
                to_pay=payment['to_pay']
            )
            saved_count += 1
        
        # DEBUG: Проверяем что сохранилось
        db.debug_payments(state.payments_upload_club, state.payments_upload_date)
        
        await query.edit_message_text(
            f"✅ ДАННЫЕ СОХРАНЕНЫ!\n\n"
            f"🏢 Клуб: {state.payments_upload_club}\n"
            f"📅 Дата: {state.payments_upload_date}\n"
            f"📊 Записей: {saved_count}\n\n"
            f"Данные можно просмотреть через кнопку ЗП"
        )
        
        # Очищаем состояние
        state.payments_upload_club = None
        state.payments_upload_date = None
        state.payments_preview_data = None
        state.payments_name_changes = None
    
    elif query.data == 'payments_save_cancel':
        await query.edit_message_text("❌ Загрузка отменена")
        state.payments_upload_club = None
        state.payments_upload_date = None
        state.payments_preview_data = None
        state.payments_name_changes = None
    
    elif query.data in ['edit_club_moskvich', 'edit_club_anora']:
        club = 'Москвич' if query.data == 'edit_club_moskvich' else 'Анора'
        
        await query.edit_message_text(f"⏳ Формирую список сотрудников {club}...")
        
        # Получаем сотрудников из НОВОЙ таблицы employees
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT code, full_name, telegram_user_id, phone, is_active
            FROM employees
            WHERE club = ?
            ORDER BY is_active DESC, code
        """, (club,))
        
        employees = cursor.fetchall()
        conn.close()
        
        if not employees:
            await query.message.reply_text(f"❌ Нет сотрудников в клубе {club}")
            return
        
        # Формируем файл
        lines = [f"СОТРУДНИКИ КЛУБА {club.upper()}\n"]
        lines.append("=" * 60 + "\n\n")
        
        active_employees = []
        fired_employees = []
        
        for code, name, tg_id, phone, is_active in employees:
            emp_dict = {
                'code': code,
                'name': name,
                'telegram_user_id': tg_id,
                'phone': phone,
                'is_active': is_active
            }
            
            if is_active:
                active_employees.append(emp_dict)
            else:
                fired_employees.append(emp_dict)
        
        # Действующие
        lines.append("✅ ДЕЙСТВУЮЩИЕ:\n\n")
        for i, emp in enumerate(active_employees, 1):
            access_icon = "🔐" if emp['telegram_user_id'] else "❌"
            phone_info = f" 📱{emp['phone']}" if emp['phone'] else ""
            lines.append(f"{i}. {emp['code']} - {emp['name']} {access_icon}{phone_info}\n")
        
        # Уволенные
        if fired_employees:
            lines.append(f"\n🗂️ УВОЛЕННЫЕ:\n\n")
            offset = len(active_employees)
            for i, emp in enumerate(fired_employees, offset + 1):
                lines.append(f"{i}. {emp['code']} - {emp['name']}\n")
        
        lines.append("\n" + "=" * 60 + "\n")
        lines.append(f"Всего: {len(employees)} | Действующих: {len(active_employees)} | Уволенных: {len(fired_employees)}")
        
        # Сохраняем во временный файл
        import tempfile
        temp_file = tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', suffix='.txt', delete=False)
        temp_file.write(''.join(lines))
        temp_file.close()
        
        # Отправляем файл
        with open(temp_file.name, 'rb') as f:
            await query.message.reply_document(
                document=f,
                filename=f"sotrudniki_{club.lower()}_edit.txt",
                caption=f"✏️ Список сотрудников клуба {club}\n\n🔐 = есть доступ к боту"
            )
        
        # Инструкция
        await query.message.reply_text(
            "✏️ ДЛЯ РЕДАКТИРОВАНИЯ:\n\n"
            "Отправьте номер сотрудника из списка\n\n"
            "📝 Пример: 5\n\n"
            "❌ Для отмены: отмена"
        )
        
        # Сохраняем список в state
        state.edit_employees_list = active_employees + fired_employees
        state.edit_employees_club = club
        state.mode = 'awaiting_employee_edit_select'
        
        # Удаляем временный файл
        import os
        os.remove(temp_file.name)
    
    # Выбор клуба для управления доступами
    # Выбор клуба для списка сотрудников
    elif query.data in ['employees_club_moskvich', 'employees_club_anora']:
        club = 'Москвич' if query.data == 'employees_club_moskvich' else 'Анора'
        await query.edit_message_text(f"👥 Формирую список сотрудников клуба {club}...")
        
        # Получаем уникальные пары (код, имя) из БД
        employees = db.get_all_employees(club)
        
        if not employees:
            await query.message.reply_text(f"❌ Нет сотрудников в клубе {club}")
            return
        
        # Сортируем по коду, потом по имени
        employees_sorted = sorted(employees, key=lambda x: (x['code'], x['name']))
        
        # Формируем текстовый файл
        lines = [f"СОТРУДНИКИ КЛУБА {club.upper()}\n"]
        lines.append("=" * 50 + "\n\n")
        
        for i, emp in enumerate(employees_sorted, 1):
            lines.append(f"{i}. {emp['code']} - {emp['name']}\n")
        
        lines.append("\n" + "=" * 50 + "\n")
        lines.append(f"Всего сотрудников: {len(employees_sorted)}")
        
        # Сохраняем во временный файл
        import tempfile
        temp_file = tempfile.NamedTemporaryFile(mode='w', encoding='utf-8', suffix='.txt', delete=False)
        temp_file.write(''.join(lines))
        temp_file.close()
        
        # Отправляем файл
        with open(temp_file.name, 'rb') as f:
            await query.message.reply_document(
                document=f,
                filename=f"sotrudniki_{club.lower()}.txt",
                caption=f"👥 Список сотрудников клуба {club}\nВсего: {len(employees_sorted)}"
            )
        
        # Инструкция по объединению
        await query.message.reply_text(
            "🔗 Для объединения сотрудников:\n\n"
            "Отправьте номера через тире или запятую\n\n"
            "📝 Примеры:\n"
            "• 1-5 (объединить 1 и 5)\n"
            "• 3-7-30 (объединить 3, 7 и 30)\n"
            "• 2,4,6 (объединить 2, 4 и 6)\n\n"
            "⚠️ Первый в списке станет главным"
        )
        
        # Сохраняем список сотрудников в state для дальнейшего использования
        state.employees_list = employees_sorted
        state.employees_club = club
        state.mode = 'awaiting_merge_employees'
        
        # Удаляем временный файл
        import os
        os.remove(temp_file.name)
    
    # Обработка кнопок стилистов
    elif query.data == 'stylist_load':
        # Загрузка расходов на стилистов
        keyboard = [
            [InlineKeyboardButton("🏢 Москвич", callback_data='stylist_load_moskvich')],
            [InlineKeyboardButton("🏢 Анора", callback_data='stylist_load_anora')]
        ]
        await query.edit_message_text(
            "💄 ЗАГРУЗКА РАСХОДОВ НА СТИЛИСТОВ\n\n"
            "Выберите клуб:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    
    elif query.data in ['stylist_load_moskvich', 'stylist_load_anora']:
        club = 'Москвич' if query.data == 'stylist_load_moskvich' else 'Анора'
        state.stylist_club = club
        state.mode = 'awaiting_stylist_period'
        
        await query.edit_message_text(
            f"💄 ЗАГРУЗКА РАСХОДОВ НА СТИЛИСТОВ\n"
            f"🏢 Клуб: {club}\n\n"
            f"📅 Укажите период или дату для расходов:\n\n"
            f"Примеры:\n"
            f"• 14.12 или 14,12 (одна дата)\n"
            f"• 14.12-20.12 или 14,12-20,12 (период)\n\n"
            f"❌ Для отмены напишите: ОТМЕНА"
        )
    
    elif query.data == 'stylist_done':
        # Кнопка ГОТОВО для стилистов
        user_id = update.effective_user.id
        state = get_user_state(user_id)
        
        if state.mode != 'awaiting_stylist_data':
            await query.answer("⚠️ Режим ввода данных стилистов не активен", show_alert=True)
            return
        
        if not state.stylist_expenses or len(state.stylist_expenses) == 0:
            await query.answer("❌ Нет данных для сохранения!", show_alert=True)
            return
        
        # Проверяем, есть ли записи требующие уточнения имени
        needs_clarification = [
            exp for exp in state.stylist_expenses 
            if exp.get('needs_selection') or exp.get('needs_input')
        ]
        
        if needs_clarification:
            # Есть записи требующие уточнения - запускаем процесс
            state.stylist_clarification_queue = needs_clarification
            state.stylist_clarification_index = 0
            state.mode = 'awaiting_stylist_clarification'
            
            await query.answer()
            await query.message.reply_text("⏳ Требуется уточнение информации...")
            # Задаем первый вопрос
            await ask_next_clarification_query(query, state)
        else:
            # Все ОК, показываем предпросмотр
            await query.answer()
            await show_stylist_preview_query(query, state)
            state.mode = 'awaiting_stylist_confirm'
        return
    
    elif query.data == 'stylist_view':
        # Просмотр расходов на стилистов
        keyboard = [
            [InlineKeyboardButton("🏢 Москвич", callback_data='stylist_view_moskvich')],
            [InlineKeyboardButton("🏢 Анора", callback_data='stylist_view_anora')]
        ]
        await query.edit_message_text(
            "💄 ПРОСМОТР РАСХОДОВ НА СТИЛИСТОВ\n\n"
            "Выберите клуб:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    
    elif query.data in ['stylist_view_moskvich', 'stylist_view_anora']:
        club = 'Москвич' if query.data == 'stylist_view_moskvich' else 'Анора'
        await handle_stylist_view(query, club)
    
    elif query.data.startswith('stylist_show_'):
        # Показать детали периода: stylist_show_CLUB_FROM_TO
        parts = query.data.replace('stylist_show_', '').split('_')
        if len(parts) >= 3:
            club = parts[0]  # Москвич или Анора
            period_from = '_'.join(parts[1:-1])  # может быть 2024-12-14
            period_to = parts[-1]
            await show_stylist_period_details(query, club, period_from, period_to)
    
    elif query.data.startswith('stylist_del_ask_'):
        # Запрос на удаление: stylist_del_ask_CLUB_FROM_TO
        parts = query.data.replace('stylist_del_ask_', '').split('_')
        if len(parts) >= 3:
            club = parts[0]
            period_from = '_'.join(parts[1:-1])
            period_to = parts[-1]
            
            user_id = query.from_user.id
            state = get_user_state(user_id)
            state.stylist_view_club = club
            state.stylist_view_from = period_from
            state.stylist_view_to = period_to
            state.mode = 'awaiting_stylist_view_delete'
            
            await query.edit_message_text(
                "🗑️ УДАЛЕНИЕ ЗАПИСЕЙ\n\n"
                "Введите номера записей для удаления:\n\n"
                "Примеры:\n"
                "• 3\n"
                "• 1 5 8\n"
                "• 1-5\n\n"
                "Или ОТМЕНА"
            )
    
    elif query.data.startswith('stylist_edit_ask_'):
        # Запрос на редактирование: stylist_edit_ask_CLUB_FROM_TO
        parts = query.data.replace('stylist_edit_ask_', '').split('_')
        if len(parts) >= 3:
            club = parts[0]
            period_from = '_'.join(parts[1:-1])
            period_to = parts[-1]
            
            user_id = query.from_user.id
            state = get_user_state(user_id)
            state.stylist_view_club = club
            state.stylist_view_from = period_from
            state.stylist_view_to = period_to
            state.mode = 'awaiting_stylist_view_edit'
            
            await query.edit_message_text(
                "✏️ РЕДАКТИРОВАНИЕ ЗАПИСИ\n\n"
                "Введите номер записи для редактирования:\n\n"
                "Пример: 4\n\n"
                "Или ОТМЕНА"
            )
    
    elif query.data.startswith('stylist_delete_'):
        # Удаление периода расходов: stylist_delete_CLUB_FROM_TO
        parts = query.data.replace('stylist_delete_', '').split('_')
        if len(parts) >= 3:
            club = parts[0]  # moskvich или anora
            club_name = 'Москвич' if club == 'moskvich' else 'Анора'
            period_from = parts[1]
            period_to = parts[2]
            
            keyboard = [
                [InlineKeyboardButton("✅ Да, удалить", callback_data=f'stylist_delete_confirm_{club}_{period_from}_{period_to}')],
                [InlineKeyboardButton("❌ Отмена", callback_data='stylist_view')]
            ]
            await query.edit_message_text(
                f"⚠️ ПОДТВЕРЖДЕНИЕ УДАЛЕНИЯ\n\n"
                f"Клуб: {club_name}\n"
                f"Период: {period_from} - {period_to}\n\n"
                f"Удалить эти расходы?",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
    
    elif query.data.startswith('stylist_delete_confirm_'):
        # Подтверждение удаления: stylist_delete_confirm_CLUB_FROM_TO
        parts = query.data.replace('stylist_delete_confirm_', '').split('_')
        if len(parts) >= 3:
            club = parts[0]
            club_name = 'Москвич' if club == 'moskvich' else 'Анора'
            period_from = parts[1]
            period_to = parts[2]
            
            deleted = db.delete_stylist_expenses_by_period(club_name, period_from, period_to)
            
            await query.edit_message_text(
                f"✅ РАСХОДЫ УДАЛЕНЫ\n\n"
                f"Клуб: {club_name}\n"
                f"Период: {period_from} - {period_to}\n"
                f"Удалено записей: {deleted}"
            )
    
    # Выбор клуба для отчёта / экспорта / списка
    elif query.data in ['report_club_moskvich', 'report_club_anora', 'report_club_both']:
        club_map = {
            'report_club_moskvich': 'москвич',
            'report_club_anora': 'анора',
            'report_club_both': 'оба'
        }
        
        # Определяем режим (отчёт, экспорт или список)
        if state.mode == 'awaiting_export_club':
            state.export_club = club_map[query.data]
            await query.edit_message_text(
                f"Экспорт: {state.export_club}\n\n"
                f"Укажите дату или период:\n"
                f"• Одна дата: 12,12\n"
                f"• Период: 10,06-11,08"
            )
            state.mode = 'awaiting_export_period'
        elif state.mode == 'awaiting_list_club':
            state.list_club = club_map[query.data]
            await query.edit_message_text(
                f"📋 Список: {state.list_club}\n\n"
                f"📅 Введите дату:\n"
                f"• 3,11\n"
                f"• 30,10"
            )
            state.mode = 'awaiting_list_date'
        elif state.mode == 'awaiting_delete_mass_club':
            state.delete_mass_club = club_map[query.data]
            state.delete_mass_date_from = None
            state.delete_mass_date_to = None
            state.delete_mass_preview = None
            await query.edit_message_text(
                f"Удаление ({state.delete_mass_club})\n\n"
                f"📅 Укажите дату или период для удаления:\n"
                f"• 5,11\n"
                f"• 2,11-5,11"
            )
            state.mode = 'awaiting_delete_mass_period'
        else:
            state.report_club = club_map[query.data]
            await query.edit_message_text(
                f"Клуб: {state.report_club}\n\n"
                f"Укажите дату или период:\n"
                f"• Одна дата: 12,12\n"
                f"• Период: 10,06-11,08"
            )
            state.mode = 'awaiting_report_period'
    
    # Выбор что удалить
    elif query.data in ['delete_nal', 'delete_beznal', 'delete_both']:
        channel_map = {
            'delete_nal': 'нал',
            'delete_beznal': 'безнал',
            'delete_both': 'обе'
        }
        choice = channel_map[query.data]
        
        await query.edit_message_text(f"Удаление: {choice.upper()}...")
        
        # Обработка удаления
        if choice in ['нал', 'безнал']:
            if choice in state.delete_records:
                db.delete_operation(state.club, state.delete_date, state.delete_code, choice)
                await query.message.reply_text(
                    f"✅ Удалено: {state.delete_code} {choice.upper()} за {state.delete_date}"
                )
            else:
                await query.message.reply_text(f"❌ Записи {choice.upper()} нет")
        
        elif choice == 'обе':
            deleted = []
            for channel in ['нал', 'безнал']:
                if channel in state.delete_records:
                    db.delete_operation(state.club, state.delete_date, state.delete_code, channel)
                    deleted.append(channel.upper())
            
            if deleted:
                await query.message.reply_text(
                    f"✅ Удалено: {state.delete_code} {', '.join(deleted)} за {state.delete_date}"
                )
            else:
                await query.message.reply_text("❌ Нет записей для удаления")
        
        state.mode = None

    elif query.data == 'delete_mass_confirm_yes':
        await query.edit_message_reply_markup(None)
        await handle_delete_mass_confirm_message(query.message, state, True)
    elif query.data == 'delete_mass_confirm_no':
        await query.edit_message_reply_markup(None)
        await handle_delete_mass_confirm_message(query.message, state, False)
    
    # Подтверждение замены объединённых сотрудников при загрузке файла
    elif query.data == 'upload_merge_yes':
        await query.edit_message_reply_markup(None)
        await query.edit_message_text(
            query.message.text + "\n\n✅ Применяю замены..."
        )
        # Устанавливаем флаг применения замен
        state.upload_file_data['apply_employee_merges'] = True
        state.upload_file_data['merge_check_done'] = True
        # Продолжаем сохранение
        await save_file_data_continue(query.message, state)
    
    elif query.data == 'upload_merge_no':
        await query.edit_message_reply_markup(None)
        await query.edit_message_text(
            query.message.text + "\n\n❌ Сохраняю как в файле..."
        )
        # НЕ применяем замены
        state.upload_file_data['apply_employee_merges'] = False
        state.upload_file_data['merge_check_done'] = True
        # Продолжаем сохранение
        await save_file_data_continue(query.message, state)
    
    # Подтверждение объединения сотрудников
    elif query.data == 'merge_employees_confirm':
        print(f"DEBUG: merge_employees_confirm callback triggered")
        print(f"DEBUG: state.merge_employee_indices={state.merge_employee_indices}")
        print(f"DEBUG: state.employees_list count={len(state.employees_list) if state.employees_list else 0}")
        print(f"DEBUG: state.employees_club={state.employees_club}")
        
        await query.edit_message_reply_markup(None)
        await handle_merge_employees_confirm(query.message, state)
    
    elif query.data == 'merge_employees_edit':
        await query.edit_message_text(
            "✏️ Введите номера заново\n\n"
            "📝 Примеры: 1-5, 3-7-30, 2,4,6"
        )
        state.mode = 'awaiting_merge_employees'
        state.merge_employee_indices = None
    
    elif query.data == 'merge_employees_cancel':
        await query.edit_message_text("❌ Объединение отменено")
        state.mode = None
        state.merge_employee_indices = None
    
    # Управление самозанятыми
    elif query.data == 'self_employed_add':
        await query.edit_message_text("➕ Введите код для добавления в самозанятые:\n\nПример: Д7")
        state.mode = 'awaiting_self_employed_add'
    elif query.data == 'self_employed_remove':
        await query.edit_message_text("➖ Введите код для удаления из самозанятых:\n\nПример: Д7")
        state.mode = 'awaiting_self_employed_remove'
    elif query.data == 'self_employed_close':
        await query.edit_message_text("✅ Закрыто")
        state.mode = None
    
    # Обработка объединения совпадений
    elif query.data == 'merge_all':
        if state.mode == 'awaiting_merge_confirm' and state.merge_candidates:
            await query.edit_message_reply_markup(None)
            await handle_merge_confirmation(update, state, 'ок', message=query.message)
        elif state.mode == 'awaiting_sb_merge_confirm' and state.sb_merge_data:
            await query.edit_message_reply_markup(None)
            await handle_sb_merge_confirmation(update, context, state, 'ок', 'ок', message=query.message)
        else:
            await query.answer("❌ Ошибка: данные не найдены", show_alert=True)
    
    elif query.data == 'merge_none':
        if state.mode == 'awaiting_merge_confirm' and state.merge_candidates:
            # Формируем команду "не" со всеми номерами
            all_numbers = ' '.join(str(i+1) for i in range(len(state.merge_candidates)))
            await query.edit_message_reply_markup(None)
            await handle_merge_confirmation(update, state, f'не {all_numbers}', message=query.message)
        elif state.mode == 'awaiting_sb_merge_confirm' and state.sb_merge_data:
            # Формируем команду "не" со всеми номерами
            sb_duplicates = state.sb_merge_data['sb_duplicates']
            all_numbers = ' '.join(str(i+1) for i in range(len(sb_duplicates)))
            await query.edit_message_reply_markup(None)
            await handle_sb_merge_confirmation(update, context, state, f'не {all_numbers}', f'не {all_numbers}', message=query.message)
        else:
            await query.answer("❌ Ошибка: данные не найдены", show_alert=True)
    
    elif query.data == 'merge_show_list':
        if state.mode == 'awaiting_merge_confirm' and state.merge_candidates:
            # Показываем список частями (по 15 записей)
            await query.answer("📄 Отправляю список...")
            candidates = state.merge_candidates
            chunk_size = 15
            
            for chunk_start in range(0, len(candidates), chunk_size):
                chunk = candidates[chunk_start:chunk_start + chunk_size]
                response = [f"📋 Совпадения ({chunk_start + 1}-{min(chunk_start + chunk_size, len(candidates))} из {len(candidates)}):\n"]
                
                for i, candidate in enumerate(chunk, chunk_start + 1):
                    response.append(f"{i}. {candidate['name']} {candidate['code']}")
                    response.append(f"   • Москвич: НАЛ {candidate['moskvich']['nal']:.0f}, БЕЗНАЛ {candidate['moskvich']['beznal']:.0f}")
                    response.append(f"   • Анора: НАЛ {candidate['anora']['nal']:.0f}, БЕЗНАЛ {candidate['anora']['beznal']:.0f}")
                    response.append("")
                
                await query.message.reply_text('\n'.join(response))
        elif state.mode == 'awaiting_sb_merge_confirm' and state.sb_merge_data:
            # Показываем список СБ частями
            await query.answer("📄 Отправляю список...")
            sb_duplicates = state.sb_merge_data['sb_duplicates']
            chunk_size = 15
            
            for chunk_start in range(0, len(sb_duplicates), chunk_size):
                chunk = sb_duplicates[chunk_start:chunk_start + chunk_size]
                response = [f"📋 СБ с похожими именами ({chunk_start + 1}-{min(chunk_start + chunk_size, len(sb_duplicates))} из {len(sb_duplicates)}):\n"]
                
                for i, group in enumerate(chunk, chunk_start + 1):
                    similarity_pct = int(group['similarity'] * 100)
                    response.append(f"{i}. Группа: {group['main_name']} (Похожесть: {similarity_pct}%)")
                    
                    # Группируем операции по именам
                    by_name = {}
                    for op in group['operations']:
                        name = op['name']
                        if name not in by_name:
                            by_name[name] = {'nal': 0, 'beznal': 0}
                        if op['channel'] == 'нал':
                            by_name[name]['nal'] += op['amount']
                        else:
                            by_name[name]['beznal'] += op['amount']
                    
                    for name in group['names']:
                        if name in by_name:
                            response.append(f"   • {name}: НАЛ {by_name[name]['nal']:.0f}, БЕЗНАЛ {by_name[name]['beznal']:.0f}")
                    response.append(f"   ИТОГО: НАЛ {group['total_nal']:.0f}, БЕЗНАЛ {group['total_beznal']:.0f}")
                    response.append("")
                
                await query.message.reply_text('\n'.join(response))
        else:
            await query.answer("❌ Ошибка: данные не найдены", show_alert=True)


def format_report_summary(totals: Dict, club_name: str, period: str, 
                         employee_count: int, merged_count: int = 0) -> str:
    """
    Форматирование краткой сводки отчёта
    totals: словарь с итогами
    club_name: название клуба или "СВОДНЫЙ"
    period: период отчёта
    employee_count: количество сотрудников
    merged_count: количество объединённых дубликатов (если есть)
    """
    lines = []
    lines.append("✅ ОТЧЁТ ГОТОВ!\n")
    lines.append(f"🏢 Клуб: {club_name}")
    lines.append(f"📅 Период: {period}")
    lines.append(f"👥 Сотрудников: {employee_count}")
    
    if merged_count > 0:
        lines.append(f"🔄 Объединено дубликатов: {merged_count}")
    
    lines.append("\n💰 ИТОГО:")
    lines.append(f"   НАЛ:      {totals['nal']:,.0f}".replace(',', ' '))
    lines.append(f"   БЕЗНАЛ:   {totals['beznal']:,.0f}".replace(',', ' '))
    lines.append(f"   10%:      {totals['minus10']:,.0f}".replace(',', ' '))
    lines.append(f"   {'─' * 25}")
    lines.append(f"   ИТОГО:    {totals['itog']:,.0f}".replace(',', ' '))
    lines.append("\n📄 Детальный отчёт в Excel файле ⬇️")
    
    return '\n'.join(lines)


async def handle_merge_employees_input(update: Update, state: UserState, text: str):
    """Обработка ввода номеров сотрудников для объединения"""
    if not state.employees_list:
        await update.message.reply_text("❌ Ошибка: список сотрудников не найден")
        state.mode = None
        return
    
    # Парсим номера (поддержка тире и запятых)
    text_normalized = text.replace(',', '-').replace(' ', '')
    parts = text_normalized.split('-')
    
    try:
        indices = [int(p) for p in parts if p.isdigit()]
    except:
        await update.message.reply_text("❌ Неверный формат. Используйте номера через тире или запятую\nПример: 1-5-8")
        return
    
    # Проверки
    if len(indices) < 2:
        await update.message.reply_text("❌ Нужно минимум 2 сотрудника для объединения")
        return
    
    if len(set(indices)) != len(indices):
        await update.message.reply_text("❌ Номера не должны повторяться")
        return
    
    max_index = len(state.employees_list)
    invalid = [i for i in indices if i < 1 or i > max_index]
    if invalid:
        await update.message.reply_text(f"❌ Неверные номера: {invalid}\nДоступны номера от 1 до {max_index}")
        return
    
    # Получаем сотрудников по индексам (индексы с 1, в массиве с 0)
    selected_employees = [state.employees_list[i-1] for i in indices]
    main_employee = selected_employees[0]
    
    # Формируем предпросмотр
    lines = ["📋 ОБЪЕДИНЕНИЕ СОТРУДНИКОВ\n"]
    lines.append(f"🏢 Клуб: {state.employees_club}\n")
    lines.append("Будут объединены:\n")
    
    for i, emp in enumerate(selected_employees):
        prefix = "← ГЛАВНЫЙ" if i == 0 else ""
        lines.append(f"{indices[i]}. {emp['code']} - {emp['name']} {prefix}\n")
    
    lines.append(f"\n⚠️ В БД все записи этих сотрудников получат:")
    lines.append(f"   КОД: {main_employee['code']}")
    lines.append(f"   ИМЯ: {main_employee['name']}\n")
    lines.append("✅ Это НАВСЕГДА изменит данные в БД!")
    
    # Сохраняем выбор
    state.merge_employee_indices = indices
    state.mode = 'awaiting_merge_employees_confirm'
    
    # Кнопка объединения
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔗 ОБЪЕДИНИТЬ", callback_data='merge_employees_confirm')],
        [InlineKeyboardButton("✏️ РЕДАКТИРОВАТЬ", callback_data='merge_employees_edit')],
        [InlineKeyboardButton("❌ ОТМЕНА", callback_data='merge_employees_cancel')]
    ])
    
    await update.message.reply_text(''.join(lines), reply_markup=keyboard)


async def handle_merge_employees_confirm(message, state: UserState):
    """Выполнение объединения сотрудников в БД"""
    print(f"DEBUG: handle_merge_employees_confirm called")
    print(f"DEBUG: merge_employee_indices={state.merge_employee_indices}")
    print(f"DEBUG: employees_list={state.employees_list}")
    
    if not state.merge_employee_indices or not state.employees_list:
        await message.reply_text("❌ Ошибка: данные не найдены")
        state.mode = None
        return
    
    # Получаем сотрудников
    selected_employees = [state.employees_list[i-1] for i in state.merge_employee_indices]
    main_employee = selected_employees[0]
    
    # Объединяем в БД
    updated_count = db.merge_employees(
        club=state.employees_club,
        main_code=main_employee['code'],
        main_name=main_employee['name'],
        employees_to_merge=selected_employees[1:]  # Все кроме главного
    )
    
    # Проверяем результат
    if updated_count == 0:
        await message.reply_text(
            "⚠️ ВНИМАНИЕ!\n\n"
            "Объединение записано, но не найдено записей в таблице operations для переноса.\n\n"
            "Возможные причины:\n"
            "• Эти сотрудники ещё не имеют записей в БД\n"
            "• Коды/имена не совпадают точно\n\n"
            "Проверьте логи для деталей."
        )
    else:
        # Формируем отчёт об успехе
        lines = ["✅ ОБЪЕДИНЕНИЕ ВЫПОЛНЕНО!\n"]
        lines.append(f"🏢 Клуб: {state.employees_club}\n")
        lines.append("Объединены:\n")
        
        for i, emp in enumerate(selected_employees[1:], 1):
            lines.append(f"• {emp['code']} - {emp['name']} → {main_employee['code']} - {main_employee['name']}\n")
        
        lines.append(f"\n📊 Обновлено записей в БД: {updated_count}")
        lines.append("\n\n✅ Теперь в отчётах эти сотрудники будут показываться как:")
        lines.append(f"   {main_employee['code']} - {main_employee['name']}")
        
        await message.reply_text(''.join(lines))
    
    # Очищаем состояние
    state.mode = None
    state.merge_employee_indices = None
    state.employees_list = None
    state.employees_club = None


async def handle_self_employed_command(update: Update, context: ContextTypes.DEFAULT_TYPE,
                                       state: UserState):
    """Команда управления самозанятыми"""
    codes = db.get_all_self_employed()
    
    if not codes:
        message = "📋 Список самозанятых пуст."
    else:
        message = f"👔 САМОЗАНЯТЫЕ ({len(codes)} чел.):\n\n"
        message += ", ".join(codes)
    
    await update.message.reply_text(
        message,
        reply_markup=get_self_employed_action_keyboard()
    )


async def handle_stylist_data_input(update: Update, state: UserState, text: str, text_lower: str):
    """Обработка ввода данных о расходах на стилистов (накопление из нескольких сообщений)"""
    from parser import DataParser
    
    # Проверяем кнопку ГОТОВО
    if text_lower == 'готово' or text_lower == '✅ готово':
        if not state.stylist_expenses or len(state.stylist_expenses) == 0:
            await update.message.reply_text(
                "❌ Нет данных для сохранения!\n\n"
                "Отправьте данные о расходах или напишите: ОТМЕНА"
            )
            return
        
        # Проверяем, есть ли записи требующие уточнения имени
        needs_clarification = [
            exp for exp in state.stylist_expenses 
            if exp.get('needs_selection') or exp.get('needs_input')
        ]
        
        if needs_clarification:
            # Есть записи требующие уточнения - запускаем процесс
            state.stylist_clarification_queue = needs_clarification
            state.stylist_clarification_index = 0
            state.mode = 'awaiting_stylist_clarification'
            
            # Задаем первый вопрос
            await ask_next_clarification(update, state)
        else:
            # Все ОК, показываем предпросмотр
            await show_stylist_preview(update, state)
            state.mode = 'awaiting_stylist_confirm'
        return
    
    # Парсим данные из текущего сообщения
    expenses, errors = DataParser.parse_stylist_expenses(text)
    
    if not expenses and not errors:
        await update.message.reply_text(
            "❌ Не найдено ни одной записи о расходах в этом сообщении!\n\n"
            "Проверьте формат данных.\n"
            "Формат: Д14Бритни 2000 или Д14 - 500\n\n"
            "После завершения нажмите: ГОТОВО"
        )
        return
    
    # Для каждого расхода БЕЗ имени ищем в БД
    for exp in expenses:
        if exp['name'] is None or exp['name'] == '':
            # Ищем имена в БД по клубу и коду
            names = db.get_employee_names_by_code(state.stylist_club, exp['code'])
            
            if len(names) == 1:
                # ✅ Одно имя - автоматически заполняем
                exp['name'] = names[0]
                exp['auto_filled'] = True
            
            elif len(names) > 1:
                # ⚠️ Несколько имен - нужен выбор
                exp['needs_selection'] = True
                exp['available_names'] = names
                exp['name'] = None  # Пока пусто
            
            else:  # len(names) == 0
                # ❓ Новый код - нужен ввод имени
                exp['needs_input'] = True
                exp['name'] = None  # Пока пусто
    
    # Добавляем распарсенные расходы к уже существующим
    if state.stylist_expenses is None:
        state.stylist_expenses = []
    if state.stylist_errors is None:
        state.stylist_errors = []
    
    state.stylist_expenses.extend(expenses)
    state.stylist_errors.extend(errors)
    
    # Подтверждаем добавление
    auto_filled_count = len([e for e in expenses if e.get('auto_filled')])
    needs_clarif_count = len([e for e in expenses if e.get('needs_selection') or e.get('needs_input')])
    
    msg = f"✅ Добавлено записей: {len(expenses)}\n"
    if auto_filled_count > 0:
        msg += f"   • Имена заполнены автоматически: {auto_filled_count}\n"
    if needs_clarif_count > 0:
        msg += f"   • Требуют уточнения: {needs_clarif_count}\n"
    msg += f"📝 Всего накоплено: {len(state.stylist_expenses)}\n"
    
    if errors:
        msg += f"\n⚠️ Ошибок в этом сообщении: {len(errors)}\n"
        for error in errors[:3]:
            msg += f"• {error}\n"
        if len(errors) > 3:
            msg += f"... и ещё {len(errors) - 3}\n"
    
    msg += "\n💬 Продолжайте отправлять данные или нажмите: ГОТОВО"
    
    await update.message.reply_text(msg)


async def show_stylist_preview(update: Update, state: UserState):
    """Показать предпросмотр расходов на стилистов с нумерацией"""
    # Проверяем какие коды не найдены в operations
    suspicious = []
    for i, exp in enumerate(state.stylist_expenses, 1):
        # Ищем код в operations для этого клуба
        ops = db.get_operations_by_period(
            state.stylist_club,
            state.stylist_period_from,
            state.stylist_period_to
        )
        codes_in_ops = set(op['code'] for op in ops)
        
        if exp['code'] not in codes_in_ops:
            # Ищем похожие коды
            similar = []
            exp_code_lower = exp['code'].lower()
            for code in codes_in_ops:
                if code.lower() in exp_code_lower or exp_code_lower in code.lower():
                    similar.append(code)
            
            suspicious.append({
                'index': i,
                'code': exp['code'],
                'name': exp['name'],
                'amount': exp['amount'],
                'similar': similar[:3]  # Максимум 3 похожих
            })
    
    preview = f"📎 ПРЕДПРОСМОТР РАСХОДОВ НА СТИЛИСТОВ\n\n"
    preview += f"🏢 Клуб: {state.stylist_club}\n"
    preview += f"📅 Период: {state.stylist_period_from} - {state.stylist_period_to}\n\n"
    
    # Если есть сомнительные - показываем предупреждение
    if suspicious:
        preview += "⚠️ ВНИМАНИЕ! Коды НЕ НАЙДЕНЫ в операциях:\n\n"
        for susp in suspicious[:5]:  # Показываем первые 5
            preview += f"{susp['index']}. {susp['code']} {susp['name']} {susp['amount']}₽\n"
            if susp['similar']:
                preview += f"   Похожие: {', '.join(susp['similar'])}\n"
        
        if len(suspicious) > 5:
            preview += f"... и ещё {len(suspicious) - 5}\n"
        
        preview += f"\nНе найдено: {len(suspicious)} из {len(state.stylist_expenses)}\n"
        preview += "-" * 45 + "\n\n"
    
    preview += f"№  | {'Код':<8} | {'Имя':<15} | Сумма\n"
    preview += "-" * 45 + "\n"
    
    total = 0
    for i, exp in enumerate(state.stylist_expenses[:50], 1):  # Показываем первые 50
        preview += f"{i:<2} | {exp['code']:<8} | {exp['name']:<15} | {exp['amount']}₽\n"
        total += exp['amount']
    
    if len(state.stylist_expenses) > 50:
        preview += f"... и ещё {len(state.stylist_expenses) - 50} записей\n"
        # Считаем полную сумму
        total = sum(exp['amount'] for exp in state.stylist_expenses)
    
    preview += "-" * 45 + "\n"
    preview += f"Всего: {len(state.stylist_expenses)} расходов на сумму {total}₽\n"
    
    if state.stylist_errors:
        preview += f"\n⚠️ Всего предупреждений: {len(state.stylist_errors)}\n"
        for error in state.stylist_errors[:3]:
            preview += f"• {error}\n"
        if len(state.stylist_errors) > 3:
            preview += f"... и ещё {len(state.stylist_errors) - 3}\n"
    
    preview += "\n✅ Что делать?\n"
    preview += "• ЗАПИСАТЬ - сохранить как есть\n"
    preview += "• ИСПРАВИТЬ [номер] - редактировать запись\n"
    preview += "• ОТМЕНА - отменить"
    
    await update.message.reply_text(preview)


async def handle_stylist_confirm(update: Update, state: UserState, text_lower: str):
    """Обработка подтверждения сохранения расходов на стилистов"""
    if text_lower == 'записать':
        # Сохраняем в БД
        success_count = 0
        for exp in state.stylist_expenses:
            success = db.add_stylist_expense(
                club=state.stylist_club,
                period_from=state.stylist_period_from,
                period_to=state.stylist_period_to,
                code=exp['code'],
                name=exp['name'],
                amount=exp['amount']
            )
            if success:
                success_count += 1
        
        total_amount = sum(exp['amount'] for exp in state.stylist_expenses)
        
        await update.message.reply_text(
            f"✅ РАСХОДЫ НА СТИЛИСТОВ СОХРАНЕНЫ!\n\n"
            f"🏢 Клуб: {state.stylist_club}\n"
            f"📅 Период: {state.stylist_period_from} - {state.stylist_period_to}\n"
            f"📝 Записей: {success_count}\n"
            f"💰 Итого: {total_amount}₽",
            reply_markup=get_main_keyboard()
        )
        
        # Очищаем state
        state.mode = None
        state.stylist_club = None
        state.stylist_period_from = None
        state.stylist_period_to = None
        state.stylist_expenses = None
        state.stylist_errors = None
        state.stylist_edit_index = None
    
    elif text_lower.startswith('исправить'):
        # Команда ИСПРАВИТЬ [номер]
        parts = text_lower.split()
        if len(parts) < 2:
            await update.message.reply_text(
                "❌ Укажите номер записи для редактирования.\n\n"
                "Пример: ИСПРАВИТЬ 3"
            )
            return
        
        try:
            index = int(parts[1]) - 1  # Преобразуем в 0-based индекс
            if index < 0 or index >= len(state.stylist_expenses):
                await update.message.reply_text(
                    f"❌ Неверный номер записи. Должен быть от 1 до {len(state.stylist_expenses)}"
                )
                return
            
            # Сохраняем индекс для редактирования
            state.stylist_edit_index = index
            exp = state.stylist_expenses[index]
            
            await update.message.reply_text(
                f"✏️ РЕДАКТИРОВАНИЕ ЗАПИСИ №{index + 1}\n\n"
                f"Текущие данные:\n"
                f"Код: {exp['code']}\n"
                f"Имя: {exp['name']}\n"
                f"Сумма: {exp['amount']}₽\n\n"
                f"📝 Введите новые данные в формате:\n"
                f"КОД ИМЯ СУММА\n\n"
                f"Пример: Н3 Влада 3000\n\n"
                f"Или напишите: ОТМЕНА"
            )
            state.mode = 'awaiting_stylist_edit_data'
        except ValueError:
            await update.message.reply_text(
                "❌ Неверный формат номера.\n\n"
                "Пример: ИСПРАВИТЬ 3"
            )
    
    else:
        await update.message.reply_text(
            "❌ Неверная команда. Введите:\n"
            "• ЗАПИСАТЬ - сохранить\n"
            "• ИСПРАВИТЬ [номер] - редактировать\n"
            "• ОТМЕНА - отменить"
        )


async def handle_stylist_edit_data(update: Update, state: UserState, text: str):
    """Обработка ввода новых данных для редактирования расхода на стилиста"""
    from parser import DataParser
    
    # Парсим одну строку
    expenses, errors = DataParser.parse_stylist_expenses(text)
    
    if not expenses or len(expenses) == 0:
        await update.message.reply_text(
            "❌ Не удалось распарсить данные!\n\n"
            "Формат: КОД ИМЯ СУММА\n"
            "Пример: Н3 Влада 3000\n\n"
            "Или напишите: ОТМЕНА"
        )
        return
    
    if len(expenses) > 1:
        await update.message.reply_text(
            "⚠️ Обнаружено несколько записей, будет использована только первая.\n\n"
            "Формат: КОД ИМЯ СУММА (одна строка)\n"
            "Пример: Н3 Влада 3000"
        )
    
    # Обновляем запись
    new_expense = expenses[0]
    state.stylist_expenses[state.stylist_edit_index] = new_expense
    
    await update.message.reply_text("✅ Запись обновлена!\n")
    
    # Показываем обновленный предпросмотр
    await show_stylist_preview(update, state)
    state.mode = 'awaiting_stylist_confirm'
    state.stylist_edit_index = None


async def ask_next_clarification(update: Update, state: UserState):
    """Задать следующий вопрос для уточнения имени сотрудника"""
    exp = state.stylist_clarification_queue[state.stylist_clarification_index]
    
    if exp.get('needs_selection'):
        # Несколько имен в БД - нужен выбор
        msg = f"⚠️ КОД {exp['code']} ИМЕЕТ НЕСКОЛЬКО ИМЕН В БАЗЕ:\n\n"
        for i, name in enumerate(exp['available_names'], 1):
            msg += f"{i}. {name}\n"
        msg += f"\nПод каким именем сохранить расход {exp['amount']}₽?\n"
        msg += "Введите номер:"
        await update.message.reply_text(msg)
    
    elif exp.get('needs_input'):
        # Новый код - нужен ввод имени
        msg = f"❓ КОД {exp['code']} НЕ НАЙДЕН В БАЗЕ\n\n"
        msg += f"Это новый сотрудник?\n"
        msg += f"Введите имя для кода {exp['code']}:"
        await update.message.reply_text(msg)


async def ask_next_clarification_query(query, state: UserState):
    """Задать следующий вопрос для уточнения имени сотрудника (через query)"""
    exp = state.stylist_clarification_queue[state.stylist_clarification_index]
    
    if exp.get('needs_selection'):
        # Несколько имен в БД - нужен выбор
        msg = f"⚠️ КОД {exp['code']} ИМЕЕТ НЕСКОЛЬКО ИМЕН В БАЗЕ:\n\n"
        for i, name in enumerate(exp['available_names'], 1):
            msg += f"{i}. {name}\n"
        msg += f"\nПод каким именем сохранить расход {exp['amount']}₽?\n"
        msg += "Введите номер:"
        await query.message.reply_text(msg)
    
    elif exp.get('needs_input'):
        # Новый код - нужен ввод имени
        msg = f"❓ КОД {exp['code']} НЕ НАЙДЕН В БАЗЕ\n\n"
        msg += f"Это новый сотрудник?\n"
        msg += f"Введите имя для кода {exp['code']}:"
        await query.message.reply_text(msg)


async def show_stylist_preview_query(query, state: UserState):
    """Показать предпросмотр расходов на стилистов с нумерацией (через query)"""
    preview = f"📎 ПРЕДПРОСМОТР РАСХОДОВ НА СТИЛИСТОВ\n\n"
    preview += f"🏢 Клуб: {state.stylist_club}\n"
    preview += f"📅 Период: {state.stylist_period_from} - {state.stylist_period_to}\n\n"
    preview += f"№  | {'Код':<8} | {'Имя':<15} | Сумма\n"
    preview += "-" * 45 + "\n"
    
    total = 0
    for i, exp in enumerate(state.stylist_expenses[:50], 1):  # Показываем первые 50
        preview += f"{i:<2} | {exp['code']:<8} | {exp['name']:<15} | {exp['amount']}₽\n"
        total += exp['amount']
    
    if len(state.stylist_expenses) > 50:
        preview += f"... и ещё {len(state.stylist_expenses) - 50} записей\n"
        # Считаем полную сумму
        total = sum(exp['amount'] for exp in state.stylist_expenses)
    
    preview += "-" * 45 + "\n"
    preview += f"Всего: {len(state.stylist_expenses)} расходов на сумму {total}₽\n"
    
    if state.stylist_errors:
        preview += f"\n⚠️ Всего предупреждений: {len(state.stylist_errors)}\n"
        for error in state.stylist_errors[:3]:
            preview += f"• {error}\n"
        if len(state.stylist_errors) > 3:
            preview += f"... и ещё {len(state.stylist_errors) - 3}\n"
    
    preview += "\n✅ Всё верно? Введите:\n"
    preview += "• ЗАПИСАТЬ - сохранить в базу\n"
    preview += "• ИСПРАВИТЬ [номер] - редактировать запись\n"
    preview += "• ОТМЕНА - отменить"
    
    await query.message.reply_text(preview)


async def handle_stylist_clarification(update: Update, state: UserState, text: str):
    """Обработка ответа пользователя на вопросы уточнения"""
    exp = state.stylist_clarification_queue[state.stylist_clarification_index]
    
    if exp.get('needs_selection'):
        # Пользователь выбрал номер из списка
        try:
            choice = int(text.strip()) - 1
            if choice < 0 or choice >= len(exp['available_names']):
                await update.message.reply_text(
                    f"❌ Неверный номер. Выберите от 1 до {len(exp['available_names'])}:"
                )
                return
            
            exp['name'] = exp['available_names'][choice]
            del exp['needs_selection']
            del exp['available_names']
            
        except ValueError:
            await update.message.reply_text("❌ Введите номер (например, 1 или 2):")
            return
    
    elif exp.get('needs_input'):
        # Пользователь ввел новое имя
        new_name = text.strip().capitalize()
        if not new_name or len(new_name) < 2:
            await update.message.reply_text("❌ Введите корректное имя:")
            return
        
        exp['name'] = new_name
        del exp['needs_input']
    
    # Переходим к следующему вопросу
    state.stylist_clarification_index += 1
    
    if state.stylist_clarification_index < len(state.stylist_clarification_queue):
        # Есть ещё вопросы
        await ask_next_clarification(update, state)
    else:
        # Все уточнения завершены
        await update.message.reply_text("✅ Все уточнения завершены!\n")
        
        state.stylist_clarification_queue = None
        state.stylist_clarification_index = None
        
        # Показываем предпросмотр
        await show_stylist_preview(update, state)
        state.mode = 'awaiting_stylist_confirm'


async def handle_stylist_view(query, club: str):
    """Просмотр загруженных расходов на стилистов"""
    periods = db.get_stylist_expenses_periods(club)
    
    if not periods:
        await query.edit_message_text(
            f"📋 РАСХОДЫ НА СТИЛИСТОВ\n\n"
            f"🏢 Клуб: {club}\n\n"
            f"❌ Нет загруженных расходов"
        )
        return
    
    # Если периодов несколько - показываем выбор
    if len(periods) > 1:
        message = f"📋 РАСХОДЫ НА СТИЛИСТОВ\n\n🏢 Клуб: {club}\n\n"
        message += "Выберите период:\n\n"
        
        keyboard = []
        for period in periods:
            period_from = period['period_from']
            period_to = period['period_to']
            count = period['count']
            total = period['total_amount']
            
            message += f"📅 {period_from} - {period_to}\n"
            message += f"   Записей: {count}, Сумма: {total}₽\n\n"
            
            # Кнопка для просмотра периода
            keyboard.append([
                InlineKeyboardButton(
                    f"📋 {period_from} - {period_to}",
                    callback_data=f'stylist_show_{club}_{period_from}_{period_to}'
                )
            ])
        
        await query.edit_message_text(
            message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        # Один период - сразу показываем детали
        period = periods[0]
        await show_stylist_period_details(query, club, period['period_from'], period['period_to'])


async def show_stylist_period_details(query_or_update, club: str, period_from: str, period_to: str):
    """Показать детали расходов стилистов за период"""
    expenses = db.get_stylist_expenses_by_period(club, period_from, period_to)
    
    if not expenses:
        message = f"📋 РАСХОДЫ НА СТИЛИСТОВ\n\n"
        message += f"🏢 Клуб: {club}\n"
        message += f"📅 Период: {period_from} - {period_to}\n\n"
        message += "❌ Нет записей"
        
        if hasattr(query_or_update, 'edit_message_text'):
            await query_or_update.edit_message_text(message)
        else:
            await query_or_update.message.reply_text(message)
        return
    
    message = f"📋 РАСХОДЫ НА СТИЛИСТОВ\n\n"
    message += f"🏢 Клуб: {club}\n"
    message += f"📅 Период: {period_from} - {period_to}\n\n"
    message += f"№  | {'Код':<8} | {'Имя':<15} | Сумма\n"
    message += "-" * 45 + "\n"
    
    total = 0
    for i, exp in enumerate(expenses[:50], 1):  # Первые 50
        message += f"{i:<2} | {exp['code']:<8} | {exp['name']:<15} | {exp['amount']}₽\n"
        total += exp['amount']
    
    if len(expenses) > 50:
        message += f"\n... и ещё {len(expenses) - 50} записей"
        total = sum(exp['amount'] for exp in expenses)
    
    message += "\n" + "-" * 45 + "\n"
    message += f"Всего: {len(expenses)} расходов на {total}₽\n\n"
    message += "Команды:\n"
    message += "• УДАЛИТЬ [номера] - удалить записи\n"
    message += "• ИСПРАВИТЬ [номер] - изменить запись"
    
    keyboard = [
        [
            InlineKeyboardButton("🗑️ Удалить записи", callback_data=f'stylist_del_ask_{club}_{period_from}_{period_to}'),
            InlineKeyboardButton("✏️ Исправить", callback_data=f'stylist_edit_ask_{club}_{period_from}_{period_to}')
        ],
        [InlineKeyboardButton("🔙 Назад", callback_data=f'stylist_view_{club.lower()}')]
    ]
    
    if hasattr(query_or_update, 'edit_message_text'):
        await query_or_update.edit_message_text(
            message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    else:
        await query_or_update.message.reply_text(
            message,
            reply_markup=InlineKeyboardMarkup(keyboard)
        )


async def handle_self_employed_add(update: Update, state: UserState, code: str):
    """Добавление кода в самозанятые"""
    from parser import DataParser
    
    code = code.strip()
    
    # Проверка формата кода
    if not DataParser.is_code(code):
        await update.message.reply_text(
            "❌ Неверный формат кода.\n"
            "Примеры: Д7, Р1, Б52, К21"
        )
        return
    
    # Нормализуем код
    normalized_code = DataParser.normalize_code(code)
    
    # Добавляем в БД
    success, message = db.add_self_employed(normalized_code)
    
    await update.message.reply_text(message)
    
    # Сбрасываем режим
    state.mode = None


async def handle_self_employed_remove(update: Update, state: UserState, code: str):
    """Удаление кода из самозанятых"""
    from parser import DataParser
    
    code = code.strip()
    
    # Проверка формата кода
    if not DataParser.is_code(code):
        await update.message.reply_text(
            "❌ Неверный формат кода.\n"
            "Примеры: Д7, Р1, Б52, К21"
        )
        return
    
    # Нормализуем код
    normalized_code = DataParser.normalize_code(code)
    
    # Удаляем из БД
    success, message = db.remove_self_employed(normalized_code)
    
    await update.message.reply_text(message)
    
    # Сбрасываем режим
    state.mode = None


async def restore_sb_names_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Команда для восстановления имен СБ из журнала"""
    user_id = update.effective_user.id
    
    # Проверка авторизации (только админы)
    if not db.is_admin(user_id):
        await update.message.reply_text("🔒 Доступ запрещён. Только для админов.")
        return
    
    try:
        await update.message.reply_text("⏳ Восстанавливаю имена СБ из журнала...")
        
        restored_count, messages = db.restore_sb_names_from_log()
        
        if restored_count > 0:
            response = [f"✅ Восстановлено записей: {restored_count}\n"]
            response.extend(messages[:20])  # Показываем первые 20
            if len(messages) > 20:
                response.append(f"\n... и ещё {len(messages) - 20} записей")
            await update.message.reply_text('\n'.join(response))
        else:
            await update.message.reply_text("ℹ️ Записей для восстановления не найдено")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка при восстановлении: {str(e)}")


async def handle_journal_command(update: Update, context: ContextTypes.DEFAULT_TYPE, 
                                 state: UserState, text: str):
    """Обработка команды журнал"""
    parts = text.split()
    
    limit = 20  # По умолчанию 20 записей
    code = None
    date = None
    
    # Парсим параметры
    # Формат: журнал [число] [код] [дата]
    # Примеры: журнал, журнал 50, журнал Д7, журнал 3,10, журнал Д7 3,10
    
    if len(parts) >= 2:
        # Проверяем второй параметр
        if parts[1].isdigit():
            limit = int(parts[1])
            if limit > 100:
                limit = 100  # Максимум 100
        else:
            # Это может быть код или дата
            from parser import DataParser
            from utils import parse_short_date
            
            # Пробуем как код
            if DataParser.is_code(parts[1]):
                code = DataParser.normalize_code(parts[1])
            else:
                # Пробуем как дату
                success, parsed_date, error = parse_short_date(parts[1])
                if success:
                    date = parsed_date
    
    if len(parts) >= 3:
        # Третий параметр
        from parser import DataParser
        from utils import parse_short_date
        
        if DataParser.is_code(parts[2]):
            code = DataParser.normalize_code(parts[2])
        else:
            success, parsed_date, error = parse_short_date(parts[2])
            if success:
                date = parsed_date
    
    # Получаем журнал
    logs = db.get_edit_log(limit=limit, code=code, date=date)
    
    if not logs:
        filter_info = []
        if code:
            filter_info.append(f"код: {code}")
        if date:
            filter_info.append(f"дата: {date}")
        
        filter_str = f" ({', '.join(filter_info)})" if filter_info else ""
        
        await update.message.reply_text(
            f"📜 Журнал изменений{filter_str}\n\n"
            f"Записей не найдено."
        )
        return
    
    # Форматируем журнал
    response_parts = []
    response_parts.append("📜 ЖУРНАЛ ИЗМЕНЕНИЙ\n")
    
    if code:
        response_parts.append(f"Фильтр: код {code}")
    if date:
        response_parts.append(f"Фильтр: дата {date}")
    
    response_parts.append(f"Показано: {len(logs)} из {limit}\n")
    
    for log in logs:
        # Форматируем дату и время
        edited_at = log['edited_at'][:16].replace('T', ' ')  # 2025-11-06T22:30:00 -> 2025-11-06 22:30
        
        action_type = log['action']
        
        # Определяем иконку по типу действия
        if 'merge' in action_type:
            icon = "🔄"
            action_text = log['action'].replace('merge_name: ', '')
        elif action_type == 'delete':
            icon = "🗑️"
            action_text = f"Удалено: {log['old_value']:.0f}"
        elif action_type == 'manual_update':
            icon = "✏️"
            action_text = f"Исправлено: {log['old_value']:.0f} → {log['new_value']:.0f}"
        elif action_type == 'update':
            icon = "➕"
            action_text = f"Добавлено: {log['old_value']:.0f} + ... = {log['new_value']:.0f}"
        elif action_type == 'replace':
            icon = "🔄"
            action_text = f"Заменено: {log['old_value']:.0f} → {log['new_value']:.0f}"
        else:
            icon = "📝"
            action_text = action_type
        
        response_parts.append(
            f"{icon} {edited_at}\n"
            f"   {log['club']} | {log['code']} | {log['channel'].upper()}\n"
            f"   {action_text}\n"
        )
    
    response_parts.append("─" * 35)
    response_parts.append(f"\n💡 Команды:")
    response_parts.append(f"• журнал 50 - показать 50 записей")
    response_parts.append(f"• журнал Д7 - по коду Д7")
    response_parts.append(f"• журнал 3,10 - за дату 03.10")
    
    await update.message.reply_text('\n'.join(response_parts))


def check_internal_duplicates(nal_data: list, beznal_data: list) -> list:
    """
    Проверка дубликатов внутри вводимых данных
    Возвращает список дубликатов (один код с именем и без имени)
    """
    from collections import defaultdict
    
    all_data = nal_data + beznal_data
    by_code = defaultdict(lambda: {'with_name': [], 'without_name': []})
    
    for item in all_data:
        code = item['code']
        if item['name']:
            by_code[code]['with_name'].append(item)
        else:
            by_code[code]['without_name'].append(item)
    
    # Ищем коды где есть И с именем И без имени
    duplicates = []
    for code, data in by_code.items():
        if data['with_name'] and data['without_name']:
            duplicates.append({
                'code': code,
                'with_name': data['with_name'],
                'without_name': data['without_name']
            })
    
    return duplicates


async def show_data_preview(update: Update, state: UserState, show_duplicates: bool = True):
    """Показать предпросмотр данных перед записью"""
    response_parts = []
    response_parts.append(f"📋 ПРЕДПРОСМОТР ДАННЫХ\n")
    response_parts.append(f"Клуб: {state.club}")
    
    if state.preview_date:
        response_parts.append(f"Дата: {state.preview_date}\n")
    
    # Показываем все данные с номерами строк
    line_num = 1
    total_nal = 0
    total_beznal = 0
    
    if state.temp_nal_data:
        response_parts.append("📗 НАЛ:")
        for item in state.temp_nal_data:
            response_parts.append(f"  {line_num}. {item['code']} {item['name']} — {item['amount']:.0f}")
            total_nal += item['amount']
            line_num += 1
        response_parts.append(f"  Итого НАЛ: {total_nal:.0f}\n")
    
    if state.temp_beznal_data:
        response_parts.append("📘 БЕЗНАЛ:")
        for item in state.temp_beznal_data:
            response_parts.append(f"  {line_num}. {item['code']} {item['name']} — {item['amount']:.0f}")
            total_beznal += item['amount']
            line_num += 1
        response_parts.append(f"  Итого БЕЗНАЛ: {total_beznal:.0f}\n")
    
    response_parts.append(f"💰 Всего: {total_nal + total_beznal:.0f}\n")
    
    # Проверка на дубликаты
    if show_duplicates:
        duplicates = check_internal_duplicates(state.temp_nal_data, state.temp_beznal_data)
        
        if duplicates:
            response_parts.append("⚠️ ВНИМАНИЕ! Найдены возможные дубликаты:\n")
            for i, dup in enumerate(duplicates, 1):
                response_parts.append(f"{i}. Код: {dup['code']}")
                
                # С именем
                names_with = set(item['name'] for item in dup['with_name'])
                for name in names_with:
                    items = [item for item in dup['with_name'] if item['name'] == name]
                    nal_sum = sum(item['amount'] for item in items if item in state.temp_nal_data)
                    bez_sum = sum(item['amount'] for item in items if item in state.temp_beznal_data)
                    response_parts.append(f"   • {name}: НАЛ {nal_sum:.0f}, БЕЗНАЛ {bez_sum:.0f}")
                
                # Без имени
                nal_no = sum(item['amount'] for item in dup['without_name'] if item in state.temp_nal_data)
                bez_no = sum(item['amount'] for item in dup['without_name'] if item in state.temp_beznal_data)
                response_parts.append(f"   • (без имени): НАЛ {nal_no:.0f}, БЕЗНАЛ {bez_no:.0f}")
                response_parts.append("")
            
            state.preview_duplicates = duplicates
    
    # Команды для пользователя
    response_parts.append("─" * 35)
    
    if not state.preview_date:
        response_parts.append("\n⏭️ СЛЕДУЮЩИЙ ШАГ:")
        response_parts.append("📅 Укажите дату в формате: 30,10 или 3,10")
        response_parts.append("\nПримеры:")
        response_parts.append("  • 3,10 → 03.10.2025")
        response_parts.append("  • 30,10 → 30.10.2025")
    else:
        response_parts.append("\n⏭️ ВЫБЕРИТЕ ДЕЙСТВИЕ:")
        response_parts.append("")
        response_parts.append("✅ ЗАПИСАТЬ")
        response_parts.append("   Сохранить данные в базу")
        response_parts.append("")
        response_parts.append("✏️ ИЗМЕНИТЬ")
        response_parts.append("   Редактировать строку по номеру")
        response_parts.append("")
        response_parts.append("❌ ОТМЕНА")
        response_parts.append("   Отменить весь ввод данных")
        
        if state.preview_duplicates:
            response_parts.append("")
            response_parts.append("─" * 35)
            response_parts.append("\n🔄 ОБЪЕДИНЕНИЕ ДУБЛИКАТОВ:")
            response_parts.append("")
            response_parts.append("• ОК → объединить все")
            response_parts.append("• ОК 1 → объединить только пункт 1")
            response_parts.append("• ОК 1 2 → объединить пункты 1 и 2")
            response_parts.append("• НЕ 1 → НЕ объединять пункт 1")
            response_parts.append("• НЕ 1 2 → НЕ объединять пункты 1 и 2")
    
    await update.message.reply_text('\n'.join(response_parts))


async def handle_preview_action(update: Update, state: UserState, text: str, text_lower: str):
    """Обработка действий в режиме предпросмотра"""
    
    # Проверяем команды объединения дубликатов
    if state.preview_duplicates and (text_lower.startswith('ок') or text_lower.startswith('не')):
        await handle_preview_duplicates(update, state, text_lower)
        return
    
    # ЗАПИСАТЬ - сохранение данных
    if text_lower == 'записать':
        await save_preview_data(update, state)
        return
    
    # ИЗМЕНИТЬ - редактирование строки
    if text_lower == 'изменить':
        total_lines = len(state.temp_nal_data) + len(state.temp_beznal_data)
        await update.message.reply_text(
            f"📝 Введите номер строки для редактирования:\n\n"
            f"📊 Доступно строк: 1-{total_lines}\n\n"
            f"Например: 1"
        )
        state.mode = 'awaiting_edit_line_number'
        return
    
    # Неизвестная команда
    await update.message.reply_text(
        "❓ Используйте команды:\n"
        "• ЗАПИСАТЬ\n"
        "• ИЗМЕНИТЬ\n"
        "• ОТМЕНА"
    )


async def handle_preview_duplicates(update: Update, state: UserState, text_lower: str):
    """Обработка объединения дубликатов в предпросмотре"""
    duplicates = state.preview_duplicates
    
    if not duplicates:
        await update.message.reply_text("❌ Дубликаты не найдены")
        return
    
    # Парсим команду
    normalized_text = text_lower.replace(',', ' ').replace('.', ' ')
    parts = normalized_text.split()
    
    if not parts:
        await update.message.reply_text("❌ Неверный формат")
        return
    
    command = parts[0]
    indices_to_merge = set()
    
    if command in ['ок', 'ok']:
        if len(parts) == 1:
            # Объединить все
            indices_to_merge = set(range(len(duplicates)))
        else:
            # Объединить указанные
            try:
                indices_to_merge = set(int(x) - 1 for x in parts[1:] if x.isdigit())
            except:
                await update.message.reply_text("❌ Неверный формат номеров")
                return
    elif command in ['не', 'нет']:
        # Не объединять указанные
        try:
            exclude_indices = set(int(x) - 1 for x in parts[1:] if x.isdigit())
            indices_to_merge = set(range(len(duplicates))) - exclude_indices
        except:
            await update.message.reply_text("❌ Неверный формат номеров")
            return
    
    # Объединяем дубликаты
    for i, dup in enumerate(duplicates):
        if i in indices_to_merge:
            code = dup['code']
            # Берём имя из записи с именем
            if dup['with_name']:
                merged_name = dup['with_name'][0]['name']
                
                # Обновляем записи без имени
                for item in dup['without_name']:
                    item['name'] = merged_name
    
    # Убираем дубликаты из списка
    state.preview_duplicates = None
    
    await update.message.reply_text(
        "✅ Дубликаты объединены!\n\n"
        "📋 Обновлённый предпросмотр:"
    )
    
    # Показываем обновлённый предпросмотр
    await show_data_preview(update, state, show_duplicates=True)


async def handle_edit_line_number(update: Update, state: UserState, text: str):
    """Обработка ввода номера строки для редактирования"""
    try:
        line_num = int(text.strip())
        
        # Проверяем диапазон
        total_lines = len(state.temp_nal_data) + len(state.temp_beznal_data)
        
        if line_num < 1 or line_num > total_lines:
            await update.message.reply_text(
                f"❌ Неверный номер строки. Введите число от 1 до {total_lines}"
            )
            return
        
        # Находим строку
        if line_num <= len(state.temp_nal_data):
            item = state.temp_nal_data[line_num - 1]
            channel = 'нал'
            index = line_num - 1
        else:
            item = state.temp_beznal_data[line_num - len(state.temp_nal_data) - 1]
            channel = 'безнал'
            index = line_num - len(state.temp_nal_data) - 1
        
        # Показываем текущие данные
        await update.message.reply_text(
            f"✏️ Редактирование строки {line_num}\n\n"
            f"📌 Текущие данные:\n"
            f"   Код: {item['code']}\n"
            f"   Имя: {item['name']}\n"
            f"   Сумма: {item['amount']:.0f}\n"
            f"   Канал: {channel.upper()}\n\n"
            f"📝 Введите новые данные в формате:\n"
            f"   КОД ИМЯ СУММА\n\n"
            f"💡 Пример: Д7 Юля 10000"
        )
        
        state.edit_line_number = line_num
        state.mode = 'awaiting_edit_line_data'
        
    except ValueError:
        await update.message.reply_text("❌ Введите корректный номер строки")


async def handle_edit_line_data(update: Update, state: UserState, text: str):
    """Обработка ввода новых данных для строки"""
    from parser import DataParser
    
    # Парсим новую строку
    success, data, error = DataParser.parse_line(text, 1)
    
    if not success:
        await update.message.reply_text(f"❌ {error}\n\nПопробуйте ещё раз")
        return
    
    # Определяем в каком списке находится строка
    line_num = state.edit_line_number
    
    if line_num <= len(state.temp_nal_data):
        # Обновляем в НАЛ
        state.temp_nal_data[line_num - 1] = data
    else:
        # Обновляем в БЕЗНАЛ
        index = line_num - len(state.temp_nal_data) - 1
        state.temp_beznal_data[index] = data
    
    await update.message.reply_text(
        "✅ Строка успешно обновлена!\n\n"
        "📋 Обновлённый предпросмотр:"
    )
    
    # Очищаем редактирование
    state.edit_line_number = None
    state.mode = 'awaiting_preview_action'
    
    # Показываем обновлённый предпросмотр
    await show_data_preview(update, state, show_duplicates=True)


async def save_preview_data(update: Update, state: UserState):
    """Сохранение данных из предпросмотра в БД"""
    if not state.preview_date:
        await update.message.reply_text("❌ Дата не указана")
        return
    
    saved_count = 0
    
    for item in state.temp_nal_data:
        db.add_or_update_operation(
            club=state.club,
            date=state.preview_date,
            code=item['code'],
            name=item['name'],
            channel='нал',
            amount=item['amount'],
            original_line=item['original_line'],
            aggregate=True
        )
        saved_count += 1
    
    for item in state.temp_beznal_data:
        db.add_or_update_operation(
            club=state.club,
            date=state.preview_date,
            code=item['code'],
            name=item['name'],
            channel='безнал',
            amount=item['amount'],
            original_line=item['original_line'],
            aggregate=True
        )
        saved_count += 1
    
    state.reset_input()
    
    await update.message.reply_text(
        f"✅ ДАННЫЕ УСПЕШНО СОХРАНЕНЫ!\n\n"
        f"🏢 Клуб: {state.club}\n"
        f"📅 Дата: {state.preview_date}\n"
        f"📊 Записей: {saved_count}\n\n"
        f"🎯 Что дальше?\n"
        f"• Введите новые данные: НАЛ / БЕЗНАЛ\n"
        f"• Посмотрите отчёт: ОТЧЁТ\n"
        f"• Или используйте другие команды ⬇️",
        reply_markup=get_main_keyboard()
    )


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка загруженных документов (Excel файлы)"""
    user_id = update.effective_user.id
    state = get_user_state(user_id)
    
    # Проверка авторизации (только админы)
    if not db.is_admin(user_id):
        await update.message.reply_text("🔒 Доступ запрещён. Только для админов.")
        return
    
    # Проверяем режим загрузки (обычный файл или ЗП)
    if state.mode == 'awaiting_payments_upload_file':
        # Обработка загрузки ЛИСТА ВЫПЛАТ
        document = update.message.document
        
        # Проверяем что это Excel файл
        if not (document.file_name.endswith('.xlsx') or document.file_name.endswith('.xls')):
            await update.message.reply_text(
                "❌ Поддерживаются только Excel файлы (.xlsx, .xls)\n"
                "Отправьте правильный файл или напишите: отмена"
            )
            return
        
        await update.message.reply_text("⏳ Обрабатываю лист выплат...")
        
        try:
            # Скачиваем файл
            file = await context.bot.get_file(document.file_id)
            file_bytes = await file.download_as_bytearray()
            
            # Парсим ЛИСТ ВЫПЛАТ
            excel_processor = ExcelProcessor()
            result = excel_processor.extract_payments_sheet(
                bytes(file_bytes), 
                db, 
                state.payments_upload_club,
                state.payments_upload_date
            )
            
            payments_data = result.get('payments', [])
            name_changes = result.get('name_changes', [])
            
            if not payments_data:
                await update.message.reply_text(
                    "❌ Не найден лист 'ЛИСТ ВЫПЛАТ' в файле\n"
                    "или он пустой.\n\n"
                    "Проверьте файл и попробуйте снова\n"
                    "или напишите: отмена"
                )
                return
            
            # Показываем предпросмотр
            preview_lines = [
                f"💰 ПРЕДПРОСМОТР ВЫПЛАТ\n",
                f"🏢 Клуб: {state.payments_upload_club}\n",
                f"📅 Дата: {state.payments_upload_date}\n",
                f"📊 Найдено записей: {len(payments_data)}\n"
            ]
            
            # Показываем предупреждение об изменении имён
            if name_changes:
                preview_lines.append(f"\n⚠️ ИЗМЕНЕНИЯ ИМЁН ({len(name_changes)}):\n")
                for change in name_changes[:5]:  # Показываем первые 5
                    preview_lines.append(
                        f"• {change['code']}: '{change['old_name']}' → '{change['new_name']}' (похожесть: {change['similarity']:.0%})\n"
                    )
                if len(name_changes) > 5:
                    preview_lines.append(f"... и ещё {len(name_changes) - 5} изменений\n")
                preview_lines.append("\n")
            
            preview_lines.append("\n")
            
            # Показываем первые 10 записей
            for i, pay in enumerate(payments_data[:10], 1):
                preview_lines.append(
                    f"{i}. {pay['code']} {pay['name']} - ИТОГО: {pay['total_shift']}\n"
                )
            
            if len(payments_data) > 10:
                preview_lines.append(f"\n... и ещё {len(payments_data) - 10} записей\n")
            
            keyboard = InlineKeyboardMarkup([
                [InlineKeyboardButton("✅ СОХРАНИТЬ", callback_data='payments_save_confirm')],
                [InlineKeyboardButton("❌ ОТМЕНА", callback_data='payments_save_cancel')]
            ])
            
            await update.message.reply_text(
                ''.join(preview_lines),
                reply_markup=keyboard
            )
            
            # Сохраняем данные в state для callback
            state.payments_preview_data = payments_data
            state.payments_name_changes = name_changes  # Сохраняем изменения имён
            
        except Exception as e:
            await update.message.reply_text(
                f"❌ Ошибка обработки файла: {str(e)}\n\n"
                f"Попробуйте снова или напишите: отмена"
            )
            state.mode = None
        
        return
    
    # Обычная загрузка файла (Примечания)
    if state.mode != 'awaiting_upload_file':
        return
    
    document = update.message.document
    
    # Проверяем что это Excel файл
    if not (document.file_name.endswith('.xlsx') or document.file_name.endswith('.xls')):
        await update.message.reply_text(
            "❌ Поддерживаются только Excel файлы (.xlsx, .xls)\n"
            "Отправьте правильный файл или напишите: отмена"
        )
        return
    
    await update.message.reply_text("⏳ Обрабатываю файл...")
    
    try:
        # Скачиваем файл
        file = await context.bot.get_file(document.file_id)
        file_bytes = await file.download_as_bytearray()
        
        # Парсим Excel
        excel_processor = ExcelProcessor()
        notes_data = excel_processor.extract_notes_entries(bytes(file_bytes))
        
        if not notes_data or (not notes_data.get('безнал') and not notes_data.get('нал')):
            await update.message.reply_text(
                "❌ Не найден блок 'Примечания' в файле\n"
                "или он пустой.\n\n"
                "Проверьте файл и попробуйте снова\n"
                "или напишите: отмена"
            )
            return
        
        # Обрабатываем данные через DataParser
        beznal_entries = notes_data.get('безнал', [])
        nal_entries = notes_data.get('нал', [])
        
        parsed_beznal = []
        parsed_nal = []
        errors = []
        
        # Парсим безнал
        for idx, entry in enumerate(beznal_entries, 1):
            if entry.get('is_total'):
                continue  # Пропускаем строку ИТОГО
            
            entry_text = entry.get('entry_text', '').strip()
            if not entry_text:
                continue
            
            success, data, error = DataParser.parse_line(entry_text, idx)
            if success:
                parsed_beznal.append(data)
            elif error and 'Пустая строка' not in error:
                errors.append(f"БЕЗНАЛ строка {idx}: {error}")
        
        # Парсим нал
        for idx, entry in enumerate(nal_entries, 1):
            if entry.get('is_total'):
                continue  # Пропускаем строку ИТОГО
            
            entry_text = entry.get('entry_text', '').strip()
            if not entry_text:
                continue
            
            success, data, error = DataParser.parse_line(entry_text, idx)
            if success:
                parsed_nal.append(data)
            elif error and 'Пустая строка' not in error:
                errors.append(f"НАЛ строка {idx}: {error}")
        
        if not parsed_beznal and not parsed_nal:
            await update.message.reply_text(
                "❌ Не удалось извлечь данные из файла\n\n"
                "Ошибки:\n" + "\n".join(errors[:5]) if errors else "Нет валидных строк"
            )
            state.mode = None
            return
        
        # Проверяем доплаты (строки начинающиеся с %)
        # ВАЖНО: Анализируем НАЛ и БЕЗНАЛ ОТДЕЛЬНО!
        beznal_analysis = DataParser.find_additional_payments(parsed_beznal)
        nal_analysis = DataParser.find_additional_payments(parsed_nal)
        
        # Сохраняем данные в состояние
        state.upload_file_data = {
            'beznal': parsed_beznal,
            'nal': parsed_nal,
            'errors': errors,
            'beznal_analysis': beznal_analysis,
            'nal_analysis': nal_analysis
        }
        
        # Показываем предпросмотр
        await show_file_preview(update, state)
        state.mode = 'awaiting_upload_confirm'
        
    except Exception as e:
        await update.message.reply_text(
            f"❌ Ошибка обработки файла: {str(e)}\n\n"
            f"Попробуйте снова или напишите: отмена"
        )
        state.mode = None


async def show_file_preview(update: Update, state: UserState):
    """Показать предпросмотр данных из файла"""
    data = state.upload_file_data
    beznal_list = data.get('beznal', [])
    nal_list = data.get('nal', [])
    errors = data.get('errors', [])
    beznal_analysis = data.get('beznal_analysis', {})
    nal_analysis = data.get('nal_analysis', {})
    
    # Заголовок
    header = []
    header.append("📎 ПРЕДПРОСМОТР ДАННЫХ ИЗ ФАЙЛА")
    header.append("")
    header.append(f"🏢 Клуб: {state.upload_file_club}")
    header.append(f"📅 Дата: {state.upload_file_date}")
    header.append("")
    
    # БЕЗНАЛ - формируем текст
    beznal_text = []
    if beznal_list:
        beznal_text.append(f"📘 БЕЗНАЛ ({len(beznal_list)} записей):")
        total_beznal = 0
        for idx, item in enumerate(beznal_list, 1):
            beznal_text.append(f"  {idx}. {item['code']} {item['name']} — {item['amount']:.0f}")
            total_beznal += item['amount']
        beznal_text.append(f"  💰 Итого безнал: {total_beznal:.0f}")
        beznal_text.append("")
    
    # НАЛ - формируем текст
    nal_text = []
    if nal_list:
        nal_text.append(f"📗 НАЛ ({len(nal_list)} записей):")
        total_nal = 0
        for idx, item in enumerate(nal_list, 1):
            nal_text.append(f"  {idx}. {item['code']} {item['name']} — {item['amount']:.0f}")
            total_nal += item['amount']
        nal_text.append(f"  💰 Итого нал: {total_nal:.0f}")
        nal_text.append("")
    
    # Ошибки
    errors_text = []
    if errors:
        errors_text.append(f"⚠️ Ошибок при парсинге: {len(errors)}")
        for error in errors[:5]:
            errors_text.append(f"  • {error}")
        if len(errors) > 5:
            errors_text.append(f"  ... и ещё {len(errors) - 5} ошибок")
        errors_text.append("")
    
    # Доплаты (строки с %) - ОТДЕЛЬНО ДЛЯ БЕЗНАЛ И НАЛ
    additional_text = []
    merge_counter = 0  # Сквозная нумерация для всех объединений
    
    # БЕЗНАЛ
    if beznal_analysis:
        beznal_merges = beznal_analysis.get('merges', [])
        beznal_not_found = beznal_analysis.get('not_found', [])
        beznal_no_code = beznal_analysis.get('no_code', [])
        
        if beznal_merges:
            additional_text.append("🔀 ДОПЛАТЫ БЕЗНАЛ:")
            additional_text.append("")
            for merge in beznal_merges:
                merge_counter += 1
                merge['merge_id'] = merge_counter  # Присваиваем ID
                code = merge['code']
                main_items = merge['main_items']
                add_item = merge['additional_item']
                total = merge['total_amount']
                
                additional_text.append(f"[{merge_counter}] Код: {code}")
                for main in main_items:
                    additional_text.append(f"     Основная: {main['name']} — {main['amount']:.0f}")
                additional_text.append(f"     Доплата: {add_item['original_line']} — {add_item['amount']:.0f}")
                additional_text.append(f"     ИТОГО: {total:.0f}")
                additional_text.append("")
        
        if beznal_not_found:
            additional_text.append("⚠️ БЕЗНАЛ - Доплаты без основной записи:")
            for item in beznal_not_found:
                additional_text.append(f"  • {item['original_line']} (код {item['code']} не найден)")
            additional_text.append("")
            
        if beznal_no_code:
            additional_text.append("❓ БЕЗНАЛ - Доплаты без кода:")
            for item in beznal_no_code:
                additional_text.append(f"  • {item['original_line']}")
            additional_text.append("")
    
    # НАЛ
    if nal_analysis:
        nal_merges = nal_analysis.get('merges', [])
        nal_not_found = nal_analysis.get('not_found', [])
        nal_no_code = nal_analysis.get('no_code', [])
        
        if nal_merges:
            additional_text.append("🔀 ДОПЛАТЫ НАЛ:")
            additional_text.append("")
            for merge in nal_merges:
                merge_counter += 1
                merge['merge_id'] = merge_counter  # Присваиваем ID
                code = merge['code']
                main_items = merge['main_items']
                add_item = merge['additional_item']
                total = merge['total_amount']
                
                additional_text.append(f"[{merge_counter}] Код: {code}")
                for main in main_items:
                    additional_text.append(f"     Основная: {main['name']} — {main['amount']:.0f}")
                additional_text.append(f"     Доплата: {add_item['original_line']} — {add_item['amount']:.0f}")
                additional_text.append(f"     ИТОГО: {total:.0f}")
                additional_text.append("")
        
        if nal_not_found:
            additional_text.append("⚠️ НАЛ - Доплаты без основной записи:")
            for item in nal_not_found:
                additional_text.append(f"  • {item['original_line']} (код {item['code']} не найден)")
            additional_text.append("")
            
        if nal_no_code:
            additional_text.append("❓ НАЛ - Доплаты без кода:")
            for item in nal_no_code:
                additional_text.append(f"  • {item['original_line']}")
            additional_text.append("")
    
    # Финал
    footer = []
    if additional_text:
        footer.append("⚠️ ВНИМАНИЕ! Обнаружены доплаты.")
        footer.append("Проверьте объединения выше.")
        footer.append("")
        footer.append("✅ Введите команду:")
        footer.append("  • ЗАПИСАТЬ - применить ВСЕ объединения")
        footer.append("  • ЗАПИСАТЬ 1 2 - применить только [1] и [2]")
        footer.append("  • ЗАПИСАТЬ БЕЗ 3 - применить все кроме [3]")
        footer.append("  • ОТМЕНА - отменить загрузку")
    else:
        footer.append("✅ Всё верно? Введите:")
        footer.append("  • ЗАПИСАТЬ - сохранить в базу")
        footer.append("  • ОТМЕНА - отменить")
    
    # Объединяем весь текст
    full_text = '\n'.join(header + beznal_text + nal_text + errors_text + additional_text + footer)
    
    # Разбиваем на части по 4000 символов если нужно
    max_length = 4000
    if len(full_text) <= max_length:
        await update.message.reply_text(full_text)
    else:
        # Разбиваем на куски
        parts = []
        current_part = []
        
        for line in (header + beznal_text + nal_text + errors_text + additional_text + footer):
            test_part = '\n'.join(current_part + [line])
            if len(test_part) > max_length and current_part:
                # Сохраняем текущую часть и начинаем новую
                parts.append('\n'.join(current_part))
                current_part = [line]
            else:
                current_part.append(line)
        
        # Добавляем последнюю часть
        if current_part:
            parts.append('\n'.join(current_part))
        
        # Отправляем все части
        for part in parts:
            await update.message.reply_text(part)


async def save_file_data_continue(message, state: UserState):
    """Продолжение сохранения файла после подтверждения замен"""
    # Вызываем save_file_data, но через Message объект
    # Создаём временный Update объект
    class FakeUpdate:
        def __init__(self, msg):
            self.message = msg
            self.effective_user = msg.from_user if hasattr(msg, 'from_user') else None
    
    fake_update = FakeUpdate(message)
    await save_file_data(fake_update, state)


async def show_merge_warning(update: Update, state: UserState, found_merges: List[Dict]):
    """Показать предупреждение о найденных объединённых сотрудниках"""
    lines = []
    lines.append("⚠️ ОБНАРУЖЕНЫ ОБЪЕДИНЁННЫЕ СОТРУДНИКИ")
    lines.append("")
    lines.append("В файле найдены сотрудники, которые ранее были объединены:")
    lines.append("")
    
    for merge in found_merges:
        lines.append(f"📌 {merge['channel'].upper()}")
        lines.append(f"   • {merge['original_code']} - {merge['original_name']}")
        lines.append(f"   → объединён с {merge['merged_code']} - {merge['merged_name']}")
        lines.append("")
    
    lines.append("💡 Что делать?")
    lines.append("")
    lines.append("✅ ДА - сохранить как объединённые")
    lines.append(f"   (данные будут записаны с новыми кодами)")
    lines.append("")
    lines.append("❌ НЕТ - сохранить как в файле")
    lines.append(f"   (игнорировать объединение)")
    
    keyboard = [
        [InlineKeyboardButton("✅ ДА, ЗАМЕНИТЬ", callback_data='upload_merge_yes')],
        [InlineKeyboardButton("❌ НЕТ, КАК В ФАЙЛЕ", callback_data='upload_merge_no')]
    ]
    
    await update.message.reply_text(
        '\n'.join(lines),
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def save_file_data(update: Update, state: UserState):
    """Сохранение данных из файла в БД с учетом объединений доплат"""
    data = state.upload_file_data
    beznal_list = data.get('beznal', [])
    nal_list = data.get('nal', [])
    beznal_analysis = data.get('beznal_analysis', {})
    nal_analysis = data.get('nal_analysis', {})
    selected_merges = data.get('selected_merges')  # None = все, [1,2] = только указанные
    
    # Сохраняем значения до очистки
    club = state.upload_file_club
    date = state.upload_file_date
    
    # НОВАЯ ЛОГИКА: Проверяем объединённых сотрудников
    if not data.get('merge_check_done'):
        found_merges = []
        
        # Проверяем безнал
        for item in beznal_list:
            if item.get('is_additional', False):
                continue
            merge_info = db.check_employee_merge(club, item['code'], item['name'])
            if merge_info:
                found_merges.append({
                    'channel': 'безнал',
                    'original_code': item['code'],
                    'original_name': item['name'],
                    'merged_code': merge_info['merged_code'],
                    'merged_name': merge_info['merged_name']
                })
        
        # Проверяем нал
        for item in nal_list:
            if item.get('is_additional', False):
                continue
            merge_info = db.check_employee_merge(club, item['code'], item['name'])
            if merge_info:
                found_merges.append({
                    'channel': 'нал',
                    'original_code': item['code'],
                    'original_name': item['name'],
                    'merged_code': merge_info['merged_code'],
                    'merged_name': merge_info['merged_name']
                })
        
        # Если нашли объединения - показываем предупреждение
        if found_merges:
            state.upload_file_data['found_merges'] = found_merges
            await show_merge_warning(update, state, found_merges)
            return  # Ждём подтверждения
        
        # Если не нашли - продолжаем сохранение
        data['merge_check_done'] = True
    
    # Создаем словари для объединений ОТДЕЛЬНО ДЛЯ БЕЗНАЛ И НАЛ
    beznal_merge_dict = {}
    nal_merge_dict = {}
    
    # БЕЗНАЛ - собираем объединения которые нужно применить
    # ВАЖНО: Для СБ используем ключ code_name, чтобы разные СБ не объединялись
    beznal_merges = beznal_analysis.get('merges', [])
    for merge in beznal_merges:
        merge_id = merge.get('merge_id')
        # Проверяем, нужно ли применять это объединение
        should_apply = False
        if selected_merges is None:
            should_apply = True  # Применяем все
        elif merge_id in selected_merges:
            should_apply = True  # Применяем только выбранные
        
        if should_apply:
            code = merge['code']
            name = merge['main_items'][0]['name'] if merge['main_items'] else ''
            
            # Для СБ используем комбинацию code_name как ключ
            if code == 'СБ' and name:
                merge_key = f"{code}_{name}"
            else:
                merge_key = code
            
            beznal_merge_dict[merge_key] = {
                'amount': merge['total_amount'],
                'name': name
            }
    
    # НАЛ - собираем объединения которые нужно применить
    # ВАЖНО: Для СБ используем ключ code_name, чтобы разные СБ не объединялись
    nal_merges = nal_analysis.get('merges', [])
    for merge in nal_merges:
        merge_id = merge.get('merge_id')
        # Проверяем, нужно ли применять это объединение
        should_apply = False
        if selected_merges is None:
            should_apply = True  # Применяем все
        elif merge_id in selected_merges:
            should_apply = True  # Применяем только выбранные
        
        if should_apply:
            code = merge['code']
            name = merge['main_items'][0]['name'] if merge['main_items'] else ''
            
            # Для СБ используем комбинацию code_name как ключ
            if code == 'СБ' and name:
                merge_key = f"{code}_{name}"
            else:
                merge_key = code
            
            nal_merge_dict[merge_key] = {
                'amount': merge['total_amount'],
                'name': name
            }
    
    # ПРИМЕНЯЕМ КАНОНИЧЕСКИЕ ИМЕНА (приоритет выше чем объединения)
    canonical_replacements = {}
    for item in beznal_list + nal_list:
        if item.get('is_additional', False):
            continue
        
        canonical = db.get_canonical_name(item['code'], club, date)
        if canonical:
            key = f"{item['code']}_{item['name']}"
            canonical_replacements[key] = {
                'code': item['code'],  # Код не меняется
                'name': canonical
            }
            print(f"DEBUG: Canonical name will be used for {item['code']}: {canonical}")
    
    # Создаём словарь замен для объединённых сотрудников
    employee_replacements = {}
    found_merges = data.get('found_merges', [])
    apply_merges = data.get('apply_employee_merges', False)  # True если пользователь нажал ДА
    
    if apply_merges and found_merges:
        for merge in found_merges:
            key = f"{merge['original_code']}_{merge['original_name']}"
            employee_replacements[key] = {
                'code': merge['merged_code'],
                'name': merge['merged_name']
            }
    
    saved_count = 0
    
    # Сохраняем безнал
    for item in beznal_list:
        # Пропускаем доплаты (is_additional=True) - они уже учтены
        if item.get('is_additional', False):
            continue
            
        code = item['code']
        name = item.get('name', '') or ''  # Убеждаемся что name не None
        
        # Проверяем каноническое имя (приоритет 1)
        key = f"{code}_{name}"
        if key in canonical_replacements:
            replacement = canonical_replacements[key]
            code = replacement['code']
            name = replacement['name']
        # Иначе проверяем объединения (приоритет 2)
        elif key in employee_replacements:
            replacement = employee_replacements[key]
            code = replacement['code']
            name = replacement['name']
        
        # Если код объединяется (доплаты) - используем итоговую сумму
        # Для СБ проверяем по комбинации code_name
        if code == 'СБ' and name:
            merge_key = f"{code}_{name}"
        else:
            merge_key = code
        
        if merge_key in beznal_merge_dict:
            # Используем сумму и имя из объединения
            amount = beznal_merge_dict[merge_key]['amount']
            name = beznal_merge_dict[merge_key]['name']  # ВАЖНО: обновляем имя из объединения
        else:
            amount = item['amount']
            # Для СБ без доплат сохраняем имя как есть
            
        db.add_or_update_operation(
            club=club,
            date=date,
            code=code,
            name=name,
            channel='безнал',
            amount=amount,
            original_line=item['original_line'],
            aggregate=True
        )
        saved_count += 1
    
    # Сохраняем нал
    for item in nal_list:
        # Пропускаем доплаты (is_additional=True) - они уже учтены
        if item.get('is_additional', False):
            continue
            
        code = item['code']
        name = item.get('name', '') or ''  # Убеждаемся что name не None
        
        # Проверяем каноническое имя (приоритет 1)
        key = f"{code}_{name}"
        if key in canonical_replacements:
            replacement = canonical_replacements[key]
            code = replacement['code']
            name = replacement['name']
        # Иначе проверяем объединения (приоритет 2)
        elif key in employee_replacements:
            replacement = employee_replacements[key]
            code = replacement['code']
            name = replacement['name']
        
        # Если код объединяется (доплаты) - используем итоговую сумму
        # Для СБ проверяем по комбинации code_name
        if code == 'СБ' and name:
            merge_key = f"{code}_{name}"
        else:
            merge_key = code
        
        if merge_key in nal_merge_dict:
            # Используем сумму и имя из объединения
            amount = nal_merge_dict[merge_key]['amount']
            name = nal_merge_dict[merge_key]['name']  # ВАЖНО: обновляем имя из объединения
        else:
            amount = item['amount']
            # Для СБ без доплат сохраняем имя как есть
            
        db.add_or_update_operation(
            club=club,
            date=date,
            code=code,
            name=name,
            channel='нал',
            amount=amount,
            original_line=item['original_line'],
            aggregate=True
        )
        saved_count += 1
    
    # Очищаем состояние
    state.upload_file_club = None
    state.upload_file_date = None
    state.upload_file_data = None
    state.mode = None
    
    await update.message.reply_text(
        f"✅ ДАННЫЕ ИЗ ФАЙЛА СОХРАНЕНЫ!\n\n"
        f"🏢 Клуб: {club}\n"
        f"📅 Дата: {date}\n"
        f"📊 Записей: {saved_count}\n\n"
        f"Используйте кнопки меню ⬇️",
        reply_markup=get_main_keyboard()
    )


def main():
    """Запуск бота"""
    # Проверяем токен
    if config.BOT_TOKEN == 'YOUR_BOT_TOKEN_HERE':
        print("❌ Ошибка: не установлен токен бота!")
        print("Установите переменную окружения TELEGRAM_BOT_TOKEN")
        print("или измените значение в config.py")
        return
    
    # Инициализация списка самозанятых (только если таблица пустая)
    initial_self_employed = [
        'Д4', 'Д5', 'Д11', 'Д15', 'Д18', 'Д20', 'Д23', 'Д33', 'Д35', 'Д38',
        'Д66', 'ОФ1', 'ОФ3', 'ОФ4', 'Б13', 'Б52', 'К2', 'К4', 'К21'
    ]
    added = db.init_self_employed_list(initial_self_employed)
    if added > 0:
        print(f"[OK] Инициализирован список самозанятых: {added} кодов")
    else:
        print(f"[OK] Список самозанятых уже существует, инициализация пропущена")
    
    # Создаем приложение
    app = Application.builder().token(config.BOT_TOKEN).build()
    
    # Регистрируем обработчики
    app.add_handler(CommandHandler("start", start_command))
    app.add_handler(CommandHandler("restore_sb", restore_sb_names_command))
    app.add_handler(CallbackQueryHandler(handle_callback_query))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    
    # Запускаем бота
    print("[BOT] Бот запущен!")
    print("Для остановки нажмите Ctrl+C")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


async def handle_stylist_view_delete(update: Update, state: UserState, text: str):
    """Обработка удаления записей стилистов при просмотре"""
    if text.lower() == 'отмена':
        state.mode = None
        await update.message.reply_text("❌ Отменено")
        return
    
    # Парсим номера
    numbers = []
    try:
        # Поддержка: "3", "1 5 8", "1-5"
        parts = text.replace(',', ' ').split()
        for part in parts:
            if '-' in part:
                # Диапазон: 1-5
                start, end = map(int, part.split('-'))
                numbers.extend(range(start, end + 1))
            else:
                numbers.append(int(part))
    except:
        await update.message.reply_text(
            "❌ Неверный формат.\n\n"
            "Примеры:\n• 3\n• 1 5 8\n• 1-5"
        )
        return
    
    # Получаем записи
    expenses = db.get_stylist_expenses_by_period(
        state.stylist_view_club,
        state.stylist_view_from,
        state.stylist_view_to
    )
    
    # Удаляем по id (записи нумеруются с 1)
    deleted = 0
    for num in numbers:
        if 1 <= num <= len(expenses):
            exp = expenses[num - 1]
            # Удаляем из БД по уникальным полям
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                DELETE FROM stylist_expenses
                WHERE club = ? AND period_from = ? AND period_to = ?
                  AND code = ? AND name = ? AND amount = ?
                LIMIT 1
            """, (
                state.stylist_view_club,
                state.stylist_view_from,
                state.stylist_view_to,
                exp['code'],
                exp['name'],
                exp['amount']
            ))
            conn.commit()
            conn.close()
            deleted += 1
    
    state.mode = None
    await update.message.reply_text(
        f"✅ Удалено записей: {deleted}\n\n"
        f"Для просмотра обновленного списка:\n"
        f"СТИЛИСТЫ → Показать расходы"
    )


async def handle_stylist_view_edit_number(update: Update, state: UserState, text: str):
    """Обработка выбора номера записи для редактирования"""
    if text.lower() == 'отмена':
        state.mode = None
        await update.message.reply_text("❌ Отменено")
        return
    
    try:
        number = int(text.strip())
    except:
        await update.message.reply_text("❌ Введите номер записи (число)")
        return
    
    # Получаем записи
    expenses = db.get_stylist_expenses_by_period(
        state.stylist_view_club,
        state.stylist_view_from,
        state.stylist_view_to
    )
    
    if number < 1 or number > len(expenses):
        await update.message.reply_text(f"❌ Номер вне диапазона (1-{len(expenses)})")
        return
    
    exp = expenses[number - 1]
    state.stylist_view_edit_index = number - 1
    state.mode = 'awaiting_stylist_view_edit_data'
    
    await update.message.reply_text(
        f"✏️ РЕДАКТИРОВАНИЕ ЗАПИСИ #{number}\n\n"
        f"Текущие данные:\n"
        f"Код: {exp['code']}\n"
        f"Имя: {exp['name']}\n"
        f"Сумма: {exp['amount']}₽\n\n"
        f"Введите новые данные в формате:\n"
        f"КОД ИМЯ СУММА\n\n"
        f"Пример: Д13 Марго 3500\n\n"
        f"Или ОТМЕНА"
    )


async def handle_stylist_view_edit_data(update: Update, state: UserState, text: str):
    """Обработка ввода новых данных при редактировании"""
    if text.lower() == 'отмена':
        state.mode = None
        await update.message.reply_text("❌ Отменено")
        return
    
    # Парсим данные
    from parser import DataParser
    expenses, errors = DataParser.parse_stylist_expenses(text)
    
    if errors or len(expenses) != 1:
        await update.message.reply_text(
            "❌ Неверный формат.\n\n"
            "Введите данные в формате:\n"
            "КОД ИМЯ СУММА\n\n"
            "Пример: Д13 Марго 3500"
        )
        return
    
    new_exp = expenses[0]
    
    # Получаем старую запись
    old_expenses = db.get_stylist_expenses_by_period(
        state.stylist_view_club,
        state.stylist_view_from,
        state.stylist_view_to
    )
    
    if state.stylist_view_edit_index >= len(old_expenses):
        await update.message.reply_text("❌ Ошибка: запись не найдена")
        state.mode = None
        return
    
    old_exp = old_expenses[state.stylist_view_edit_index]
    
    # Удаляем старую и добавляем новую
    conn = db.get_connection()
    cursor = conn.cursor()
    
    # Удаляем старую
    cursor.execute("""
        DELETE FROM stylist_expenses
        WHERE club = ? AND period_from = ? AND period_to = ?
          AND code = ? AND name = ? AND amount = ?
        LIMIT 1
    """, (
        state.stylist_view_club,
        state.stylist_view_from,
        state.stylist_view_to,
        old_exp['code'],
        old_exp['name'],
        old_exp['amount']
    ))
    
    # Добавляем новую
    from datetime import datetime
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute("""
        INSERT INTO stylist_expenses 
        (club, period_from, period_to, code, name, amount, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    """, (
        state.stylist_view_club,
        state.stylist_view_from,
        state.stylist_view_to,
        new_exp['code'],
        new_exp['name'],
        new_exp['amount'],
        now
    ))
    
    conn.commit()
    conn.close()
    
    state.mode = None
    await update.message.reply_text(
        f"✅ ЗАПИСЬ ОБНОВЛЕНА\n\n"
        f"Было: {old_exp['code']} {old_exp['name']} {old_exp['amount']}₽\n"
        f"Стало: {new_exp['code']} {new_exp['name']} {new_exp['amount']}₽\n\n"
        f"Для просмотра обновленного списка:\n"
        f"СТИЛИСТЫ → Показать расходы"
    )


if __name__ == '__main__':
    main()

